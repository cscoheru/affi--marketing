import asyncio
import contextvars
import io
import json
import logging
import os
import re
import time
from datetime import datetime, date, timezone
from pathlib import Path

from dotenv import load_dotenv
from fastapi import Depends as _Depends, FastAPI, HTTPException, Request, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.openapi.docs import get_swagger_ui_html
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from swagger_ui_bundle import swagger_ui_path
from jose import JWTError, jwt as _jwt
from psycopg.types.json import Json

from app.core.auth import (
    get_current_admin, verify_password, create_access_token,
    hash_password, get_current_user, ALGORITHM, JWT_SECRET,
)
from app.core import admin_crud
from app.api import assessment

load_dotenv()

# ── 多租户 Context（从 connection.py 导入共享实例）──

load_dotenv()

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

_swagger_static_dir = str(swagger_ui_path).lstrip("/")

app = FastAPI(
    title="国窖1573智能销售兵工厂",
    description="AI驱动的高端白酒销售培训系统",
    version="2.1.0",
    docs_url=None,  # 禁用默认，用自定义路由
    redoc_url=None,
)
app.mount("/static/swagger-ui", StaticFiles(directory=swagger_ui_path), name="swagger-static")

@app.get("/docs", include_in_schema=False)
async def custom_swagger_ui_html(request: Request) -> HTMLResponse:
    return get_swagger_ui_html(
        openapi_url=request.url_for("openapi"),
        title=app.title + " - Swagger UI",
        swagger_js_url="/static/swagger-ui/swagger-ui-bundle.js",
        swagger_css_url="/static/swagger-ui/swagger-ui.css",
        swagger_favicon_url="/static/swagger-ui/favicon-32x32.png",
    )

@app.get("/redoc", include_in_schema=False)
async def custom_redoc_html(request: Request) -> HTMLResponse:
    from fastapi.openapi.docs import get_redoc_html
    return get_redoc_html(
        openapi_url=request.url_for("openapi"),
        title=app.title + " - ReDoc",
        redoc_favicon_url=f"{_swagger_static}/favicon-32x32.png",
    )

# CORS — allow_origins=["*"] + credentials=True 不兼容，浏览器会拒绝
# 改用 allow_origin_regex 匹配所有来源（效果等同但兼容 credentials）
app.add_middleware(
    CORSMiddleware,
    allow_origin_regex=r"https?://.*",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── 多租户 + 角色中间件（行级隔离） ──
@app.middleware("http")
async def tenant_schema_middleware(request: Request, call_next):
    """从请求头提取租户和角色，设置 contextvar"""
    slug = request.headers.get("X-Tenant-Slug", "guojiao")
    _tenant_schema.set(slug)
    role_id = request.headers.get("X-Role-Id")
    _tenant_role.set(int(role_id) if role_id else None)
    logger.info(f"[tenant] {request.url.path} → {slug}")
    response = await call_next(request)
    return response

# ── 数据源 ──
DATA_PATH = Path(os.getenv("DATA_PATH", Path(__file__).parent.parent.parent / "frontend" / "src" / "data" / "enriched-data.json"))

from app.core.database.connection import db, admin_db, admin_table, _tenant_schema, _tenant_role
from app.core.admin_crud import _apply_role_filter





def _load_json_fallback() -> dict:
    """JSON 文件降级加载"""
    with open(DATA_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def _try_db(func):
    """数据库优先 + JSON 降级的装饰器工厂"""
    def wrapper(*args, **kwargs):
        try:
            return func(db(), *args, **kwargs)
        except Exception as e:
            logger.warning(f"数据库请求失败，降级到 JSON: {e}")
            return func(None, *args, **kwargs)
    return wrapper


# ═══════════ Snake_case → camelCase 转换 ═══════════

def _to_camel(snake: str) -> str:
    """snake_case → camelCase（处理 trigger_keywords → triggerKeywords）"""
    parts = snake.split("_")
    return parts[0] + "".join(p.capitalize() for p in parts[1:])


def _camel_to_snake(name: str) -> str:
    """camelCase → snake_case"""
    import re as _re
    s1 = _re.sub("(.)([A-Z][a-z]+)", r"\1_\2", name)
    return _re.sub("([a-z0-9])([A-Z])", r"\1_\2", s1).lower()


def _body_to_snake(body: dict) -> dict:
    """将请求体的 camelCase 键转为 snake_case"""
    return {_camel_to_snake(k): v for k, v in body.items()}


def _row_to_camel(row: dict) -> dict:
    """将数据库行的 snake_case 键转为 camelCase"""
    return {_to_camel(k): v for k, v in row.items()}


# ═══════════ 健康检查 ═══════════
def _get_tenant_display() -> str:
    """获取当前租户显示名称"""
    try:
        sb = admin_table()
        result = sb.table("tenant_configs").select("display_name").eq("tenant_id", _tenant_schema.get("guojiao")).execute().data
        return result[0]["display_name"] if result else "国窖1573智能销售兵工厂"
    except Exception:
        return "国窖1573智能销售兵工厂"


@app.get("/health")
async def health_check():
    return {
        "status": "ok",
        "service": _get_tenant_display(),
        "version": "2.1.0",
    }


# ═══════════ 用户认证 API（手机端）════════════
@app.post("/api/auth/register")
async def user_register(body: dict):
    """用户注册（手机号 + 密码 + 角色）"""
    phone = (body.get("phone") or "").strip()
    password = body.get("password") or ""
    role_id = body.get("role_id") or body.get("roleId")
    if not phone or len(phone) < 11:
        raise HTTPException(status_code=400, detail="请输入有效手机号")
    if len(password) < 6:
        raise HTTPException(status_code=400, detail="密码至少6位")
    if not role_id:
        raise HTTPException(status_code=400, detail="请选择角色")
    sb = db()
    # 验证角色存在
    roles = sb.table("roles").select("id").eq("id", role_id).eq("is_active", True).execute().data
    if not roles:
        raise HTTPException(status_code=400, detail="无效的角色")
    existing = sb.table("users").select("id").eq("phone", phone).execute().data
    if existing:
        raise HTTPException(status_code=409, detail="该手机号已注册")
    password_hash = hash_password(password)
    nickname = body.get("nickname") or f"用户{phone[-4:]}"
    result = sb.table("users").insert({
        "phone": phone, "password_hash": password_hash, "nickname": nickname,
        "role_id": int(role_id),
    }).execute()
    if not result.data:
        raise HTTPException(status_code=500, detail="注册失败")
    user = result.data[0]
    token = create_access_token({"sub": str(user["id"]), "type": "user"})
    return {"user": _row_to_camel(user), "access_token": token, "token_type": "bearer"}


@app.post("/api/auth/login")
async def user_login(body: dict):
    """用户登录（手机号 + 密码）"""
    phone = (body.get("phone") or "").strip()
    password = body.get("password") or ""
    if not phone or not password:
        raise HTTPException(status_code=400, detail="手机号和密码不能为空")
    sb = db()
    users = sb.table("users").select("*").eq("phone", phone).execute().data
    if not users:
        raise HTTPException(status_code=401, detail="手机号未注册")
    user = users[0]
    if not verify_password(password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="密码错误")
    sb.table("users").update({"last_login": "now()"}).eq("id", user["id"]).execute()
    token = create_access_token({"sub": str(user["id"]), "type": "user"})
    return {"user": _row_to_camel(user), "access_token": token, "token_type": "bearer"}


@app.get("/api/auth/me")
async def user_me(request: Request):
    """获取当前用户信息（可选认证）"""
    auth_header = request.headers.get("Authorization", "")
    if not auth_header or not auth_header.startswith("Bearer "):
        return {"user": None, "logged_in": False}
    token = auth_header[7:]
    try:
        payload = _jwt.decode(token, JWT_SECRET, algorithms=[ALGORITHM])
        if payload.get("type") != "user":
            return {"user": None, "logged_in": False}
        user_id = payload.get("sub")
        if not user_id:
            return {"user": None, "logged_in": False}
    except JWTError:
        return {"user": None, "logged_in": False}
    sb = db()
    try:
        users = sb.table("users").select("*").eq("id", user_id).execute().data
        if users:
            u = users[0]
            return {"user": {"id": u["id"], "phone": u.get("phone", ""), "nickname": u.get("nickname", ""), "current_level": u.get("current_level", 1), "total_xp": u.get("total_xp", 0)}, "logged_in": True}
    except Exception:
        pass
    return {"user": None, "logged_in": False}


# ═══════════ 公共：角色列表（注册用） ═══════════
@app.get("/api/roles")
async def public_roles():
    """获取当前租户的角色列表（注册时选择）"""
    sb = db()
    rows = sb.table("roles").select("id, name, slug, description, icon, color, sort_order").eq("is_active", True).order("sort_order").execute().data
    return {"roles": [_row_to_camel(r) for r in rows]}


# ═══════════ 搜索反馈（匿名） ═══════════
@app.post("/api/feedback")
async def submit_feedback(body: dict):
    """记录搜索结果反馈（有用/无用），匿名 device_id"""
    sb = db()
    try:
        sb.table("search_feedbacks").insert({
            "query": body.get("query", "")[:500],
            "matched_tags": body.get("tags", []),
            "useful": body.get("useful", True),
            "device_id": body.get("device_id", ""),
        }).execute()
        return {"status": "ok"}
    except Exception as e:
        print(f"[feedback] insert error: {e}")
        return {"status": "ok"}  # 静默失败，不打断用户体验


# ═══════════ 学习记录（匿名） ═══════════
@app.post("/api/user/learning")
async def record_learning(body: dict):
    """记录用户学习行为（浏览/完成/搜索），匿名 device_id"""
    sb = db()
    try:
        sb.table("learning_records").insert({
            "device_id": body.get("device_id", "")[:100],
            "action_type": body.get("action_type", "view"),
            "item_id": body.get("item_id", ""),
            "item_type": body.get("item_type", "resource"),
            "metadata": body.get("metadata", {}),
        }).execute()
        return {"status": "ok"}
    except Exception as e:
        print(f"[learning] insert error: {e}")
        return {"status": "ok"}


# ═════════════ 后台情报采集调度器 ═════════════
_intel_scheduler_running = False
_intel_scheduler_task = None


async def _intelligence_scheduler():
    """后台循环：每 10 分钟检查是否有源需要采集"""
    global _intel_scheduler_running
    _intel_scheduler_running = True
    logger.info("情报采集调度器已启动")

    while _intel_scheduler_running:
        try:
            # 检查情报采集功能是否启用
            from app.core.feature_flags import is_feature_enabled
            if not is_feature_enabled("content.intelligence"):
                await asyncio.sleep(600)
                continue
            sb = db()
            # 获取所有活跃源
            sources = sb.table("intelligence_sources").select(
                "id, fetch_interval, last_fetched_at"
            ).eq("active", True).execute().data or []

            now = time.time()
            for source in sources:
                interval = source.get("fetch_interval", 360) * 60  # 分钟 → 秒
                last_fetched = source.get("last_fetched_at")
                if last_fetched:
                    last_ts = datetime.fromisoformat(last_fetched.replace("Z", "+00:00")).timestamp()
                else:
                    last_ts = 0

                if now - last_ts >= interval:
                    logger.info(f"触发采集: source={source['id']}")
                    try:
                        await asyncio.to_thread(run_collection, source["id"])
                    except Exception as e:
                        logger.error(f"采集异常 ({source['id']}): {e}")

        except Exception as e:
            logger.error(f"调度器循环异常: {e}")

        # 每 10 分钟检查一次
        await asyncio.sleep(600)


@app.on_event("startup")
async def start_intelligence_scheduler():
    """FastAPI 启动时启动情报采集调度器"""
    import threading
    loop = asyncio.get_event_loop()

    async def _run():
        await asyncio.sleep(30)  # 启动后 30 秒再开始采集
        await _intelligence_scheduler()

    global _intel_scheduler_task
    _intel_scheduler_task = asyncio.create_task(_run())


# ═══════════ 标签接口 ═══════════
@app.get("/api/tags")
async def get_tags(dimension: str | None = None):
    """获取标签字典"""
    try:
        sb = db()
        query = sb.table("tags").select("*")
        # 角色过滤：如果有角色上下文，只返回该角色的标签 + 共享标签
        role_id = _tenant_role.get()
        if role_id:
            query = query.or_(f"role_id.eq.{role_id},role_id.is.null")
        if dimension:
            query = query.eq("dimension", dimension)
        rows = query.execute().data
        tags = [_row_to_camel(r) for r in rows]
        return {"tags": tags, "total": len(tags)}
    except Exception as e:
        logger.warning(f"数据库请求失败，降级到 JSON: {e}")
        data = _load_json_fallback()
        tags = data.get("tagDict", [])
        if dimension:
            tags = [t for t in tags if t.get("dimension") == dimension]
        return {"tags": tags, "total": len(tags)}


@app.get("/api/tags/associations")
async def get_tag_associations():
    """获取传导规则（B/V → P/K）和痛点-技能关联"""
    try:
        sb = db()
        # 传导规则
        传导_rows = sb.table("tag_associations").select("source_tag_id, target_tag_id, weight, relation_type").execute().data
        # 痛点→技能关联
        ps_rows = sb.table("painpoint_skill_associations").select("painpoint_id, skill_id, weight, source_scenario_id").execute().data
        return {
            "conduction_rules": [_row_to_camel(r) for r in 传导_rows],
            "painpoint_skill_rules": [_row_to_camel(r) for r in ps_rows],
        }
    except Exception as e:
        return {"conduction_rules": [], "painpoint_skill_rules": [], "error": str(e)}


# ═══════════ 资源接口 ═══════════
@app.get("/api/resources")
async def get_resources(
    tags: str | None = None,
    type: str | None = None,
    search: str | None = None,
    page: int = 1,
    limit: int = 50,
):
    """获取资源库"""
    try:
        sb = db()

        # 先查 resource_ids（按标签过滤）
        resource_ids = None
        if tags:
            tag_list = [t.strip() for t in tags.split(",")]
            # 每个标签都要匹配（AND 逻辑）
            for tag_id in tag_list:
                res = sb.table("resource_tags").select("resource_id").eq("tag_id", tag_id).execute().data
                ids = {r["resource_id"] for r in res}
                if resource_ids is None:
                    resource_ids = ids
                else:
                    resource_ids &= ids
            if not resource_ids:
                return {"resources": [], "total": 0, "page": page, "limit": limit}

        # 查 resources
        query = sb.table("resources").select("*")
        if resource_ids is not None:
            query = query.in_("id", list(resource_ids))
        if type:
            query = query.eq("type", type)
        if search:
            query = query.ilike("title", f"%{search}%").or_("content", f"ilike.%{search}%")

        total_res = query.execute()
        all_rows = total_res.data
        total = len(all_rows)
        start = (page - 1) * limit
        page_rows = all_rows[start:start + limit]

        # 补充 tags 关联
        resources = []
        for r in page_rows:
            res = sb.table("resource_tags").select("tag_id").eq("resource_id", r["id"]).execute().data
            r["tags"] = [rt["tag_id"] for rt in res]
            resources.append(_row_to_camel(r))

        return {"resources": resources, "total": total, "page": page, "limit": limit}
    except Exception as e:
        logger.warning(f"数据库请求失败，降级到 JSON: {e}")
        data = _load_json_fallback()
        resources = data.get("resourceLibrary", [])
        if tags:
            tag_list = [t.strip() for t in tags.split(",")]
            resources = [r for r in resources if all(t in r.get("tags", []) for t in tag_list)]
        if type:
            resources = [r for r in resources if r.get("type") == type]
        if search:
            s = search.lower()
            resources = [r for r in resources if s in r.get("title", "").lower() or s in r.get("content", "").lower()]
        total = len(resources)
        start = (page - 1) * limit
        resources = resources[start:start + limit]
        return {"resources": resources, "total": total, "page": page, "limit": limit}


# ═══════════ AI 分析接口 ═══════════
@app.post("/api/analyze")
async def analyze_query(request: dict):
    """AI 智能匹配 — DashScope Qwen + 关键词降级"""
    t0 = time.time()
    query = request.get("query", "")

    from app.core.llm.prompt_builder import build_system_prompt, apply_auto_association, apply_conduction
    from app.core.llm.dashscope_client import analyze_with_fallback

    # 获取 tag_dict（数据库优先）
    tag_dict = _get_tag_dict()

    # 获取租户名称用于 prompt
    tenant_name = "国窖1573"
    try:
        sb = db()
        # 如果当前租户是数字化中心，调整 prompt
    except Exception:
        pass

    system_prompt = build_system_prompt(tag_dict, tenant_name)
    llm_result = analyze_with_fallback(system_prompt, query)

    conduction_map = {}
    if llm_result:
        tags = apply_auto_association(llm_result.tags)
        # 应用传导规则 B/V → P/K
        tags, conduction_map = apply_conduction(tags)
        latency_ms = llm_result.latency_ms
        reasoning = llm_result.reasoning
        model_used = llm_result.model_used
        confidence = llm_result.confidence
        logger.info(f"LLM 匹配成功: model={model_used}, tags={tags}, conduction={conduction_map}, latency={latency_ms}ms")
    else:
        tags, confidence, reasoning = _keyword_fallback(query)
        model_used = "keyword-fallback"
        logger.warning(f"降级到关键词匹配: tags={tags}")

    # 获取匹配资源
    matched_resources = _get_resources_by_tags(tags, conduction_map)
    total_latency = int((time.time() - t0) * 1000)

    result = {
        "tags": tags,
        "confidence": confidence,
        "reasoning": reasoning,
        "resources": matched_resources,
        "model_used": model_used,
        "latency_ms": total_latency,
    }
    if conduction_map:
        result["conduction_map"] = conduction_map
    return result


# ═══════════ AI 分析接口 v2（向量搜索 + LLM 精排）═══════════
@app.post("/api/analyze/v2")
async def analyze_query_v2(request: dict):
    """AI 智能匹配 v2 — pgvector 语义搜索 + LLM 精排（flag 关闭时降级到 v1）"""
    from app.core.feature_flags import is_feature_enabled
    if not is_feature_enabled("ai.search_v2"):
        return await analyze_query(request)

    t0 = time.time()
    query = request.get("query", "")

    from app.core.llm.prompt_builder import apply_auto_association
    from app.core.llm.dashscope_client import analyze_with_fallback
    from app.core.llm.rerank_prompt import build_rerank_prompt
    from app.core.vector.search import search_tags

    # Step 1: Vector search for top-10 candidates (run in thread to avoid blocking)
    candidates = await asyncio.to_thread(search_tags, query, 10)

    if not candidates:
        logger.warning("Vector search returned no candidates, falling back to v1")
        return await analyze_query(request)

    # Step 2: High-similarity fast path — skip LLM when top result is very confident
    top_similarity = max(c.similarity for c in candidates)
    if top_similarity >= 0.85:
        tags = _select_top_by_dimension(candidates)
        confidence = round(top_similarity, 2)
        reasoning = f"向量语义匹配（相似度 {confidence:.2f}，跳过 LLM 精排）"
        model_used = "v2-vector-only"
        logger.info(f"V2 快速路径: similarity={top_similarity:.2f}, tags={tags}")
    else:
        # Step 3: LLM rerank from candidates (run in thread to avoid blocking)
        candidate_dicts = [
            {
                "id": c.id, "label": c.label, "dimension": c.dimension,
                "description": c.description,
                "trigger_keywords": c.trigger_keywords,
                "auto_association": c.auto_association,
            }
            for c in candidates
        ]
        system_prompt = build_rerank_prompt(candidate_dicts)
        llm_result = await asyncio.to_thread(analyze_with_fallback, system_prompt, query)

        if llm_result:
            tags = apply_auto_association(llm_result.tags)
            reasoning = llm_result.reasoning
            model_used = "v2-vector+llm"
            confidence = llm_result.confidence
            logger.info(f"V2 匹配成功: tags={tags}, candidates={len(candidates)}, latency={llm_result.latency_ms}ms")
        else:
            tags = _select_top_by_dimension(candidates)
            confidence = 0.6
            reasoning = "向量语义匹配结果（LLM 精排暂不可用）"
            model_used = "v2-vector-only"
            logger.warning(f"LLM 精排失败，使用向量直选: tags={tags}")

    matched_resources = await asyncio.to_thread(_get_resources_by_tags, tags)

    # 获取匹配技能关联的场景（技能池跨场景共享）
    skill_tags = [t for t in tags if t.startswith("K")]
    matched_scenarios = []
    if skill_tags:
        try:
            sb = db()
            for kid in skill_tags:
                sc = sb.rpc("get_skill_scenarios", {"skill_id_param": kid}).execute()
                if sc.data:
                    matched_scenarios.extend(sc.data)
            # 去重
            seen_ids = set()
            unique = []
            for s in matched_scenarios:
                if s["scenario_id"] not in seen_ids:
                    seen_ids.add(s["scenario_id"])
                    unique.append(s)
            matched_scenarios = unique[:5]  # 最多返回 5 个
        except Exception as e:
            logger.warning(f"Failed to get matched scenarios: {e}")

    # Phase 3: 获取父级技能作为扩展学习推荐
    extended_tags = []
    if skill_tags:
        try:
            sb = db()
            for kid in skill_tags:
                tag_row = sb.table("tags").select("id", "label", "parent_id", "granularity").eq("id", kid).execute().data
                if tag_row and tag_row[0].get("parent_id"):
                    parent_id = tag_row[0]["parent_id"]
                    parent_row = sb.table("tags").select("id", "label", "granularity").eq("id", parent_id).eq("is_published", True).execute().data
                    if parent_row:
                        extended_tags.append({
                            "id": parent_row[0]["id"],
                            "label": parent_row[0]["label"],
                            "granularity": parent_row[0].get("granularity", "skill"),
                            "reason": f"「{tag_row[0]['label']}」的上级能力",
                        })
        except Exception:
            pass

    total_latency = int((time.time() - t0) * 1000)

    result = {
        "tags": tags, "confidence": confidence, "reasoning": reasoning,
        "resources": matched_resources, "model_used": model_used,
        "latency_ms": total_latency,
    }
    if matched_scenarios:
        result["matched_scenarios"] = matched_scenarios
    if extended_tags:
        result["extended_tags"] = extended_tags
    return result


def _select_top_by_dimension(candidates) -> list[str]:
    """Select top candidate per dimension from vector search results."""
    seen = set()
    tags = []
    for c in sorted(candidates, key=lambda x: x.similarity, reverse=True):
        if c.dimension not in seen:
            seen.add(c.dimension)
            tags.append(c.id)
    return tags


def _get_tag_dict() -> list[dict]:
    """获取标签字典（数据库优先，JSON降级）— 五维标签"""
    try:
        sb = db()
        rows = sb.table("tags").select("*").eq("is_published", True).execute().data
        tag_dict = []
        for r in rows:
            tag_dict.append({
                "id": r["id"],
                "label": r["label"],
                "dimension": r["dimension"],
                "description": r.get("description", ""),
                "triggerKeywords": r.get("trigger_keywords", []),
                "autoAssociation": r.get("auto_association", []),
                "category": r.get("category", ""),
                "alias": r.get("alias", []),
            })
        return tag_dict
    except Exception as e:
        logger.warning(f"获取 tag_dict 失败，降级到 JSON: {e}")
        return _load_json_fallback().get("tagDict", [])


def _get_resources_by_tags(tags: list[str], conduction_map: dict | None = None) -> list[dict]:
    """根据标签获取匹配资源

    匹配策略:
    - S/P 核心标签: AND 逻辑（必须全部匹配）
    - K 技能标签: OR 逻辑（匹配任一即可）
    - B/V 传导来的标签: OR 逻辑
    """
    try:
        sb = db()
        if not tags:
            return []

        # 分离核心标签（S/P）和技能标签（K）
        core_tags = {t for t in tags if t.startswith(("S", "P"))}
        skill_tags = {t for t in tags if t.startswith("K")}
        bv_tags = {t for t in tags if t.startswith(("B", "V"))}
        # 传导目标标签（从 B/V 传导来的 P/K）
        conduction_targets = set()
        if conduction_map:
            for targets in conduction_map.values():
                conduction_targets.update(targets)

        # 查所有包含任一非B/V标签的 resource_id
        query_tags = list(core_tags | skill_tags | conduction_targets)
        if not query_tags:
            return []
        res = sb.table("resource_tags").select("resource_id, tag_id").in_("tag_id", query_tags).execute().data

        # 资源匹配: 核心 AND + 技能 OR
        resource_tag_map: dict[str, set[str]] = {}
        for r in res:
            resource_tag_map.setdefault(r["resource_id"], set()).add(r["tag_id"])

        matched_ids = []
        for rid, res_tags in resource_tag_map.items():
            # 核心 S/P 标签必须全部匹配（AND）
            if core_tags and not core_tags.issubset(res_tags):
                continue
            matched_ids.append(rid)

        if not matched_ids:
            # 如果核心 AND 太严格，降级为 OR
            matched_ids = list(resource_tag_map.keys())

        if not matched_ids:
            return []
        # 查 resources，限制最多 50 个
        rows = sb.table("resources").select("*").in_("id", matched_ids[:50]).eq("is_published", True).execute().data
        resources = []
        for r in rows:
            tag_res = sb.table("resource_tags").select("tag_id").eq("resource_id", r["id"]).execute().data
            r["tags"] = [t["tag_id"] for t in tag_res]
            resources.append(_row_to_camel(r))
        return resources
    except Exception as e:
        logger.warning(f"获取资源失败，降级到 JSON: {e}")
        all_resources = _load_json_fallback().get("resourceLibrary", [])
        return [r for r in all_resources if any(t in r.get("tags", []) for t in tags)]


def _keyword_fallback(query: str) -> tuple[list[str], float, str]:
    """关键词匹配 Fallback（五维标签 + 传导扩展）"""
    q = query.lower()
    matched: set[str] = set()

    keyword_map = [
        (r"冷|不理|态度|破冰|拒绝", ["S01", "P03"]),
        (r"贵|利润|钱|价格|赚|便宜|价差|不划算", ["P01", "V01"]),
        (r"五粮液|竞品|排他|签了|剑南春|茅台", ["P02"]),
        (r"宴席|酒席|婚宴|升学|寿宴|办酒", ["S02", "V02"]),
        (r"库存|压货|资金|周转", ["P04", "V05"]),
        (r"团购|企业|单位|批量|采购", ["S03"]),
        (r"餐饮|酒店|餐厅", ["S04"]),
        (r"没人要|不认识|不好卖|卖不开", ["P05", "V03"]),
        (r"跌价|亏|赔|倒挂|风险", ["P06"]),
        (r"品牌|渗透|知名|认知|市场覆盖", ["B01"]),
        (r"渠道|扩张|终端开发|铺货|收入|增长", ["B02"]),
        (r"份额|竞争|排名|领先", ["B03"]),
        (r"开瓶|消费|扫码|真实消费|动销数据", ["B04"]),
        (r"复购|持续|稳定|忠诚|长期合作", ["B05"]),
        (r"面子|档次|高端|排面|送礼", ["V02"]),
        (r"信任|国宝|历史|认证|450年", ["V03"]),
        (r"支持|动销|帮忙|品鉴会|活动", ["V04"]),
        (r"保障|以旧换新|退货|风险保障", ["V05"]),
    ]
    for pattern, tag_ids in keyword_map:
        if re.search(pattern, q):
            for t in tag_ids:
                matched.add(t)

    from app.core.llm.prompt_builder import apply_auto_association, apply_conduction
    tags, conduction_map = apply_conduction(list(matched))

    return tags, 0.7, "关键词匹配结果（AI 模型暂不可用）"


# ═══════════ 场景树接口 ═══════════
@app.get("/api/scenario/tree")
async def get_scenario_tree():
    """获取完整场景树（返回 {L1:[], L2:[], L3:[]} 格式）"""
    try:
        sb = db()
        query = sb.table("scenario_tree_nodes").select("*").order("sort_order")
        role_id = _tenant_role.get()
        if role_id:
            query = query.or_(f"role_id.eq.{role_id},role_id.is.null")
        rows = query.execute().data
        l1, l2, l3 = [], [], []
        for r in rows:
            node = {
                "id": r["id"],
                "title": r["title"],
                "description": r.get("description", ""),
                "icon": r.get("icon", ""),
            }
            if r["level"] == 1:
                l1.append(node)
            elif r["level"] == 2:
                node["parentId"] = r["parent_id"]
                node["gradient"] = r.get("gradient", "")
                l2.append(node)
            elif r["level"] == 3:
                node["parentId"] = r["parent_id"]
                node["tags"] = r.get("tags", [])
                node["painpoint"] = r.get("painpoint", "")
                if r.get("skill"):
                    node["skill"] = r["skill"]
                if r.get("danger_level"):
                    node["dangerLevel"] = r["danger_level"]
                l3.append(node)

        return {"L1": l1, "L2": l2, "L3": l3}
    except Exception as e:
        logger.warning(f"数据库请求失败，降级到 JSON: {e}")
        data = _load_json_fallback()
        return {
            "L1": data["scenarioTree"]["L1"],
            "L2": data["scenarioTree"]["L2"],
            "L3": data["scenarioTree"]["L3"],
        }


# ═══════════ 四维度接口（新增）═══════════
@app.get("/api/scenario/dimensions")
async def get_scenario_dimensions(l3_node_id: str | None = None):
    """获取四维度数据"""
    try:
        sb = db()
        query = sb.table("scenario_dimensions").select("*")
        if l3_node_id:
            query = query.eq("l3_node_id", l3_node_id)
        role_id = _tenant_role.get()
        if role_id:
            # 过滤：只返回该角色关联的 L3 场景维度
            role_scenario_ids = sb.table("scenario_tree_nodes").select("id").eq("role_id", role_id).execute().data or []
            if role_scenario_ids:
                valid_ids = [s["id"] for s in role_scenario_ids]
                query = query.in_("l3_node_id", valid_ids)
        rows = query.execute().data

        # 按 l3_node_id 分组
        result = {}
        for r in rows:
            node_id = r["l3_node_id"]
            dim_type = r["dimension_type"]
            if node_id not in result:
                result[node_id] = {"painPoints": [], "customerExperience": [], "companyValue": [], "skills": []}

            if dim_type == "painPoints":
                result[node_id]["painPoints"].append({
                    "tagId": r.get("tag_id"),
                    "summary": r.get("summary", ""),
                    "why": r.get("why", ""),
                    "howToEliminate": r.get("how_to_eliminate", ""),
                    "howToImprove": r.get("how_to_improve", ""),
                })
            elif dim_type == "customerExperience":
                result[node_id]["customerExperience"].append({
                    "title": r.get("title", ""),
                    "description": r.get("description", ""),
                })
            elif dim_type == "companyValue":
                result[node_id]["companyValue"].append({
                    "title": r.get("title", ""),
                    "description": r.get("description", ""),
                })
            elif dim_type == "skills":
                result[node_id]["skills"].append({
                    "tagId": r.get("tag_id"),
                    "summary": r.get("summary", ""),
                })

        if l3_node_id:
            return result.get(l3_node_id, {"painPoints": [], "customerExperience": [], "companyValue": [], "skills": []})
        return result
    except Exception as e:
        logger.warning(f"数据库请求失败，降级到 JSON: {e}")
        data = _load_json_fallback()
        dims = data.get("scenarioDimensions", {})
        if l3_node_id:
            return dims.get(l3_node_id, {"painPoints": [], "customerExperience": [], "companyValue": [], "skills": []})
        return dims


# ═══════════ 痛点矩阵接口（从场景维度数据自动聚合）═══════════
@app.get("/api/painpoint/matrix")
async def get_painpoint_matrix(focus_scenario: str | None = None):
    """痛点矩阵：按 L2 聚合其下所有 L3 的 painPoints 维度 + 关联技能，dangerLevel 自动计算"""
    try:
        sb = db()
        role_id = _tenant_role.get()

        # 1. 获取所有已发布的 L2 节点
        query = sb.table("scenario_tree_nodes").select("id", "title", "icon", "gradient").eq("level", 2).eq("is_published", True)
        if focus_scenario:
            query = query.eq("id", focus_scenario)
        if role_id:
            query = query.or_(f"role_id.eq.{role_id},role_id.is.null")
        l2_nodes = query.order("sort_order").execute().data or []

        # 2. 获取所有 L3 节点（用于 tag 频率统计）
        l3_query = sb.table("scenario_tree_nodes").select("id", "tags", "parent_id", "title").eq("level", 3).eq("is_published", True)
        if role_id:
            l3_query = l3_query.or_(f"role_id.eq.{role_id},role_id.is.null")
        all_l3 = l3_query.execute().data or []

        # 3. 痛点标签频率统计
        tag_freq: dict[str, int] = {}
        for node in all_l3:
            for t in (node.get("tags") or []):
                if t.startswith("P"):
                    tag_freq[t] = tag_freq.get(t, 0) + 1
        max_freq = max(tag_freq.values(), default=1)

        # 4. 获取所有 L2 下 L3 的 painPoints 维度数据
        l3_ids = [n["id"] for n in all_l3]
        dims = []
        if l3_ids:
            dims = sb.table("scenario_dimensions").select("l3_node_id", "dimension_type", "tag_id", "summary", "why", "how_to_eliminate", "sort_order").eq("is_published", True).eq("dimension_type", "painPoints").in_("l3_node_id", l3_ids).order("sort_order").execute().data or []

        # 5. 获取痛点-技能关联
        pp_tags_in_dims = list({d["tag_id"] for d in dims if d.get("tag_id")})
        skill_map: dict[str, list[str]] = {}
        if pp_tags_in_dims:
            try:
                assoc_result = sb.rpc("get_painpoint_associations", {"painpoint_ids": pp_tags_in_dims}).execute()
                if assoc_result.data:
                    for row in assoc_result.data:
                        pid = row["painpoint_id"]
                        sid = row["skill_id"]
                        if pid not in skill_map:
                            skill_map[pid] = []
                        skill_map[pid].append(sid)
            except Exception:
                pass

        # 6. 获取标签名称映射
        all_tag_ids = list({d["tag_id"] for d in dims if d.get("tag_id")}) + [s for sl in skill_map.values() for s in sl]
        tag_labels: dict[str, str] = {}
        if all_tag_ids:
            tag_rows = sb.table("tags").select("id", "label").in_("id", all_tag_ids).execute().data or []
            tag_labels = {t["id"]: t["label"] for t in tag_rows}

        # 7. 按 L2 分组聚合
        l3_parent_map = {n["id"]: n["parent_id"] for n in all_l3}
        l3_title_map = {n["id"]: n["title"] for n in all_l3}

        steps = []
        for idx, l2 in enumerate(l2_nodes):
            l2_id = l2["id"]
            # 该 L2 下所有 L3 的 painPoints
            l3_under = [n["id"] for n in all_l3 if n.get("parent_id") == l2_id]
            l2_dims = [d for d in dims if d["l3_node_id"] in l3_under]

            if not l2_dims:
                continue

            # 聚合痛点
            pp_summaries = []
            pp_tag_ids = []
            for d in l2_dims:
                scene_title = l3_title_map.get(d["l3_node_id"], "")
                summary = d.get("summary", "")
                if summary:
                    pp_summaries.append(f"[{scene_title}] {summary}")
                if d.get("tag_id"):
                    pp_tag_ids.append(d["tag_id"])

            # 聚合客户抗拒（从 why 字段提取）
            customer_resistances = [d.get("why", "") for d in l2_dims if d.get("why")]
            customer_text = "；".join(customer_resistances[:3]) if customer_resistances else ""

            # 聚合技能标签
            all_skill_ids: list[str] = []
            for tid in pp_tag_ids:
                all_skill_ids.extend(skill_map.get(tid, []))
            all_skill_ids = list(dict.fromkeys(all_skill_ids))  # 去重保序

            # dangerLevel 自动计算
            if pp_tag_ids:
                pp_freqs = [tag_freq.get(t, 0) for t in pp_tag_ids if t in tag_freq]
                top_freq = max(pp_freqs, default=1)
                danger = round(top_freq / max_freq * 100) if max_freq > 0 else 0
            else:
                danger = 30

            if danger >= 80:
                badge = "极高危"
            elif danger >= 60:
                badge = "高危"
            elif danger >= 40:
                badge = "中等"
            else:
                badge = "可控"

            steps.append({
                "step": idx + 1,
                "title": l2["title"],
                "subtitle": f"{len(l3_under)} 个场景 · {len(pp_tag_ids)} 个痛点",
                "icon": l2.get("icon", "fa-solid fa-layer-group"),
                "dangerLevel": danger,
                "badge": badge,
                "customerResistance": customer_text,
                "painPoint": "\n".join(pp_summaries[:3]),
                "dataInsight": "",
                "painpointTags": pp_tag_ids,
                "skillTags": all_skill_ids,
            })

        return {"steps": steps, "focusScenario": focus_scenario or (l2_nodes[0]["title"] if l2_nodes else "")}
    except Exception as e:
        logger.warning(f"痛点矩阵聚合失败，降级到 JSON: {e}")
        return _load_json_fallback().get("painPointMatrix", {})


# ═══════════ 等级配置接口 ═══════════
@app.get("/api/levels")
async def get_levels():
    """获取等级配置"""
    try:
        sb = db()

        rows = sb.table("level_config").select("*").order("level").execute().data
        levels = []
        for r in rows:
            levels.append({
                "id": r["level"],
                "shortName": r.get("short_name", ""),
                "fullName": r.get("full_name", ""),
                "period": r.get("period", ""),
                "challenges": r.get("challenges", []),
                "loots": r.get("loots", []),
                "subLevels": r.get("sub_levels", []),
            })

        return {"levels": levels}
    except Exception as e:
        logger.warning(f"数据库请求失败，降级到 JSON: {e}")
        return {"levels": _load_json_fallback().get("levelConfig", [])}


# ═══════════ Feature Flag 系统 ═══════════
from app.core.feature_flags import (
    get_feature_flags, get_all_flags, require_feature,
    invalidate_cache, get_subscription, TIER_ORDER,
)


# ═══════════ Admin API（JWT 认证）═══════════

@app.post("/admin/api/auth/login")
async def admin_login(request: dict):
    """管理员登录，返回 JWT"""
    username = request.get("username", "")
    password = request.get("password", "")
    if not username or not password:
        raise HTTPException(status_code=400, detail="用户名和密码不能为空")

    sb = admin_db()
    res = sb.table("admin_users").select("*").eq("username", username).eq("is_active", True).execute()
    if not res.data:
        raise HTTPException(status_code=401, detail="用户名或密码错误")

    user = res.data[0]
    if not verify_password(password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="用户名或密码错误")

    token = create_access_token(data={"sub": str(user["id"])})
    return {
        "access_token": token,
        "token_type": "bearer",
        "admin": {
            "id": user["id"],
            "username": user["username"],
            "display_name": user.get("display_name", ""),
            "role": user.get("role", "editor"),
        },
    }


@app.get("/admin/api/auth/me")
async def admin_me(admin: dict = _Depends(get_current_admin)):
    return admin


@app.get("/admin/api/dashboard/stats")
async def admin_dashboard_stats(admin: dict = _Depends(get_current_admin)):
    """Dashboard 统计（按租户 + 角色）"""
    sb = db()
    tables = ["tags", "resources", "scenario_tree_nodes", "scenario_dimensions", "pain_point_matrix", "level_config"]
    stats = {}
    role_id = _tenant_role.get()
    for t in tables:
        try:
            q = sb.table(t).select("*", count="exact")
            q = _apply_role_filter(q, t)
            total_r = q.execute()
            total = total_r.count or 0
            qp = sb.table(t).select("*", count="exact").eq("is_published", True)
            qp = _apply_role_filter(qp, t)
            pub_r = qp.execute()
            published = pub_r.count or 0
            stats[t] = {"total": total, "published": published, "draft": total - published}
        except Exception:
            stats[t] = {"total": 0, "published": 0, "draft": 0}
    return stats


# ── Tags CRUD ──
@app.get("/admin/api/tags")
async def admin_list_tags(dimension: str | None = None, page: int = 1, page_size: int = 50,
                          admin: dict = _Depends(get_current_admin)):
    filters = {"dimension": dimension} if dimension else None
    # 按角色过滤：匹配 tags.role_id 或 source_scenario_id 关联的场景
    role_id = _tenant_role.get()
    if role_id:
        sb = db()
        try:
            # Fetch tags matching role_id
            q = sb.table("tags").select("*").eq("role_id", role_id)
            if dimension:
                q = q.eq("dimension", dimension)
            rows = q.execute().data or []
            # Also include tags with NULL role_id (shared across roles)
            qs = sb.table("tags").select("*").not_.is_("role_id", "0")
            if dimension:
                qs = qs.eq("dimension", dimension)
            shared = [r for r in (qs.execute().data or []) if r.get("role_id") is None]
            # Merge, deduplicate by id
            seen = {}
            for r in rows + shared:
                seen[r["id"]] = r
            all_rows = list(seen.values())
            tags = [_row_to_camel(r) for r in all_rows]
            return {"items": tags, "total": len(tags), "page": page, "page_size": page_size}
        except Exception as e:
            import traceback
            traceback.print_exc()
    return admin_crud.admin_list("tags", filters=filters, order="sort_order", page=page, page_size=page_size)


@app.get("/admin/api/tags/{tag_id}")
async def admin_get_tag(tag_id: str, admin: dict = _Depends(get_current_admin)):
    row = admin_crud.admin_get("tags", tag_id)
    if not row:
        raise HTTPException(status_code=404, detail="标签不存在")
    return _row_to_camel(row)


@app.post("/admin/api/tags")
async def admin_create_tag(body: dict, admin: dict = _Depends(get_current_admin)):
    data = _body_to_snake(body)
    result = admin_crud.admin_create("tags", data, admin["id"])
    if result:
        _upsert_tag_embedding(result)
    return result


@app.put("/admin/api/tags/{tag_id}")
async def admin_update_tag(tag_id: str, body: dict, admin: dict = _Depends(get_current_admin)):
    data = _body_to_snake(body)
    result = admin_crud.admin_update("tags", tag_id, data, admin["id"])
    if result:
        _upsert_tag_embedding(result)
    return result


def _upsert_tag_embedding(tag: dict):
    """为单个标签生成/更新 embedding（静默失败）"""
    try:
        from app.core.vector.embedding_client import embed_texts
        from app.core.vector.search import build_tag_text
        text = build_tag_text(tag)
        emb = embed_texts([text])[0]
        db().table("tags").update({"embedding": emb}).eq("id", tag["id"]).execute()
        logger.info(f"Embedding updated for tag {tag.get('id', '?')}")
    except Exception as e:
        logger.warning(f"Failed to generate embedding for tag: {e}")


@app.delete("/admin/api/tags/{tag_id}")
async def admin_delete_tag(tag_id: str, admin: dict = _Depends(get_current_admin)):
    admin_crud.admin_delete("tags", tag_id, admin["id"])
    return {"ok": True}


# ═══════════ 自动标签生成（场景 CRUD 调用）═══════════

def _auto_tag_for_scenario(sb, node_id: str, node: dict) -> dict:
    """为 L3 场景自动生成痛点/技能标签，并建立场景-技能引用。返回 {tags_created, skills_linked}"""
    result = {"tags_created": 0, "skills_linked": 0}
    level = node.get("level")
    if level != 3:
        return result

    # 1. 痛点标签生成
    painpoint_text = (node.get("painpoint") or "").strip()
    if painpoint_text:
        existing = sb.table("tags").select("id").eq("source_scenario_id", node_id).eq("dimension", "painpoint").execute().data
        if not existing:
            # 生成 ID
            all_p = sb.table("tags").select("id").eq("dimension", "painpoint").execute().data
            seq = max((int(t["id"][1:]) for t in all_p if t["id"].startswith("P") and t["id"][1:].isdigit()), default=0) + 1
            pid = f"P{seq:02d}"
            tag_data = {
                "id": pid, "label": painpoint_text[:50], "dimension": "painpoint",
                "description": painpoint_text, "trigger_keywords": [], "auto_association": [],
                "source_scenario_id": node_id, "granularity": "skill",
                "is_published": True, "is_active": True,
            }
            sb.table("tags").insert(tag_data).execute()
            result["tags_created"] += 1
            logger.info(f"Auto-created painpoint tag {pid} for scenario {node_id}")
            # 异步生成 embedding（静默失败）
            _upsert_tag_embedding(tag_data)

    # 2. 技能标签生成 + 场景引用
    skill_text = (node.get("skill") or "").strip()
    node_tags = node.get("tags") or []

    # 2a. 从 tags[] 中提取已有 K* 标签，建立引用
    for tid in node_tags:
        if tid.startswith("K"):
            tag_exists = sb.table("tags").select("id").eq("id", tid).execute().data
            if tag_exists:
                try:
                    sb.table("scenario_skills").insert({
                        "scenario_id": node_id, "skill_id": tid, "context": "primary",
                    }).execute()
                    result["skills_linked"] += 1
                except Exception:
                    pass  # 已存在

    # 2b. 为 skill 文本创建标签（如果不存在）
    if skill_text:
        existing_k = sb.table("tags").select("id").eq("source_scenario_id", node_id).eq("dimension", "skill").execute().data
        if not existing_k:
            all_k = sb.table("tags").select("id").eq("dimension", "skill").execute().data
            seq = max((int(t["id"][1:]) for t in all_k if t["id"].startswith("K") and t["id"][1:].isdigit()), default=0) + 1
            kid = f"K{seq:02d}"
            tag_data = {
                "id": kid, "label": skill_text[:50], "dimension": "skill",
                "description": skill_text, "trigger_keywords": [], "auto_association": [],
                "source_scenario_id": node_id, "granularity": "skill",
                "is_published": True, "is_active": True,
            }
            sb.table("tags").insert(tag_data).execute()
            result["tags_created"] += 1
            logger.info(f"Auto-created skill tag {kid} for scenario {node_id}")
            _upsert_tag_embedding(tag_data)
            # 建立引用
            try:
                sb.table("scenario_skills").insert({
                    "scenario_id": node_id, "skill_id": kid, "context": "primary",
                }).execute()
                result["skills_linked"] += 1
            except Exception:
                pass

    # 3. 自动生成痛点-技能关联
    _generate_painpoint_associations(sb, node_id)

    return result


def _generate_painpoint_associations(sb, scenario_id: str):
    """为场景自动生成痛点-技能关联（静默失败）"""
    try:
        # 获取该场景的痛点标签
        p_tags = sb.table("tags").select("id").eq("source_scenario_id", scenario_id).eq("dimension", "painpoint").eq("is_published", True).execute().data
        # 获取该场景的技能引用
        s_refs = sb.table("scenario_skills").select("skill_id").eq("scenario_id", scenario_id).execute().data

        if not p_tags or not s_refs:
            return

        for p in p_tags:
            for s in s_refs:
                try:
                    sb.table("painpoint_skill_associations").insert({
                        "painpoint_id": p["id"],
                        "skill_id": s["skill_id"],
                        "source_scenario_id": scenario_id,
                        "weight": 1.0,
                    }).execute()
                except Exception:
                    pass  # 已存在
    except Exception as e:
        logger.warning(f"Failed to generate associations for {scenario_id}: {e}")


@app.post("/admin/api/scenario-tree/{node_id}/auto-tag")
async def admin_auto_tag_scenario(node_id: str, admin: dict = _Depends(get_current_admin)):
    """手动触发场景自动标签生成"""
    sb = db()
    rows = sb.table("scenario_tree_nodes").select("*").eq("id", node_id).execute().data
    if not rows:
        raise HTTPException(status_code=404, detail="场景不存在")
    result = _auto_tag_for_scenario(sb, node_id, rows[0])
    return {"ok": True, **result}


@app.post("/admin/api/tags/generate-embeddings")
async def admin_generate_embeddings(body: dict, admin: dict = _Depends(get_current_admin)):
    """重新生成全部标签的 embedding"""
    from app.core.vector.embedding_client import embed_texts
    from app.core.vector.search import build_tag_text

    sb = db()
    tag_ids = body.get("tagIds")  # optional: specific tags
    if tag_ids:
        rows = sb.table("tags").select("*").in_("id", tag_ids).execute().data
    else:
        rows = sb.table("tags").select("*").eq("is_published", True).execute().data

    if not rows:
        return {"updated": 0, "message": "没有需要处理的标签"}

    texts = [build_tag_text(row) for row in rows]
    embeddings = embed_texts(texts)

    updated = 0
    for row, emb in zip(rows, embeddings):
        sb.table("tags").update({"embedding": emb}).eq("id", row["id"]).execute()
        updated += 1

    return {"updated": updated, "message": f"已生成 {updated} 个标签的向量"}


@app.post("/admin/api/tags/{tag_id}/publish")
async def admin_publish_tag(tag_id: str, admin: dict = _Depends(get_current_admin)):
    return admin_crud.admin_toggle_publish("tags", tag_id, admin["id"], True)


@app.post("/admin/api/tags/{tag_id}/unpublish")
async def admin_unpublish_tag(tag_id: str, admin: dict = _Depends(get_current_admin)):
    return admin_crud.admin_toggle_publish("tags", tag_id, admin["id"], False)


# ── Resources CRUD ──
@app.get("/admin/api/resources")
async def admin_list_resources(type: str | None = None, page: int = 1, page_size: int = 50,
                                admin: dict = _Depends(get_current_admin)):
    filters = {"type": type} if type else None
    return admin_crud.admin_list("resources", filters=filters, order="created_at", page=page, page_size=page_size)


@app.get("/admin/api/resources/{res_id}")
async def admin_get_resource(res_id: str, admin: dict = _Depends(get_current_admin)):
    row = admin_crud.admin_get("resources", res_id)
    if not row:
        raise HTTPException(status_code=404, detail="资源不存在")
    # 补充 tags 和 skills
    sb = db()
    tags_res = sb.table("resource_tags").select("tag_id").eq("resource_id", res_id).execute().data
    row["tags"] = [t["tag_id"] for t in tags_res]
    skills_res = sb.table("resource_skills").select("skill_item_id, skill_dimension_id").eq("resource_id", res_id).execute().data or []
    row["skillIds"] = [s["skill_item_id"] for s in skills_res if s.get("skill_item_id")]
    row["skillDimensionIds"] = [s["skill_dimension_id"] for s in skills_res if s.get("skill_dimension_id")]
    return _row_to_camel(row)


@app.post("/admin/api/resources")
async def admin_create_resource(body: dict, admin: dict = _Depends(get_current_admin)):
    data = _body_to_snake(body)
    tag_ids = data.pop("tags", [])
    result = admin_crud.admin_create("resources", data, admin["id"])
    # 同步 resource_tags
    if result and tag_ids:
        sb = db()
        tag_rows = [{"resource_id": result["id"], "tag_id": tid} for tid in tag_ids]
        sb.table("resource_tags").upsert(tag_rows, on_conflict="resource_id,tag_id").execute()
    # 自动推荐标签（不影响已有 tags）
    suggested = _suggest_tags_for_resource(data.get("title", ""), data.get("content", ""))
    if suggested:
        result["suggested_tags"] = suggested
    return result


@app.put("/admin/api/resources/{res_id}")
async def admin_update_resource(res_id: str, body: dict, admin: dict = _Depends(get_current_admin)):
    data = _body_to_snake(body)
    tag_ids = data.pop("tags", None)
    result = admin_crud.admin_update("resources", res_id, data, admin["id"])
    # 同步 resource_tags（仅在 tags 字段存在时更新）
    if tag_ids is not None:
        sb = db()
        sb.table("resource_tags").delete().eq("resource_id", res_id).execute()
        if tag_ids:
            tag_rows = [{"resource_id": res_id, "tag_id": tid} for tid in tag_ids]
            sb.table("resource_tags").upsert(tag_rows, on_conflict="resource_id,tag_id").execute()
    # 更新时也推荐标签
    suggested = _suggest_tags_for_resource(data.get("title", ""), data.get("content", ""))
    if suggested:
        result["suggested_tags"] = suggested
    return result


@app.post("/admin/api/resources/{res_id}/apply-tags")
async def admin_apply_suggested_tags(res_id: str, body: dict, admin: dict = _Depends(get_current_admin)):
    """确认并写入系统推荐的标签"""
    tag_ids = body.get("tag_ids", [])
    if not tag_ids:
        raise HTTPException(status_code=400, detail="tag_ids 不能为空")
    sb = db()
    # 清除旧标签，写入新标签
    sb.table("resource_tags").delete().eq("resource_id", res_id).execute()
    tag_rows = [{"resource_id": res_id, "tag_id": tid} for tid in tag_ids]
    sb.table("resource_tags").upsert(tag_rows, on_conflict="resource_id,tag_id").execute()
    # 记录审计
    admin_crud.log_audit(sb, admin["id"], "resource_tags", res_id, "update", new_values={"tag_ids": tag_ids})
    return {"ok": True, "applied": len(tag_ids)}


@app.delete("/admin/api/resources/{res_id}")
async def admin_delete_resource(res_id: str, admin: dict = _Depends(get_current_admin)):
    admin_crud.admin_delete("resources", res_id, admin["id"])
    return {"ok": True}


@app.post("/admin/api/resources/{res_id}/publish")
async def admin_publish_resource(res_id: str, admin: dict = _Depends(get_current_admin)):
    return admin_crud.admin_toggle_publish("resources", res_id, admin["id"], True)


@app.post("/admin/api/resources/{res_id}/unpublish")
async def admin_unpublish_resource(res_id: str, admin: dict = _Depends(get_current_admin)):
    return admin_crud.admin_toggle_publish("resources", res_id, admin["id"], False)


# ── Scenario Tree CRUD ──
@app.get("/admin/api/scenario-tree")
async def admin_list_scenario_tree(level: int | None = None, parent_id: str | None = None,
                                    page: int = 1, page_size: int = 50,
                                    admin: dict = _Depends(get_current_admin)):
    filters = {}
    if level is not None:
        filters["level"] = level
    if parent_id is not None:
        filters["parent_id"] = parent_id
    return admin_crud.admin_list("scenario_tree_nodes", filters=filters or None,
                                 order="id", page=page, page_size=page_size)


@app.get("/admin/api/scenario-tree/{node_id}")
async def admin_get_scenario_node(node_id: str, admin: dict = _Depends(get_current_admin)):
    row = admin_crud.admin_get("scenario_tree_nodes", node_id)
    if not row:
        raise HTTPException(status_code=404, detail="场景节点不存在")
    return _row_to_camel(row)


@app.post("/admin/api/scenario-tree/batch")
async def admin_batch_scenario_nodes(body: dict, admin: dict = _Depends(get_current_admin)):
    """批量创建场景树节点（先 L1，再 L2，最后 L3，保证父子顺序）"""
    sb = db()
    nodes = body.get("nodes", [])
    if not nodes:
        raise HTTPException(status_code=400, detail="nodes 不能为空")

    role_id = _tenant_role.get()

    # 按层级排序：L1 → L2 → L3
    sorted_nodes = sorted(nodes, key=lambda n: n.get("level", 3))
    results = []
    for node in sorted_nodes:
        data = _body_to_snake(node)
        node_id = (data.get("id") or "").strip()
        if not node_id:
            raise HTTPException(status_code=400, detail="节点 ID 不能为空")
        data["id"] = node_id
        if isinstance(data.get("tags"), list):
            data["tags"] = "{" + ",".join(str(t) for t in data["tags"]) + "}"
        data.setdefault("is_active", True)
        data.setdefault("is_published", False)
        if role_id and "role_id" not in data:
            data["role_id"] = role_id
        try:
            result = sb.table("scenario_tree_nodes").insert(data).execute()
            action = "create"
        except Exception as e:
            if "duplicate key" in str(e):
                # 已存在则更新
                result = sb.table("scenario_tree_nodes").update(data).eq("id", data["id"]).execute()
                action = "update"
            else:
                raise
        if result.data:
            results.append(result.data[0])
            admin_crud.log_audit(sb, admin["id"], "scenario_tree_nodes", str(data["id"]), action, new_values=data)
            # L3 节点自动生成标签
            if data.get("level") == 3:
                _auto_tag_for_scenario(sb, str(data["id"]), data)

    return {"created": len(results), "items": results}


@app.post("/admin/api/scenario-tree")
async def admin_create_scenario_node(body: dict, admin: dict = _Depends(get_current_admin)):
    data = _body_to_snake(body)
    node_id = (data.get("id") or "").strip()
    if not node_id:
        raise HTTPException(status_code=400, detail="节点 ID 不能为空")
    # 检查 ID 是否重复
    sb = db()
    existing = sb.table("scenario_tree_nodes").select("id").eq("id", node_id).execute().data
    if existing:
        raise HTTPException(status_code=409, detail=f"节点 ID '{node_id}' 已存在")
    data["id"] = node_id
    if isinstance(data.get("tags"), list):
        data["tags"] = "{" + ",".join(str(t) for t in data["tags"]) + "}"
    result = admin_crud.admin_create("scenario_tree_nodes", data, admin["id"])
    # L3 节点自动生成标签
    if data.get("level") == 3:
        sb = db()
        _auto_tag_for_scenario(sb, str(data.get("id", "")), data)
    return result


@app.put("/admin/api/scenario-tree/{node_id}")
async def admin_update_scenario_node(node_id: str, body: dict, admin: dict = _Depends(get_current_admin)):
    data = _body_to_snake(body)
    # tags 列是 TEXT[]，Supabase 客户端发 jsonb 会报类型错误，需转为 PostgreSQL 数组字面量
    if isinstance(data.get("tags"), list):
        data["tags"] = "{" + ",".join(str(t) for t in data["tags"]) + "}"

    # 处理 ID 变更（仅允许原 ID 为空时设置新 ID）
    new_id = (data.pop("id", None) or "").strip()
    actual_id = node_id  # 用于后续更新的 ID
    if new_id and new_id != node_id:
        if node_id:
            raise HTTPException(status_code=400, detail="已有 ID 的节点不允许修改 ID，请删除后重建")
        # 原 ID 为空，先设置新 ID
        sb = db()
        existing = sb.table("scenario_tree_nodes").select("id").eq("id", new_id).execute().data
        if existing:
            raise HTTPException(status_code=409, detail=f"节点 ID '{new_id}' 已存在")
        sb.table("scenario_tree_nodes").update({"id": new_id}).eq("id", "").execute()
        actual_id = new_id

    result = admin_crud.admin_update("scenario_tree_nodes", actual_id, data, admin["id"])
    # L3 节点更新时重新生成标签
    if data.get("level") == 3:
        sb = db()
        _auto_tag_for_scenario(sb, node_id, data)
    return result


@app.put("/admin/api/scenario-tree/{node_id}/level")
async def admin_change_scenario_level(node_id: str, body: dict, admin: dict = _Depends(get_current_admin)):
    """调整场景树节点层级 (L1↔L2↔L3)"""
    new_level = body.get("level")
    if new_level not in (1, 2, 3):
        raise HTTPException(status_code=400, detail="层级必须是 1、2 或 3")

    current = admin_crud.admin_get("scenario_tree_nodes", node_id)
    if not current:
        raise HTTPException(status_code=404, detail="节点不存在")

    old_level = current.get("level")
    if old_level == new_level:
        return current

    update_data: dict = {"level": new_level}
    # 升级到 L1 时清除父级
    if new_level == 1:
        update_data["parent_id"] = None

    # 级联处理子节点：避免子节点挂在错误层级的父级下
    if old_level == 1 and new_level == 2:
        # 原 L1 降为 L2，其 L2 子节点需重新挂到其他 L1 下（清除 parentId 让用户手动调整）
        sb = db()
        children = sb.table("scenario_tree_nodes").select("id").eq("parent_id", node_id).eq("level", 2).execute().data
        for child in children:
            sb.table("scenario_tree_nodes").update({"parent_id": None}).eq("id", child["id"]).execute()
    elif old_level == 2 and new_level == 1:
        # 原 L2 升为 L1，其 L3 子节点的 parent 指向一个 L2，现在该 L2 变成了 L1，需清除
        sb = db()
        children = sb.table("scenario_tree_nodes").select("id").eq("parent_id", node_id).eq("level", 3).execute().data
        for child in children:
            sb.table("scenario_tree_nodes").update({"parent_id": None}).eq("id", child["id"]).execute()

    result = admin_crud.admin_update("scenario_tree_nodes", node_id, update_data, admin["id"])
    return result


@app.delete("/admin/api/scenario-tree/{node_id}")
async def admin_delete_scenario_node(node_id: str, admin: dict = _Depends(get_current_admin)):
    admin_crud.admin_delete("scenario_tree_nodes", node_id, admin["id"])
    return {"ok": True}


@app.post("/admin/api/scenario-tree/{node_id}/publish")
async def admin_publish_scenario_node(node_id: str, admin: dict = _Depends(get_current_admin)):
    return admin_crud.admin_toggle_publish("scenario_tree_nodes", node_id, admin["id"], True)


@app.post("/admin/api/scenario-tree/{node_id}/unpublish")
async def admin_unpublish_scenario_node(node_id: str, admin: dict = _Depends(get_current_admin)):
    return admin_crud.admin_toggle_publish("scenario_tree_nodes", node_id, admin["id"], False)


# ── Scenario Dimensions CRUD ──
@app.get("/admin/api/dimensions")
async def admin_list_dimensions(l3_node_id: str | None = None, dimension_type: str | None = None,
                                page: int = 1, page_size: int = 50,
                                admin: dict = _Depends(get_current_admin)):
    filters = {}
    if l3_node_id:
        filters["l3_node_id"] = l3_node_id
    if dimension_type:
        filters["dimension_type"] = dimension_type
    return admin_crud.admin_list("scenario_dimensions", filters=filters or None,
                                 order="sort_order", page=page, page_size=page_size)


@app.get("/admin/api/dimensions/{dim_id}")
async def admin_get_dimension(dim_id: int, admin: dict = _Depends(get_current_admin)):
    row = admin_crud.admin_get("scenario_dimensions", dim_id)
    if not row:
        raise HTTPException(status_code=404, detail="维度数据不存在")
    return _row_to_camel(row)


@app.post("/admin/api/dimensions/batch")
async def admin_batch_dimensions(body: dict, admin: dict = _Depends(get_current_admin)):
    """批量创建四维度数据（为单个 L3 节点一次性创建全部四维度）"""
    sb = db()
    l3_node_id = body.get("l3NodeId")
    items = body.get("items", [])
    if not l3_node_id or not items:
        raise HTTPException(status_code=400, detail="l3NodeId 和 items 不能为空")

    results = []
    for i, item in enumerate(items):
        data = _body_to_snake(item)
        data["l3_node_id"] = l3_node_id
        data.setdefault("sort_order", i)
        data.setdefault("is_published", False)
        try:
            result = sb.table("scenario_dimensions").insert(data).execute()
            action = "create"
        except Exception as e:
            if "duplicate key" in str(e):
                result = sb.table("scenario_dimensions").update(data).eq("id", data["id"]).execute()
                action = "update"
            else:
                raise
        if result.data:
            results.append(result.data[0])
            admin_crud.log_audit(sb, admin["id"], "scenario_dimensions", str(result.data[0]["id"]), action, new_values=data)

    return {"created": len(results), "items": results}


@app.post("/admin/api/dimensions")
async def admin_create_dimension(body: dict, admin: dict = _Depends(get_current_admin)):
    data = _body_to_snake(body)
    return admin_crud.admin_create("scenario_dimensions", data, admin["id"])


@app.put("/admin/api/dimensions/{dim_id}")
async def admin_update_dimension(dim_id: int, body: dict, admin: dict = _Depends(get_current_admin)):
    data = _body_to_snake(body)
    return admin_crud.admin_update("scenario_dimensions", dim_id, data, admin["id"])


@app.delete("/admin/api/dimensions/{dim_id}")
async def admin_delete_dimension(dim_id: int, admin: dict = _Depends(get_current_admin)):
    admin_crud.admin_delete("scenario_dimensions", dim_id, admin["id"])
    return {"ok": True}


@app.post("/admin/api/dimensions/{dim_id}/publish")
async def admin_publish_dimension(dim_id: int, admin: dict = _Depends(get_current_admin)):
    return admin_crud.admin_toggle_publish("scenario_dimensions", dim_id, admin["id"], True)


@app.post("/admin/api/dimensions/{dim_id}/unpublish")
async def admin_unpublish_dimension(dim_id: int, admin: dict = _Depends(get_current_admin)):
    return admin_crud.admin_toggle_publish("scenario_dimensions", dim_id, admin["id"], False)


# ── Pain Point Matrix CRUD ──
@app.get("/admin/api/pain-point-matrix")
async def admin_list_pain_point_matrix(focus_scenario: str | None = None,
                                        page: int = 1, page_size: int = 50,
                                        admin: dict = _Depends(get_current_admin)):
    filters = {"focus_scenario": focus_scenario} if focus_scenario else None
    return admin_crud.admin_list("pain_point_matrix", filters=filters,
                                 order="step", page=page, page_size=page_size)


@app.get("/admin/api/pain-point-matrix/{matrix_id}")
async def admin_get_pain_point(matrix_id: int, admin: dict = _Depends(get_current_admin)):
    row = admin_crud.admin_get("pain_point_matrix", matrix_id)
    if not row:
        raise HTTPException(status_code=404, detail="痛点矩阵记录不存在")
    return _row_to_camel(row)


@app.post("/admin/api/pain-point-matrix")
async def admin_create_pain_point(body: dict, admin: dict = _Depends(get_current_admin)):
    data = _body_to_snake(body)
    return admin_crud.admin_create("pain_point_matrix", data, admin["id"])


@app.put("/admin/api/pain-point-matrix/{matrix_id}")
async def admin_update_pain_point(matrix_id: int, body: dict, admin: dict = _Depends(get_current_admin)):
    data = _body_to_snake(body)
    return admin_crud.admin_update("pain_point_matrix", matrix_id, data, admin["id"])


@app.delete("/admin/api/pain-point-matrix/{matrix_id}")
async def admin_delete_pain_point(matrix_id: int, admin: dict = _Depends(get_current_admin)):
    admin_crud.admin_delete("pain_point_matrix", matrix_id, admin["id"])
    return {"ok": True}


@app.post("/admin/api/pain-point-matrix/{matrix_id}/publish")
async def admin_publish_pain_point(matrix_id: int, admin: dict = _Depends(get_current_admin)):
    return admin_crud.admin_toggle_publish("pain_point_matrix", matrix_id, admin["id"], True)


@app.post("/admin/api/pain-point-matrix/{matrix_id}/unpublish")
async def admin_unpublish_pain_point(matrix_id: int, admin: dict = _Depends(get_current_admin)):
    return admin_crud.admin_toggle_publish("pain_point_matrix", matrix_id, admin["id"], False)


# ── Level Config CRUD ──
@app.get("/admin/api/levels")
async def admin_list_levels(page: int = 1, page_size: int = 50,
                            admin: dict = _Depends(get_current_admin)):
    return admin_crud.admin_list("level_config", order="level", page=page, page_size=page_size)


@app.get("/admin/api/levels/{level_id}")
async def admin_get_level(level_id: int, admin: dict = _Depends(get_current_admin)):
    row = admin_crud.admin_get("level_config", level_id)
    if not row:
        raise HTTPException(status_code=404, detail="等级配置不存在")
    return _row_to_camel(row)


@app.post("/admin/api/levels")
async def admin_create_level(body: dict, admin: dict = _Depends(get_current_admin)):
    data = _body_to_snake(body)
    return admin_crud.admin_create("level_config", data, admin["id"])


@app.put("/admin/api/levels/{level_id}")
async def admin_update_level(level_id: int, body: dict, admin: dict = _Depends(get_current_admin)):
    data = _body_to_snake(body)
    return admin_crud.admin_update("level_config", level_id, data, admin["id"])


@app.delete("/admin/api/levels/{level_id}")
async def admin_delete_level(level_id: int, admin: dict = _Depends(get_current_admin)):
    admin_crud.admin_delete("level_config", level_id, admin["id"])
    return {"ok": True}


@app.post("/admin/api/levels/{level_id}/publish")
async def admin_publish_level(level_id: int, admin: dict = _Depends(get_current_admin)):
    return admin_crud.admin_toggle_publish("level_config", level_id, admin["id"], True)


@app.post("/admin/api/levels/{level_id}/unpublish")
async def admin_unpublish_level(level_id: int, admin: dict = _Depends(get_current_admin)):
    return admin_crud.admin_toggle_publish("level_config", level_id, admin["id"], False)


# ── Sub-Levels (子关卡) ──

@app.get("/admin/api/sub-levels")
async def admin_list_sub_levels(level_id: int | None = None, admin: dict = _Depends(get_current_admin)):
    sb = db()
    # Pre-load resource titles for enrichment
    res_map = {r["id"]: r.get("title", "") for r in sb.table("resources").select("id, title").execute().data}
    q = sb.table("sub_levels").select("*")
    if level_id is not None:
        q = q.eq("level_id", level_id)
    result = q.order("level_id").order("sort_order").execute()
    items = []
    for r in result.data:
        # Load challenges
        chals = sb.table("sub_level_challenges").select("*").eq("sub_level_id", r["id"]).order("sort_order").execute().data
        # Load skills
        skills = sb.table("skill_level_map").select("skill_item_id, is_required, sort_order").eq("sub_level_id", r["id"]).order("sort_order").execute().data
        # Load learning tasks with packs
        tasks = sb.table("learning_tasks").select("*").eq("sub_level_id", r["id"]).order("sort_order").execute().data
        for t in tasks:
            packs = sb.table("learning_packs").select("*").eq("learning_task_id", t["id"]).order("sort_order").execute().data
            for p in packs:
                p_items = sb.table("learning_pack_items").select("resource_id, is_required, sort_order").eq("pack_id", p["id"]).order("sort_order").execute().data
                # Enrich with resource title
                for pi in p_items:
                    pi["resource_title"] = res_map.get(pi["resource_id"], "")
                p["items"] = p_items
            t["packs"] = packs
        r["challenges"] = chals
        r["skills"] = skills
        r["learning_tasks"] = tasks
        items.append(_row_to_camel(r))
    return {"items": items}


@app.put("/admin/api/sub-levels/{sub_level_id}")
async def admin_update_sub_level(sub_level_id: str, body: dict, admin: dict = _Depends(get_current_admin)):
    sb = db()
    # Update basic info
    updates = {}
    for k in ["title", "description", "sort_order", "is_published"]:
        snake_k = _camel_to_snake(k) if k[0].islower() else k.lower()
        if snake_k in body:
            updates[snake_k] = body[snake_k]
        elif k in body:
            updates[k] = body[k]
    if updates:
        sb.table("sub_levels").update(updates).eq("id", sub_level_id).execute()

    # Update challenges if provided
    if "challenges" in body:
        sb.table("sub_level_challenges").delete().eq("sub_level_id", sub_level_id).execute()
        for i, ch in enumerate(body["challenges"]):
            desc = ch if isinstance(ch, str) else ch.get("description", ch.get("text", ""))
            sb.table("sub_level_challenges").insert({"sub_level_id": sub_level_id, "description": desc, "sort_order": i}).execute()

    # Update skills if provided
    if "skillIds" in body or "skill_ids" in body:
        skill_ids = body.get("skillIds", body.get("skill_ids", []))
        sb.table("skill_level_map").delete().eq("sub_level_id", sub_level_id).execute()
        # Get level_id for this sub_level
        sl = sb.table("sub_levels").select("level_id").eq("id", sub_level_id).execute().data
        level_id = sl[0]["level_id"] if sl else 0
        for i, sid in enumerate(skill_ids):
            sb.table("skill_level_map").insert({
                "skill_item_id": sid, "level_id": level_id, "sub_level_id": sub_level_id,
                "is_required": True, "sort_order": i
            }).execute()

    # Update learning tasks if provided
    if "learningTasks" in body or "learning_tasks" in body:
        tasks = body.get("learningTasks", body.get("learning_tasks", []))
        # Delete existing tasks (cascades to packs and items)
        sb.table("learning_tasks").delete().eq("sub_level_id", sub_level_id).execute()
        for ti, task in enumerate(tasks):
            task_row = {
                "sub_level_id": sub_level_id,
                "challenge_id": task.get("challengeId", task.get("challenge_id")),
                "title": task.get("title", ""),
                "objective": task.get("objective", ""),
                "estimated_minutes": task.get("estimatedMinutes", task.get("estimated_minutes", 0)),
                "difficulty": task.get("difficulty", "medium"),
                "xp_points": task.get("xpPoints", task.get("xp_points", 0)),
                "sort_order": ti,
                "is_published": True,
            }
            result = sb.table("learning_tasks").insert(task_row).execute()
            task_id = result.data[0]["id"] if result.data else None
            if not task_id:
                continue
            # Insert packs
            for pack in task.get("packs", []):
                pack_row = {
                    "learning_task_id": task_id,
                    "pack_type": pack.get("packType", pack.get("pack_type", "pre_job")),
                    "title": pack.get("title", ""),
                    "description": pack.get("description", ""),
                    "sort_order": 0,
                }
                p_result = sb.table("learning_packs").insert(pack_row).execute()
                pack_id = p_result.data[0]["id"] if p_result.data else None
                if not pack_id:
                    continue
                # Insert pack items
                for ri, res_id in enumerate(pack.get("resourceIds", pack.get("resource_ids", []))):
                    sb.table("learning_pack_items").insert({
                        "pack_id": pack_id, "resource_id": res_id,
                        "is_required": True, "sort_order": ri
                    }).execute()

    return {"status": "ok"}


# ── Batch Publish ──
@app.post("/admin/api/batch/publish")
async def admin_batch_publish(body: dict, admin: dict = _Depends(get_current_admin)):
    """批量发布/取消发布"""
    sb = db()
    items = body.get("items", [])  # [{table: str, id: str|number}]
    publish = body.get("publish", True)
    action = "publish" if publish else "unpublish"

    results = []
    errors = []
    table_map = {
        "tags": "tags", "resources": "resources", "scenario-tree": "scenario_tree_nodes",
        "dimensions": "scenario_dimensions", "pain-point-matrix": "pain_point_matrix",
        "levels": "level_config",
    }

    for item in items:
        table_name = table_map.get(item.get("table"), item.get("table"))
        record_id = item.get("id")
        if not table_name or not record_id:
            errors.append(f"无效参数: {item}")
            continue
        try:
            result = sb.table(table_name).update({"is_published": publish}).eq("id", record_id).execute()
            if result.data:
                results.append(result.data[0])
                admin_crud.log_audit(sb, admin["id"], table_name, str(record_id), action)
            else:
                errors.append(f"未找到: {table_name}/{record_id}")
        except Exception as e:
            errors.append(f"{table_name}/{record_id}: {str(e)}")

    return {"success": len(results), "failed": len(errors), "errors": errors}


# ── File Upload（本地文件系统）──
UPLOAD_DIR = Path(os.getenv("UPLOAD_DIR", Path(__file__).parent.parent.parent / "uploads"))
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")

@app.post("/admin/api/resources/upload")
async def admin_upload_resource(body: dict, admin: dict = _Depends(get_current_admin),
                                 _=require_feature("content.file_upload")):
    """上传文件到本地文件系统（前端 base64 编码），返回公开 URL"""
    import base64 as _base64
    import uuid as _uuid

    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

    filename = body.get("filename", "")
    media_type = body.get("mediaType", "")
    file_b64 = body.get("file", "")

    if not filename or not media_type:
        raise HTTPException(status_code=400, detail="filename 和 mediaType 不能为空")
    if not file_b64:
        raise HTTPException(status_code=400, detail="file 数据不能为空")

    try:
        file_bytes = _base64.b64decode(file_b64)
    except Exception:
        raise HTTPException(status_code=400, detail="file base64 解码失败")

    # 确定子目录
    type_dir_map = {"image": "images", "audio": "audio", "video": "video"}
    primary_type = media_type.split("/")[0]
    subdir = type_dir_map.get(primary_type, "files")

    # 生成唯一路径
    unique_name = f"{_uuid.uuid4().hex[:8]}_{filename}"
    file_dir = UPLOAD_DIR / subdir
    file_dir.mkdir(parents=True, exist_ok=True)
    file_path = file_dir / unique_name

    try:
        file_path.write_bytes(file_bytes)
        url = f"/uploads/{subdir}/{unique_name}"
        return {"url": url, "path": str(file_path), "filename": filename}
    except Exception as e:
        logger.error(f"文件上传失败: {e}")
        raise HTTPException(status_code=500, detail=f"文件上传失败: {str(e)}")


# ═══════════ 角色管理 API ═══════════

@app.get("/admin/api/roles")
async def admin_list_roles(
    all_roles: bool = False,
    admin: dict = _Depends(get_current_admin),
):
    """获取角色列表。all_roles=true 时跳过角色过滤，返回租户全部角色"""
    sb = db()
    role_id = _tenant_role.get()
    q = sb.table("roles").select("*").order("sort_order")
    if not all_roles and role_id:
        q = q.eq("id", role_id)
    return {"roles": q.execute().data or []}

@app.post("/admin/api/roles")
async def admin_create_role(body: dict, admin: dict = _Depends(get_current_admin)):
    """创建角色"""
    sb = db()
    data = {k: v for k, v in body.items() if k in ("name", "slug", "description", "icon", "color", "sort_order")}
    if not data.get("name") or not data.get("slug"):
        raise HTTPException(status_code=400, detail="name 和 slug 必填")
    result = sb.table("roles").insert(data).execute()
    return {"role": result.data[0] if result.data else None}

@app.put("/admin/api/roles/{role_id}")
async def admin_update_role(role_id: int, body: dict, admin: dict = _Depends(get_current_admin)):
    """更新角色"""
    sb = db()
    allowed = {"name", "slug", "description", "icon", "color", "sort_order", "is_active"}
    data = {k: v for k, v in body.items() if k in allowed}
    sb.table("roles").update(data).eq("id", role_id).execute()
    return {"status": "ok"}

@app.delete("/admin/api/roles/{role_id}")
async def admin_delete_role(role_id: int, admin: dict = _Depends(get_current_admin)):
    """删除角色"""
    sb = db()
    # 检查是否有关联数据
    items = sb.table("skill_items").select("id").eq("role_id", role_id).execute().data or []
    if items:
        raise HTTPException(status_code=400, detail=f"该角色下有 {len(items)} 个能力项，请先移除")
    sb.table("roles").delete().eq("id", role_id).execute()
    return {"status": "ok"}


# ═══════════ 资源-技能关联 API ═══════════

@app.get("/admin/api/resource-skills")
async def admin_get_resource_skills(resource_id: int, admin: dict = _Depends(get_current_admin)):
    """获取资源关联的技能"""
    sb = db()
    skills = sb.table("resource_skills").select("*, skill_dimensions(name), skill_items(name, dimension_id)").eq("resource_id", resource_id).order("resource_skills.sort_order").execute().data or []
    return {"skills": skills}

@app.put("/admin/api/resources/{resource_id}/skills")
async def admin_set_resource_skills(resource_id: int, body: dict, admin: dict = _Depends(get_current_admin)):
    """设置资源的技能关联（全量替换）"""
    sb = db()
    skill_ids = body.get("skill_ids", [])
    # 先清除旧关联
    sb.table("resource_skills").delete().eq("resource_id", resource_id).execute()
    # 插入新关联
    for i, sid in enumerate(skill_ids):
        if sid:
            sb.table("resource_skills").insert({"resource_id": resource_id, "skill_item_id": sid, "sort_order": i}).execute()
    return {"status": "ok", "count": len(skill_ids)}


# ═══════════ 技能管理体系 API ═══════════

@app.get("/admin/api/skill-dimensions")
async def admin_list_skill_dimensions(admin: dict = _Depends(get_current_admin)):
    sb = db()
    q = _apply_role_filter(sb.table("skill_dimensions").select("*").order("sort_order"), "skill_dimensions")
    result = q.execute()
    return {"dimensions": result.data or []}

@app.post("/admin/api/skill-dimensions")
async def admin_create_skill_dimension(body: dict, admin: dict = _Depends(get_current_admin)):
    sb = db()
    role_id = _tenant_role.get()
    data = {k: v for k, v in body.items() if k in ("id", "name", "description", "icon", "sort_order", "is_active")}
    if role_id and "role_id" not in data:
        data["role_id"] = role_id
    result = sb.table("skill_dimensions").insert(data).execute()
    return {"dimension": result.data[0] if result.data else None}

@app.put("/admin/api/skill-dimensions/{dim_id}")
async def admin_update_skill_dimension(dim_id: str, body: dict, admin: dict = _Depends(get_current_admin)):
    sb = db()
    data = {k: v for k, v in body.items() if k in ("name", "description", "icon", "sort_order", "is_active")}
    sb.table("skill_dimensions").update(data).eq("id", dim_id).execute()
    return {"status": "ok"}

@app.delete("/admin/api/skill-dimensions/{dim_id}")
async def admin_delete_skill_dimension(dim_id: str, admin: dict = _Depends(get_current_admin)):
    sb = db()
    sb.table("skill_dimensions").delete().eq("id", dim_id).execute()
    return {"status": "ok"}

@app.get("/admin/api/skill-items")
async def admin_list_skill_items(dimension_id: str | None = None, admin: dict = _Depends(get_current_admin)):
    sb = db()
    q = _apply_role_filter(sb.table("skill_items").select("*, skill_dimensions(name)"), "skill_items")
    if dimension_id:
        q = q.eq("dimension_id", dimension_id)
    result = q.order("skill_items.sort_order").execute()
    items = []
    for r in (result.data or []):
        r["dimension_name"] = r.get("skill_dimensions", {}).get("name", "") if isinstance(r.get("skill_dimensions"), dict) else ""
        items.append(r)
    return {"items": items}

@app.post("/admin/api/skill-items")
async def admin_create_skill_item(body: dict, admin: dict = _Depends(get_current_admin)):
    sb = db()
    role_id = _tenant_role.get()
    data = {k: v for k, v in body.items() if k in ("id", "dimension_id", "name", "description", "sort_order", "is_active")}
    if role_id and "role_id" not in data:
        data["role_id"] = role_id
    result = sb.table("skill_items").insert(data).execute()
    return {"item": result.data[0] if result.data else None}

@app.put("/admin/api/skill-items/{item_id}")
async def admin_update_skill_item(item_id: str, body: dict, admin: dict = _Depends(get_current_admin)):
    sb = db()
    data = {k: v for k, v in body.items() if k in ("name", "description", "dimension_id", "sort_order", "is_active")}
    sb.table("skill_items").update(data).eq("id", item_id).execute()
    return {"status": "ok"}

@app.delete("/admin/api/skill-items/{item_id}")
async def admin_delete_skill_item(item_id: str, admin: dict = _Depends(get_current_admin)):
    sb = db()
    sb.table("skill_items").delete().eq("id", item_id).execute()
    return {"status": "ok"}

@app.get("/admin/api/skill-level-map")
async def admin_get_skill_level_map(skill_item_id: str | None = None, level_id: int | None = None, admin: dict = _Depends(get_current_admin)):
    sb = db()
    q = sb.table("skill_level_map").select("*")
    if skill_item_id:
        q = q.eq("skill_item_id", skill_item_id)
    if level_id is not None:
        q = q.eq("level_id", level_id)
    result = q.execute()
    return {"mappings": result.data or []}

@app.put("/admin/api/skill-level-map")
async def admin_update_skill_level_map(body: dict, admin: dict = _Depends(get_current_admin)):
    sb = db()
    skill_item_id = body.get("skill_item_id")
    levels = body.get("levels", [])
    if not skill_item_id:
        return {"status": "error", "detail": "skill_item_id required"}
    sb.table("skill_level_map").delete().eq("skill_item_id", skill_item_id).execute()
    if levels:
        rows = [{"skill_item_id": skill_item_id, **l} for l in levels]
        sb.table("skill_level_map").insert(rows).execute()
    return {"status": "ok"}

@app.get("/admin/api/skills/tree")
async def admin_get_skill_tree(admin: dict = _Depends(get_current_admin)):
    sb = db()
    role_id = _tenant_role.get()

    q = sb.table("skill_dimensions").select("*").order("sort_order")
    if role_id:
        q = q.eq("skill_dimensions.role_id", role_id)
    dims = q.execute().data or []

    qi = sb.table("skill_items").select("*").order("sort_order")
    if role_id:
        qi = qi.eq("skill_items.role_id", role_id)
    items = qi.execute().data or []

    qb = sb.table("skill_behaviors").select("*")
    if role_id:
        qb = qb.eq("skill_behaviors.role_id", role_id)
    behaviors = qb.execute().data or []

    maps_result = sb.table("skill_level_map").select("*").execute().data or []
    maps = maps_result

    levels_result = sb.table("level_config").select("*").order("level").execute().data or []
    level_names = {l["level"]: l["short_name"] for l in levels_result}
    # Build sub_level lookup: {sub_level_id: {title, level_id, level_name}}
    sub_level_map = {}
    for l in levels_result:
        for sub in (l.get("sub_levels") or []):
            sub_id = sub.get("id", "")
            if sub_id:
                sub_level_map[sub_id] = {"title": sub.get("title", ""), "level_id": l["level"], "level_name": l["short_name"]}
    actions = sb.table("scenario_actions").select("skill_item_id").not_().is_("skill_item_id", "null").execute().data or []
    action_counts: dict = {}
    for a in actions:
        sid = a.get("skill_item_id")
        if sid:
            action_counts[sid] = action_counts.get(sid, 0) + 1
    # Group behaviors by skill_item_id
    behaviors_by_item: dict = {}
    for b in behaviors:
        bid = b["skill_item_id"]
        if bid not in behaviors_by_item:
            behaviors_by_item[bid] = []
        behaviors_by_item[bid].append(b)
    tree = []
    for d in dims:
        d_items = []
        for item in items:
            if item["dimension_id"] != d["id"]:
                continue
            item_maps = [m for m in maps if m["skill_item_id"] == item["id"]]
            item_behaviors = behaviors_by_item.get(item["id"], [])
            d_items.append({
                **item,
                "level_mappings": [{"level_id": m["level_id"], "is_required": m["is_required"], "level_name": level_names.get(m["level_id"], f"Lv.{m['level_id']}")} for m in item_maps],
                "behaviors": [{
                    "id": b["id"], "behavior": b["behavior"],
                    "level_id": b["level_id"], "sub_level_id": b.get("sub_level_id"),
                    "is_required": b.get("is_required", False),
                    "sub_level_name": sub_level_map.get(b.get("sub_level_id"), {}).get("title", "") if b.get("sub_level_id") else None,
                    "level_name": level_names.get(b["level_id"], f"Lv.{b['level_id']}"),
                } for b in item_behaviors],
                "action_count": action_counts.get(item["id"], 0),
            })
        tree.append({**d, "items": d_items})
    return {"tree": tree}

@app.get("/api/skills/tree")
async def public_skill_tree():
    sb = db()
    dims = sb.table("skill_dimensions").select("*").eq("is_active", True).order("sort_order").execute().data or []
    items = sb.table("skill_items").select("*").eq("is_active", True).order("sort_order").execute().data or []
    tree = []
    for d in dims:
        d_items = [item for item in items if item["dimension_id"] == d["id"]]
        tree.append({**d, "items": d_items})
    return {"tree": tree}


@app.post("/admin/api/skills/demote-dimension")
async def admin_demote_dimension(body: dict, admin: dict = _Depends(get_current_admin)):
    """L1 维度降级为 L2 技能项"""
    sb = db()
    dim_id = body.get("dimension_id")
    target_dim_id = body.get("target_dimension_id")
    if not dim_id or not target_dim_id:
        raise HTTPException(status_code=400, detail="dimension_id 和 target_dimension_id 必填")
    if dim_id == target_dim_id:
        raise HTTPException(status_code=400, detail="不能降级到自身")
    dim = sb.table("skill_dimensions").select("*").eq("id", dim_id).execute().data
    if not dim:
        raise HTTPException(status_code=404, detail="维度不存在")
    dim = dim[0]
    new_item_id = f"SI_{dim_id}"
    existing = sb.table("skill_items").select("id").eq("id", new_item_id).execute().data
    if existing:
        new_item_id = f"SI_{dim_id}_{int(time.time())}"
    sb.table("skill_items").update({"dimension_id": target_dim_id}).eq("dimension_id", dim_id).execute()
    sb.table("skill_items").insert({
        "id": new_item_id, "dimension_id": target_dim_id,
        "name": dim["name"], "description": dim.get("description", ""),
        "sort_order": dim.get("sort_order", 0), "is_active": True,
    }).execute()
    sb.table("skill_dimensions").delete().eq("id", dim_id).execute()
    return {"status": "ok", "new_item_id": new_item_id, "children_moved_to": target_dim_id}

@app.post("/admin/api/skills/promote-item")
async def admin_promote_skill_item(body: dict, admin: dict = _Depends(get_current_admin)):
    """L2 技能项升级为 L1 维度"""
    sb = db()
    item_id = body.get("item_id")
    if not item_id:
        raise HTTPException(status_code=400, detail="item_id 必填")
    item = sb.table("skill_items").select("*").eq("id", item_id).execute().data
    if not item:
        raise HTTPException(status_code=404, detail="技能项不存在")
    item = item[0]
    new_dim_id = item_id.replace("SI", "SD") if item_id.startswith("SI") else f"SD_{item_id}"
    existing = sb.table("skill_dimensions").select("id").eq("id", new_dim_id).execute().data
    if existing:
        new_dim_id = f"SD_{item_id}_{int(time.time())}"
    sb.table("skill_dimensions").insert({
        "id": new_dim_id, "name": item["name"],
        "description": item.get("description", ""), "icon": "",
        "sort_order": item.get("sort_order", 0), "is_active": True,
    }).execute()
    sb.table("skill_items").delete().eq("id", item_id).execute()
    return {"status": "ok", "new_dimension_id": new_dim_id}

@app.put("/admin/api/skills/move-item")
async def admin_move_skill_item(body: dict, admin: dict = _Depends(get_current_admin)):
    """将 L2 技能项从一个维度移到另一个维度"""
    sb = db()
    item_id = body.get("item_id")
    target_dim_id = body.get("target_dimension_id")
    if not item_id or not target_dim_id:
        raise HTTPException(status_code=400, detail="item_id 和 target_dimension_id 必填")
    sb.table("skill_items").update({"dimension_id": target_dim_id}).eq("id", item_id).execute()
    return {"status": "ok"}


# ── L3 关键行为 CRUD ──

@app.get("/admin/api/skill-behaviors")
async def admin_list_behaviors(skill_item_id: str, admin: dict = _Depends(get_current_admin)):
    sb = db()
    q = _apply_role_filter(sb.table("skill_behaviors").select("*"), "skill_behaviors").eq("skill_item_id", skill_item_id).order("level_id").order("sort_order")
    return {"behaviors": q.execute().data or []}

@app.post("/admin/api/skill-behaviors")
async def admin_create_behavior(body: dict, admin: dict = _Depends(get_current_admin)):
    sb = db()
    data = {
        "skill_item_id": body.get("skill_item_id"),
        "behavior": body.get("behavior", ""),
        "level_id": body.get("level_id"),
        "sub_level_id": body.get("sub_level_id"),
        "is_required": body.get("is_required", False),
        "sort_order": body.get("sort_order", 0),
    }
    if not data["skill_item_id"] or not data["behavior"]:
        raise HTTPException(status_code=400, detail="skill_item_id 和 behavior 必填")
    result = sb.table("skill_behaviors").insert(data).execute()
    return {"behavior": result.data[0] if result.data else None}

@app.put("/admin/api/skill-behaviors/{behavior_id}")
async def admin_update_behavior(behavior_id: int, body: dict, admin: dict = _Depends(get_current_admin)):
    sb = db()
    allowed = {"behavior", "level_id", "sub_level_id", "is_required", "sort_order"}
    data = {k: v for k, v in body.items() if k in allowed}
    sb.table("skill_behaviors").update(data).eq("id", behavior_id).execute()
    return {"status": "ok"}

@app.delete("/admin/api/skill-behaviors/{behavior_id}")
async def admin_delete_behavior(behavior_id: int, admin: dict = _Depends(get_current_admin)):
    sb = db()
    sb.table("skill_behaviors").delete().eq("id", behavior_id).execute()
    return {"status": "ok"}

@app.put("/admin/api/skill-behaviors/batch")
async def admin_batch_update_behaviors(body: dict, admin: dict = _Depends(get_current_admin)):
    """批量更新行为（用于拖拽排序等场景）"""
    sb = db()
    behaviors = body.get("behaviors", [])
    for b in behaviors:
        if b.get("id"):
            update_data = {}
            if "level_id" in b: update_data["level_id"] = b["level_id"]
            if "sub_level_id" in b: update_data["sub_level_id"] = b["sub_level_id"]
            if "is_required" in b: update_data["is_required"] = b["is_required"]
            if "sort_order" in b: update_data["sort_order"] = b["sort_order"]
            if "behavior" in b: update_data["behavior"] = b["behavior"]
            if update_data:
                sb.table("skill_behaviors").update(update_data).eq("id", b["id"]).execute()
    return {"status": "ok"}


# ═══════════ 技能全局池 API ═══════════

@app.get("/admin/api/scenarios/{node_id}/skills")
async def admin_get_scenario_skills(node_id: str, admin: dict = _Depends(get_current_admin)):
    """获取场景引用的技能列表"""
    sb = db()
    result = sb.rpc("get_scenario_skills", {"scenario_id_param": node_id}).execute()
    return {"skills": result.data or []}


@app.post("/admin/api/scenarios/{node_id}/skills")
async def admin_set_scenario_skills(node_id: str, body: dict, admin: dict = _Depends(get_current_admin)):
    """批量设置场景技能引用（先删后插）"""
    sb = db()
    skills = body.get("skills", [])
    # 删除旧引用
    sb.table("scenario_skills").delete().eq("scenario_id", node_id).execute()
    # 插入新引用
    for s in skills:
        sb.table("scenario_skills").insert({
            "scenario_id": node_id,
            "skill_id": s["skill_id"],
            "context": s.get("context", "primary"),
        }).execute()
    return {"ok": True, "count": len(skills)}


@app.get("/admin/api/skills")
async def admin_list_skills(search: str = "", limit: int = 50, offset: int = 0,
                            admin: dict = _Depends(get_current_admin)):
    """获取全局技能池"""
    sb = db()
    result = sb.rpc("get_global_skill_pool", {
        "search_query": search, "match_limit": limit, "offset_val": offset,
    }).execute()
    return {"skills": result.data or []}


@app.get("/admin/api/skills/{skill_id}/scenarios")
async def admin_get_skill_scenarios(skill_id: str, admin: dict = _Depends(get_current_admin)):
    """获取引用该技能的所有场景"""
    sb = db()
    result = sb.rpc("get_skill_scenarios", {"skill_id_param": skill_id}).execute()
    return {"scenarios": result.data or []}


# ═══════════ 标签层次 API（Phase 3）════════════

@app.get("/api/tags/hierarchy")
async def get_tag_hierarchy(dimension: str | None = None):
    """获取标签树结构（前端技能层级展示）"""
    sb = db()
    query = sb.table("tags").select("id", "label", "dimension", "description", "parent_id", "granularity", "is_published").eq("is_published", True)
    if dimension:
        query = query.eq("dimension", dimension)
    rows = query.order("sort_order").execute().data

    # 构建树
    nodes = {r["id"]: {**r} for r in rows}
    roots = []
    for r in rows:
        node = nodes[r["id"]]
        parent_id = node.pop("parent_id")
        node["children"] = []
        if parent_id and parent_id in nodes:
            nodes[parent_id]["children"].append(node)
        elif not parent_id:
            roots.append(node)

    return {"hierarchy": roots}


@app.put("/admin/api/tags/{tag_id}/hierarchy")
async def admin_update_tag_hierarchy(tag_id: str, body: dict, admin: dict = _Depends(get_current_admin)):
    """更新标签层级关系（设置 parent_id 和 granularity）"""
    sb = db()
    update_data = {}
    if "parentId" in body:
        update_data["parent_id"] = body["parentId"]
    if "granularity" in body:
        update_data["granularity"] = body["granularity"]
    if not update_data:
        raise HTTPException(status_code=400, detail="无更新字段")
    result = sb.table("tags").update(update_data).eq("id", tag_id).execute()
    admin_crud.log_audit(sb, admin["id"], "tags", tag_id, "update", new_values=update_data)
    return {"ok": True, **(result.data[0] if result.data else {})}


# ═══════════ 资源标签智能推荐（Phase 4）════════════

def _suggest_tags_for_resource(title: str, content: str) -> list[dict]:
    """分析资源内容，推荐相关标签（静默失败）"""
    text = f"{title} {content}".strip()
    if len(text) < 10:
        return []
    try:
        from app.core.vector.search import search_tags
        candidates = search_tags(text, 8)
        return [
            {"id": c.id, "label": c.label, "dimension": c.dimension, "similarity": round(c.similarity, 3)}
            for c in candidates if c.dimension in ("skill", "painpoint")
        ][:6]
    except Exception:
        return []

@app.post("/admin/api/resources/suggest-tags")
async def admin_suggest_tags(body: dict, admin: dict = _Depends(get_current_admin),
                               _=require_feature("ai.tag_suggestion")):
    """AI 分析资源内容，推荐相关标签"""
    title = body.get("title", "")
    content = body.get("content", "")
    text = f"{title} {content}".strip()
    if len(text) < 10:
        raise HTTPException(status_code=400, detail="标题或内容太短")

    try:
        from app.core.vector.search import search_tags
        candidates = await asyncio.to_thread(search_tags, text, 8)
        if not candidates:
            return {"suggested_tags": []}

        # 过滤：只返回 skill 和 painpoint 标签
        relevant = [c for c in candidates if c.dimension in ("skill", "painpoint")]
        result = [
            {"id": c.id, "label": c.label, "dimension": c.dimension, "similarity": round(c.similarity, 3)}
            for c in relevant[:6]
        ]
        return {"suggested_tags": result}
    except Exception as e:
        logger.warning(f"Tag suggestion failed: {e}")
        return {"suggested_tags": []}

# ═════════════ 情报采集系统 API ═════════════
from app.core.intelligence import (
    run_collection, test_source, promote_to_resource,
    generate_resource_preview,
    classify_item, CATEGORY_LABELS, CATEGORIES,
)


# ── 情报源 CRUD ──
@app.get("/admin/api/intelligence/sources")
async def admin_list_sources(admin: dict = _Depends(get_current_admin),
                               _=require_feature("content.intelligence")):
    """获取情报源列表"""
    sb = db()
    rows = sb.table("intelligence_sources").select("*").order("created_at", desc=True).execute().data or []
    return {"items": [_row_to_camel(r) for r in rows], "total": len(rows)}


@app.post("/admin/api/intelligence/sources")
async def admin_create_source(body: dict, admin: dict = _Depends(get_current_admin),
                                _=require_feature("content.intelligence")):
    """创建情报源"""
    data = _body_to_snake(body)
    result = admin_crud.admin_create("intelligence_sources", data, admin["id"])
    return result


@app.put("/admin/api/intelligence/sources/{source_id}")
async def admin_update_source(source_id: str, body: dict, admin: dict = _Depends(get_current_admin),
                                _=require_feature("content.intelligence")):
    """更新情报源"""
    data = _body_to_snake(body)
    result = admin_crud.admin_update("intelligence_sources", source_id, data, admin["id"])
    return result


@app.delete("/admin/api/intelligence/sources/{source_id}")
async def admin_delete_source(source_id: str, admin: dict = _Depends(get_current_admin),
                                _=require_feature("content.intelligence")):
    """删除情报源"""
    admin_crud.admin_delete("intelligence_sources", source_id, admin["id"])
    return {"ok": True}


@app.post("/admin/api/intelligence/sources/{source_id}/test")
async def admin_test_source(source_id: str, admin: dict = _Depends(get_current_admin),
                              _=require_feature("content.intelligence")):
    """测试情报源抓取（不写入数据库）"""
    sb = db()
    rows = sb.table("intelligence_sources").select("*").eq("id", source_id).execute().data
    if not rows:
        raise HTTPException(status_code=404, detail="情报源不存在")
    result = await asyncio.to_thread(test_source, rows[0])
    return result


@app.post("/admin/api/intelligence/sources/{source_id}/fetch")
async def admin_fetch_source(source_id: str, admin: dict = _Depends(get_current_admin),
                               _=require_feature("content.intelligence")):
    """手动触发采集"""
    result = await asyncio.to_thread(run_collection, source_id)
    return result


# ── 情报条目管理 ──
@app.get("/admin/api/intelligence/items")
async def admin_list_items(
    source_id: str | None = None,
    ai_category: str | None = None,
    is_starred: str | None = None,
    page: int = 1,
    page_size: int = 50,
    admin: dict = _Depends(get_current_admin),
    _=require_feature("content.intelligence"),
):
    """获取情报条目列表"""
    sb = db()
    query = sb.table("intelligence_items").select("*, intelligence_sources(name)", count="exact")
    if source_id:
        query = query.eq("source_id", source_id)
    if ai_category:
        query = query.eq("ai_category", ai_category)
    if is_starred is not None:
        query = query.eq("is_starred", is_starred.lower() == "true")

    start = (page - 1) * page_size
    end = page * page_size - 1
    result = query.order("created_at", desc=True).range(start, end).execute()

    items = []
    for r in result.data:
        item = _row_to_camel(r)
        # 嵌套源名称
        if r.get("intelligence_sources"):
            item["sourceName"] = r["intelligence_sources"].get("name", "")
        items.append(item)

    return {"items": items, "total": result.count, "page": page, "page_size": page_size}


@app.get("/admin/api/intelligence/items/{item_id}")
async def admin_get_item(item_id: str, admin: dict = _Depends(get_current_admin),
                          _=require_feature("content.intelligence")):
    sb = db()
    row = sb.table("intelligence_items").select("*").eq("id", item_id).execute().data
    if not row:
        raise HTTPException(status_code=404, detail="情报条目不存在")
    return _row_to_camel(row[0])


@app.put("/admin/api/intelligence/items/{item_id}")
async def admin_update_item(item_id: str, body: dict, admin: dict = _Depends(get_current_admin),
                               _=require_feature("content.intelligence")):
    """更新情报条目（标记精选等）"""
    data = _body_to_snake(body)
    result = admin_crud.admin_update("intelligence_items", item_id, data, admin["id"])
    return result


@app.delete("/admin/api/intelligence/items/{item_id}")
async def admin_delete_item(item_id: str, admin: dict = _Depends(get_current_admin),
                               _=require_feature("content.intelligence")):
    """删除情报条目"""
    admin_crud.admin_delete("intelligence_items", item_id, admin["id"])
    return {"ok": True}


@app.post("/admin/api/intelligence/items/{item_id}/classify")
async def admin_reclassify_item(item_id: str, admin: dict = _Depends(get_current_admin),
                                   _=require_feature("content.intelligence.classify")):
    """AI 重新分类"""
    sb = db()
    row = sb.table("intelligence_items").select("*").eq("id", item_id).execute().data
    if not row:
        raise HTTPException(status_code=404, detail="情报条目不存在")

    classification = await asyncio.to_thread(
        classify_item, row[0]["title"], row[0].get("content", "") or ""
    )
    result = sb.table("intelligence_items").update({
        "ai_category": classification["category"],
        "summary": classification["summary"],
        "value_score": classification["value_score"],
        "ai_tags": classification["tags"],
    }).eq("id", item_id).execute()

    return {"ok": True, "classification": classification}


@app.post("/admin/api/intelligence/items/{item_id}/preview")
async def admin_preview_item(item_id: str, body: dict, admin: dict = _Depends(get_current_admin),
                                 _=require_feature("content.intelligence.promote")):
    """AI 预览改写内容（不创建 resource）"""
    sb = db()
    row = sb.table("intelligence_items").select("*").eq("id", item_id).execute().data
    if not row:
        raise HTTPException(status_code=404, detail="情报条目不存在")

    resource_type = body.get("resourceType", "实战案例")
    preview = await asyncio.to_thread(generate_resource_preview, row[0], resource_type)
    return {"ok": True, **preview}


@app.post("/admin/api/intelligence/items/{item_id}/promote")
async def admin_promote_item(item_id: str, body: dict, admin: dict = _Depends(get_current_admin),
                                _=require_feature("content.intelligence.promote")):
    """推送情报为资源草稿（支持用户编辑后的内容）"""
    sb = db()
    row = sb.table("intelligence_items").select("*").eq("id", item_id).execute().data
    if not row:
        raise HTTPException(status_code=404, detail="情报条目不存在")
    if row[0].get("is_promoted"):
        raise HTTPException(status_code=400, detail="该情报已推送为资源")

    resource_type = body.get("resourceType", "实战案例")
    custom_content = body.get("content")  # 用户在预览弹窗中编辑后的内容
    resource = await asyncio.to_thread(promote_to_resource, row[0], resource_type, custom_content)
    if not resource:
        raise HTTPException(status_code=500, detail="推送失败")

    admin_crud.log_audit(sb, admin["id"], "intelligence_items", item_id, "promote",
                         new_values={"resourceType": resource_type})
    return {"ok": True, "resource": _row_to_camel(resource)}


@app.post("/admin/api/intelligence/items/batch-promote")
async def admin_batch_promote(body: dict, admin: dict = _Depends(get_current_admin),
                                _=require_feature("content.intelligence.promote")):
    """批量推送精选情报为资源"""
    resource_type = body.get("resourceType", "实战案例")
    sb = db()

    # 获取所有精选且未推送的条目
    rows = sb.table("intelligence_items").select("*").eq("is_starred", True).eq("is_promoted", False).execute().data
    if not rows:
        return {"ok": True, "promoted": 0, "message": "没有需要推送的精选情报"}

    promoted = 0
    errors = []
    for row in rows:
        resource = await asyncio.to_thread(promote_to_resource, row, resource_type)
        if resource:
            promoted += 1
        else:
            errors.append(row["id"])

    return {"ok": True, "promoted": promoted, "errors": errors}


@app.post("/admin/api/intelligence/items/{item_id}/publish")
async def admin_publish_item(item_id: str, admin: dict = _Depends(get_current_admin)):
    """发布情报（前端可见）"""
    from datetime import datetime, timezone
    sb = db()
    result = sb.table("intelligence_items").update({
        "published": True,
        "published_at": datetime.now(timezone.utc).isoformat(),
    }).eq("id", item_id).execute()
    admin_crud.log_audit(sb, admin["id"], "intelligence_items", item_id, "publish")
    return {"ok": True}


@app.post("/admin/api/intelligence/items/{item_id}/unpublish")
async def admin_unpublish_item(item_id: str, admin: dict = _Depends(get_current_admin)):
    """取消发布情报"""
    sb = admin_table()
    sb.table("intelligence_items").update({"published": False, "published_at": None}).eq("id", item_id).execute()
    admin_crud.log_audit(sb, admin["id"], "intelligence_items", item_id, "unpublish")
    return {"ok": True}


# ── 前端公开 API ──
@app.get("/api/intelligence/trending")
async def get_trending(limit: int = 10):
    """获取最新已发布情报（前端"今日战情"用）"""
    try:
        sb = admin_table()
        rows = (
            sb.table("intelligence_items")
            .select("id, title, summary, ai_category, published_at")
            .eq("published", True)
            .order("published_at", desc=True)
            .limit(limit)
            .execute().data or []
        )
        items = [_row_to_camel(r) for r in rows]
        return {"items": items, "total": len(items)}
    except Exception as e:
        logger.warning(f"获取 trending 失败: {e}")
        return {"items": [], "total": 0}


@app.get("/api/intelligence/stats")
async def get_intelligence_stats():
    """情报系统统计（公开）"""
    try:
        sb = db()
        source_total = sb.table("intelligence_sources").select("id", count="exact").execute().count
        source_active = sb.table("intelligence_sources").select("id", count="exact").eq("active", True).execute().count
        item_total = sb.table("intelligence_items").select("id", count="exact").execute().count
        item_published = sb.table("intelligence_items").select("id", count="exact").eq("published", True).execute().count
        return {
            "sourceTotal": source_total, "sourceActive": source_active,
            "itemTotal": item_total, "itemPublished": item_published,
        }
    except Exception:
        return {"sourceTotal": 0, "sourceActive": 0, "itemTotal": 0, "itemPublished": 0}


# ═══════════ Feature Flag 公开 API ═══════════
@app.get("/api/feature-flags")
async def public_feature_flags():
    """获取当前租户可用的 feature flags（前端初始化时调用）"""
    tenant = _tenant_schema.get("guojiao")
    return get_feature_flags(tenant)


# ── Audit Log ──
@app.get("/admin/api/audit-log")
async def admin_audit_log(table_name: str | None = None, action: str | None = None,
                          page: int = 1, page_size: int = 50,
                          admin: dict = _Depends(get_current_admin)):
    filters = {}
    if table_name:
        filters["table_name"] = table_name
    if action:
        filters["action"] = action
    return admin_crud.admin_list("content_audit_log", filters=filters or None,
                                 order="created_at", page=page, page_size=page_size)


# ═══════════ Feature Flag Admin API ═══════════
@app.get("/admin/api/feature-flags")
async def admin_list_feature_flags(category: str | None = None, admin: dict = _Depends(get_current_admin)):
    """获取所有 feature flags（Admin 管理面板，不受 tier 限制）"""
    all_flags = get_all_flags()
    if category:
        all_flags = {k: v for k, v in all_flags.items() if v.get("category") == category}
    # 按 sort_order 排序
    items = sorted(all_flags.values(), key=lambda x: x.get("sort_order", 0))
    return {"items": items, "total": len(items)}


@app.post("/admin/api/feature-flags")
async def admin_create_feature_flag(body: dict, admin: dict = _Depends(get_current_admin)):
    """创建 feature flag"""
    sb = admin_table()
    data = _body_to_snake(body)
    result = sb.table("feature_flags").insert(data).execute()
    invalidate_cache()
    admin_crud.log_audit(sb, admin["id"], "feature_flags", data.get("id", ""), "create", new_values=data)
    return {"ok": True, **(result.data[0] if result.data else {})}


@app.put("/admin/api/feature-flags/{flag_id}")
async def admin_update_feature_flag(flag_id: str, body: dict, admin: dict = _Depends(get_current_admin)):
    """更新 feature flag"""
    sb = admin_table()
    data = _body_to_snake(body)
    data["updated_at"] = datetime.now(timezone.utc).isoformat() if hasattr(datetime, 'timezone') else datetime.utcnow().isoformat()
    result = sb.table("feature_flags").update(data).eq("id", flag_id).execute()
    invalidate_cache()
    admin_crud.log_audit(sb, admin["id"], "feature_flags", flag_id, "update", new_values=data)
    return {"ok": True, **(result.data[0] if result.data else {})}


@app.post("/admin/api/feature-flags/{flag_id}/toggle")
async def admin_toggle_feature_flag(flag_id: str, admin: dict = _Depends(get_current_admin)):
    """快速切换 feature flag 开关状态"""
    sb = admin_table()
    # 获取当前状态
    row = sb.table("feature_flags").select("is_enabled").eq("id", flag_id).execute().data
    if not row:
        raise HTTPException(status_code=404, detail=f"Flag '{flag_id}' 不存在")
    new_state = not row[0]["is_enabled"]
    sb.table("feature_flags").update({
        "is_enabled": new_state,
        "updated_at": datetime.utcnow().isoformat(),
    }).eq("id", flag_id).execute()
    invalidate_cache()
    admin_crud.log_audit(sb, admin["id"], "feature_flags", flag_id,
                         "toggle", new_values={"is_enabled": new_state})
    return {"ok": True, "flag_id": flag_id, "is_enabled": new_state}


# ── 订阅管理 ──
@app.get("/admin/api/subscription")
async def admin_get_subscription(admin: dict = _Depends(get_current_admin)):
    """查看当前订阅"""
    tenant = _tenant_schema.get("guojiao")
    sub = get_subscription(tenant)
    all_flags = get_all_flags()
    # 附加可选功能列表
    optional_flags = [
        {"id": k, "name": v["name"], "description": v.get("description", ""), "tier": v.get("tier")}
        for k, v in all_flags.items() if v.get("is_optional")
    ]
    return {**_row_to_camel(sub), "optional_features": optional_flags}


@app.put("/admin/api/subscription")
async def admin_update_subscription(body: dict, admin: dict = _Depends(get_current_admin)):
    """修改订阅 tier 和 custom_features"""
    sb = admin_table()
    data = {}
    if "tier" in body:
        data["tier"] = body["tier"]
    if "customFeatures" in body:
        data["custom_features"] = body["customFeatures"]
    if "maxUsers" in body:
        data["max_users"] = body["maxUsers"]
    if not data:
        raise HTTPException(status_code=400, detail="无更新字段")

    tenant = _tenant_schema.get("guojiao")
    # 确保 subscription 记录存在
    existing = sb.table("tenant_subscriptions").select("tenant_id").eq("tenant_id", tenant).execute().data
    if not existing:
        sb.table("tenant_subscriptions").insert({"tenant_id": tenant, "tier": data.get("tier", "pro"), "is_active": True}).execute()
    result = sb.table("tenant_subscriptions").update(data).eq("tenant_id", tenant).execute()
    invalidate_cache()
    admin_crud.log_audit(sb, admin["id"], "tenant_subscriptions", tenant, "update", new_values=data)
    return {"ok": True, **(result.data[0] if result.data else {})}


# ═══════════ 动作体系 API（future.action_system）═══════════

# ── 公共 API（手机端）──
@app.get("/api/scenarios/{l3_node_id}/actions")
async def get_scenario_actions(l3_node_id: str):
    """获取场景下所有已发布的动作"""
    sb = db()
    rows = sb.table("scenario_actions").select("*").eq("scenario_id", l3_node_id).eq("is_published", True).eq("is_active", True).order("sort_order").execute().data or []
    return {"actions": [_row_to_camel(r) for r in rows], "total": len(rows)}


@app.get("/api/actions/{action_id}")
async def get_action_detail(action_id: str):
    """获取单个动作详情"""
    sb = db()
    rows = sb.table("scenario_actions").select("*").eq("id", action_id).eq("is_published", True).execute().data
    if not rows:
        raise HTTPException(status_code=404, detail="动作不存在")
    return _row_to_camel(rows[0])


@app.post("/api/user/actions/{action_id}/complete")
async def complete_action(action_id: str, body: dict):
    """标记动作完成"""
    sb = db()
    user_id = body.get("userId")
    self_rating = body.get("selfRating", 5)
    notes = body.get("notes", "")

    if not user_id:
        raise HTTPException(status_code=400, detail="userId 不能为空")

    # 获取动作信息
    action = sb.table("scenario_actions").select("scenario_id").eq("id", action_id).execute().data
    if not action:
        raise HTTPException(status_code=404, detail="动作不存在")

    try:
        result = sb.table("user_action_progress").upsert({
            "user_id": user_id,
            "action_id": action_id,
            "scenario_id": action[0]["scenario_id"],
            "completed_at": datetime.utcnow().isoformat(),
            "self_rating": self_rating,
            "notes": notes,
        }, on_conflict="user_id,action_id").execute()
        return {"ok": True, **(result.data[0] if result.data else {})}
    except Exception as e:
        logger.error(f"Failed to complete action: {e}")
        raise HTTPException(status_code=500, detail="操作失败")


@app.get("/api/user/{user_id}/profile")
async def get_user_profile(user_id: str):
    """获取用户画像数据（整合动作完成情况 + 等级信息）"""
    sb = db()

    # 用户基础信息
    user = sb.table("users").select("id, nickname, current_level, total_xp, created_at").eq("id", user_id).execute().data
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    user = user[0]

    # 已完成的动作
    progress = sb.table("user_action_progress").select("*").eq("user_id", user_id).not_().is_("completed_at", "null").execute().data or []
    completed_action_ids = {p["action_id"] for p in progress}
    completed_scenario_ids = {p["scenario_id"] for p in progress}

    # 已通过的练习（加权：每次通过计 0.5 个行动）
    practice_sessions = sb.table("practice_sessions").select("id").eq("user_id", user_id).eq("passed", True).execute().data or []
    passed_practice_bonus = len(practice_sessions) // 2  # 2次通过 = 1个行动

    # 所有已发布的动作总数
    all_actions = sb.table("scenario_actions").select("id, scenario_id, value_impact, experience_impact, difficulty").eq("is_published", True).execute().data or []
    total_actions = len(all_actions)

    # ── 一期：纯完成率六维画像 ──
    all_scenario_ids = {a["scenario_id"] for a in all_actions}
    scenario_coverage = len(completed_scenario_ids) / len(all_scenario_ids) * 10 if all_scenario_ids else 0

    high_value_actions = [a for a in all_actions if a["value_impact"] >= 7]
    high_value_completed = sum(1 for a in high_value_actions if a["id"] in completed_action_ids)
    painpoint_analysis = high_value_completed / len(high_value_actions) * 10 if high_value_actions else 0

    completed_actions_data = [a for a in all_actions if a["id"] in completed_action_ids]
    value_mining = sum(a["value_impact"] for a in completed_actions_data) / len(completed_actions_data) if completed_actions_data else 0

    skill_mastery = len(completed_action_ids) / total_actions * 10 if total_actions else 0
    action_completion = len(completed_action_ids) / total_actions * 10 if total_actions else 0

    from datetime import timedelta
    seven_days_ago = (datetime.utcnow() - timedelta(days=7)).isoformat()
    recent_progress = [p for p in progress if p.get("completed_at") and p["completed_at"] > seven_days_ago]
    learning_activity = min(len(recent_progress) / 3 * 10, 10)

    radar_base = {
        "scenarioCoverage": round(scenario_coverage, 1),
        "painpointAnalysis": round(painpoint_analysis, 1),
        "valueMining": round(value_mining, 1),
        "skillMastery": round(skill_mastery, 1),
        "actionCompletion": round(action_completion, 1),
        "learningActivity": round(learning_activity, 1),
    }

    # ── Phase 2: 增强画像（加权评分） ──
    # 维度得分 = 动作完成率 × 0.4 + 指标达成率 × 0.4 + 自评一致性 × 0.2
    radar = radar_base
    try:
        from app.core.feature_flags import is_feature_enabled
        if is_feature_enabled("phase2.enhanced_profile"):
            # 获取用户最近 3 个月的指标数据
            three_months_ago = (datetime.utcnow() - timedelta(days=90)).strftime("%Y-%m-%d")
            metrics = sb.table("business_metrics").select("*").eq("user_id", user_id).gte("period_start", three_months_ago).execute().data or []

            # 指标达成率：各指标最近值 vs 基准
            benchmarks = sb.table("metric_benchmarks").select("*").eq("benchmark_level", "average").execute().data or []
            bench_map = {b["metric_type"]: b["min_value"] for b in benchmarks}

            metric_score = 0
            metric_count = 0
            for m in metrics:
                bt = m["metric_type"]
                if bt in bench_map and bench_map[bt] > 0:
                    ratio = float(m["metric_value"]) / float(bench_map[bt])
                    metric_score += min(ratio, 1.5) / 1.5 * 10  # 超过基准150%封顶
                    metric_count += 1
            avg_metric_score = metric_score / metric_count if metric_count > 0 else 5.0  # 无指标数据时给中等分

            # 自评一致性：self_rating 与动作难度的相关性
            self_ratings = [p.get("self_rating", 5) for p in progress]
            action_difficulties = {"easy": 3, "medium": 5, "hard": 7, "expert": 9}
            action_diff_vals = []
            for p in progress:
                action_id = p["action_id"]
                action = next((a for a in all_actions if a["id"] == action_id), None)
                if action:
                    action_diff_vals.append(action_difficulties.get(action.get("difficulty", "medium"), 5))

            consistency = 5.0  # 默认中等
            if len(self_ratings) == len(action_diff_vals) and len(self_ratings) >= 3:
                # 计算相关系数的简化版：差异越小越一致
                diffs = [abs(sr - dv) for sr, dv in zip(self_ratings, action_diff_vals)]
                avg_diff = sum(diffs) / len(diffs)
                consistency = max(0, 10 - avg_diff * 2)  # 平均差1分=8分, 差2分=6分, 差5分=0分

            # 加权计算各维度
            completion_rate = len(completed_action_ids) / total_actions if total_actions else 0
            radar = {
                "scenarioCoverage": round(completion_rate * 10 * 0.4 + avg_metric_score * 0.4 + consistency * 0.2, 1),
                "painpointAnalysis": round(painpoint_analysis * 0.4 + avg_metric_score * 0.4 + consistency * 0.2, 1),
                "valueMining": round(value_mining * 0.4 + avg_metric_score * 0.4 + consistency * 0.2, 1),
                "skillMastery": round(skill_mastery * 0.4 + avg_metric_score * 0.4 + consistency * 0.2, 1),
                "actionCompletion": round(action_completion * 0.4 + avg_metric_score * 0.4 + consistency * 0.2, 1),
                "learningActivity": round(learning_activity * 0.4 + avg_metric_score * 0.4 + consistency * 0.2, 1),
            }
    except Exception:
        pass  # flag 检查失败时降级为一期算法

    # 推荐学习路径（基于薄弱环节）
    learning_path = []
    for sid in sorted(all_scenario_ids):
        scenario_actions = [a for a in all_actions if a["scenario_id"] == sid]
        completed_in_scenario = sum(1 for a in scenario_actions if a["id"] in completed_action_ids)
        if completed_in_scenario < len(scenario_actions):
            next_action = next((a for a in scenario_actions if a["id"] not in completed_action_ids), None)
            if next_action:
                resources = sb.table("resource_tags").select("resource_id").eq("tag_id", sid).execute().data or []
                resource_ids = [r["resource_id"] for r in resources[:3]]
                learning_path.append({
                    "scenarioId": sid,
                    "completedCount": completed_in_scenario,
                    "totalCount": len(scenario_actions),
                    "nextActionId": next_action["id"],
                    "resourceIds": resource_ids,
                })

    return {
        "user": _row_to_camel(user),
        "stats": {
            "completedActions": len(completed_action_ids) + passed_practice_bonus,
            "totalActions": total_actions,
            "completedScenarios": len(completed_scenario_ids),
            "totalScenarios": len(all_scenario_ids),
        },
        "radar": radar,
        "radarBase": radar_base,  # 一期原始分数（用于对比）
        "learningPath": learning_path,
    }


# ── Admin API ──
@app.get("/admin/api/actions")
async def admin_list_actions(scenario_id: str | None = None, difficulty: str | None = None,
                              page: int = 1, page_size: int = 50,
                              admin: dict = _Depends(get_current_admin),
                              _=require_feature("future.action_system")):
    """获取动作列表（支持场景和难度过滤）"""
    sb = db()
    # PostgREST embedded resource 的字段会覆盖父表同名字段
    # 用显式列选择 + 单独查询 L3 标题，避免 id/title 被覆盖
    query = sb.table("scenario_actions").select("id, scenario_id, title, description, tags, steps, value_impact, experience_impact, difficulty, sort_order, is_active, is_published, created_at, updated_at, skill_item_id", count="exact")
    if scenario_id:
        query = query.eq("scenario_id", scenario_id)
    if difficulty:
        query = query.eq("difficulty", difficulty)
    # 角色过滤：通过关联场景的 role_id
    role_id = _tenant_role.get()
    if role_id and not scenario_id:
        role_scenarios = sb.table("scenario_tree_nodes").select("id").eq("role_id", role_id).execute().data or []
        role_scenario_ids = [s["id"] for s in role_scenarios]
        if role_scenario_ids:
            query = query.in_("scenario_id", role_scenario_ids)
        else:
            return {"items": [], "total": 0, "page": page, "page_size": page_size}

    start = (page - 1) * page_size
    end = page * page_size - 1
    result = query.order("scenario_actions.scenario_id, scenario_actions.sort_order").range(start, end).execute()

    # 批量查询 L3 标题
    scenario_ids = list({r["scenario_id"] for r in (result.data or [])})
    scenario_titles = {}
    if scenario_ids:
        rows = sb.table("scenario_tree_nodes").select("id, title").in_("id", scenario_ids).execute().data or []
        scenario_titles = {r["id"]: r["title"] for r in rows}

    items = []
    for r in (result.data or []):
        item = _row_to_camel(r)
        item["scenarioTitle"] = scenario_titles.get(r["scenario_id"], "")
        items.append(item)

    return {"items": items, "total": result.count, "page": page, "page_size": page_size}


@app.get("/admin/api/actions/{action_id}")
async def admin_get_action(action_id: str, admin: dict = _Depends(get_current_admin),
                            _=require_feature("future.action_system")):
    """获取动作详情"""
    row = admin_crud.admin_get("scenario_actions", action_id)
    if not row:
        raise HTTPException(status_code=404, detail="动作不存在")
    # 补充场景名称
    sb = db()
    scenario = sb.table("scenario_tree_nodes").select("title").eq("id", row["scenario_id"]).execute().data
    if scenario:
        row["scenario_title"] = scenario[0]["title"]
    return _row_to_camel(row)


@app.post("/admin/api/actions")
async def admin_create_action(body: dict, admin: dict = _Depends(get_current_admin),
                               _=require_feature("future.action_system")):
    """创建动作"""
    data = _body_to_snake(body)
    result = admin_crud.admin_create("scenario_actions", data, admin["id"])
    return result


@app.put("/admin/api/actions/{action_id}")
async def admin_update_action(action_id: str, body: dict, admin: dict = _Depends(get_current_admin),
                               _=require_feature("future.action_system")):
    """更新动作（含九宫格坐标拖拽）"""
    data = _body_to_snake(body)
    result = admin_crud.admin_update("scenario_actions", action_id, data, admin["id"])
    return result


@app.delete("/admin/api/actions/{action_id}")
async def admin_delete_action(action_id: str, admin: dict = _Depends(get_current_admin),
                               _=require_feature("future.action_system")):
    """删除动作"""
    admin_crud.admin_delete("scenario_actions", action_id, admin["id"])
    return {"ok": True}


@app.post("/admin/api/actions/batch")
async def admin_batch_actions(body: dict, admin: dict = _Depends(get_current_admin),
                               _=require_feature("future.action_system")):
    """批量创建动作（按 L3 节点）"""
    sb = db()
    actions = body.get("actions", [])
    if not actions:
        raise HTTPException(status_code=400, detail="actions 不能为空")

    results = []
    for action in actions:
        data = _body_to_snake(action)
        data.setdefault("is_active", True)
        data.setdefault("is_published", False)
        try:
            result = sb.table("scenario_actions").insert(data).execute()
            action_type = "create"
        except Exception as e:
            if "duplicate key" in str(e):
                result = sb.table("scenario_actions").update(data).eq("id", data.get("id")).execute()
                action_type = "update"
            else:
                raise
        if result.data:
            results.append(result.data[0])
            admin_crud.log_audit(sb, admin["id"], "scenario_actions", str(data.get("id", "")), action_type, new_values=data)

    return {"created": len(results), "items": results}


@app.post("/admin/api/actions/{action_id}/publish")
async def admin_publish_action(action_id: str, admin: dict = _Depends(get_current_admin),
                                _=require_feature("future.action_system")):
    return admin_crud.admin_toggle_publish("scenario_actions", action_id, admin["id"], True)


@app.post("/admin/api/actions/{action_id}/unpublish")
async def admin_unpublish_action(action_id: str, admin: dict = _Depends(get_current_admin),
                                  _=require_feature("future.action_system")):
    return admin_crud.admin_toggle_publish("scenario_actions", action_id, admin["id"], False)


@app.get("/admin/api/user-action-progress")
async def admin_list_user_progress(user_id: str | None = None, scenario_id: str | None = None,
                                    page: int = 1, page_size: int = 50,
                                    admin: dict = _Depends(get_current_admin),
                                    _=require_feature("future.talent_profile")):
    """查看用户动作完成情况"""
    sb = db()
    query = sb.table("user_action_progress").select("*, users(nickname), scenario_actions(title, value_impact, experience_impact)", count="exact")
    if user_id:
        query = query.eq("user_id", user_id)
    if scenario_id:
        query = query.eq("scenario_id", scenario_id)

    start = (page - 1) * page_size
    end = page * page_size - 1
    result = query.order("user_action_progress.completed_at", desc=True).range(start, end).execute()

    items = []
    for r in result.data:
        item = _row_to_camel(r)
        if r.get("users"):
            item["userNickname"] = r["users"].get("nickname", "")
        if r.get("scenario_actions"):
            item["actionTitle"] = r["scenario_actions"].get("title", "")
            item["valueImpact"] = r["scenario_actions"].get("value_impact")
            item["experienceImpact"] = r["scenario_actions"].get("experience_impact")
        items.append(item)

    return {"items": items, "total": result.count, "page": page, "page_size": page_size}


@app.get("/admin/api/users")
async def admin_list_users(page: int = 1, page_size: int = 50,
                            admin: dict = _Depends(get_current_admin),
                            _=require_feature("future.talent_profile")):
    """获取用户列表（人才档案用）"""
    sb = db()
    query = sb.table("users").select("*", count="exact")
    start = (page - 1) * page_size
    end = page * page_size - 1
    result = query.order("created_at", desc=True).range(start, end).execute()

    # 附加动作完成统计
    items = []
    for r in result.data:
        item = _row_to_camel(r)
        progress_count = sb.table("user_action_progress").select("id", count="exact").eq("user_id", r["id"]).not_().is_("completed_at", "null").execute().count
        item["completedActions"] = progress_count
        items.append(item)

    return {"items": items, "total": result.count, "page": page, "page_size": page_size}


@app.put("/admin/api/users/{user_id}")
async def admin_update_user(user_id: str, body: dict, admin: dict = _Depends(get_current_admin)):
    """更新用户信息"""
    sb = db()
    data = {}
    if "nickname" in body:
        data["nickname"] = body["nickname"]
    if "role_id" in body:
        data["role_id"] = body["role_id"]
    if not data:
        raise HTTPException(status_code=400, detail="没有要更新的字段")
    sb.table("users").update(data).eq("id", user_id).execute()
    return {"ok": True}


@app.delete("/admin/api/users/{user_id}")
async def admin_delete_user(user_id: str, admin: dict = _Depends(get_current_admin)):
    """删除用户"""
    sb = db()
    sb.table("team_assignments").delete().eq("user_id", user_id).execute()
    sb.table("users").delete().eq("id", user_id).execute()
    return {"ok": True}


@app.get("/admin/api/users/{user_id}/profile")
async def admin_get_user_profile(user_id: str, admin: dict = _Depends(get_current_admin),
                                  _=require_feature("future.talent_profile")):
    """获取用户完整画像（Admin 版，含更多详情）"""
    return await get_user_profile(user_id)


# ── Phase 1 补齐：数据完整性 API ──

@app.get("/admin/api/data-health")
async def admin_data_health(admin: dict = _Depends(get_current_admin)):
    """数据完整性仪表盘 — 返回各维度覆盖率"""
    sb = db()

    # L1/L2/L3 各层级统计
    l1_nodes = sb.table("scenario_tree_nodes").select("id").eq("level", 1).eq("is_active", True).execute().data
    l2_nodes = sb.table("scenario_tree_nodes").select("id, parent_id").eq("level", 2).eq("is_active", True).execute().data
    l3_nodes = sb.table("scenario_tree_nodes").select("id, parent_id, title").eq("level", 3).eq("is_active", True).execute().data
    l3_total = len(l3_nodes)

    # 有动作的 L3
    actions = sb.table("scenario_actions").select("scenario_id", count="exact").eq("is_active", True).execute().data
    l3_with_actions = set(a["scenario_id"] for a in actions) if actions else set()
    l3_with_actions_count = len([n for n in l3_nodes if n["id"] in l3_with_actions])

    # 有标签的 L3
    tag_links = sb.table("resource_tags").select("resource_id").execute().data
    tagged_ids = set(t["resource_id"] for t in tag_links) if tag_links else set()
    # resource_tags 关联的是 resources 表，检查 L3 场景的 tags 字段
    l3_with_tags = 0
    for n in l3_nodes:
        if n.get("tags") and len(n["tags"]) > 0:
            l3_with_tags += 1

    # 有资源的动作
    actions_total = len(actions)
    # 检查 actions 表中是否有 resource_ids 字段
    actions_with_resources = 0
    if actions:
        for a in actions:
            if a.get("resource_ids") and len(a["resource_ids"]) > 0:
                actions_with_resources += 1

    # 缺失动作的 L3 场景
    missing_actions = [n for n in l3_nodes if n["id"] not in l3_with_actions]
    missing_actions.sort(key=lambda x: x.get("title", ""))

    # 按 L2 分组统计
    l2_map = {n["id"]: n for n in l2_nodes}
    by_l2 = {}
    for l2 in l2_nodes:
        l2_id = l2["id"]
        l2_l3s = [n for n in l3_nodes if n.get("parent_id") == l2_id]
        l2_l3_with_actions = [n for n in l2_l3s if n["id"] in l3_with_actions]
        by_l2[l2_id] = {
            "l2_title": l2.get("title", l2_id),
            "l3_count": len(l2_l3s),
            "l3_with_actions": len(l2_l3_with_actions),
            "action_coverage": round(len(l2_l3_with_actions) / max(len(l2_l3s), 1), 2),
        }

    return {
        "l1_count": len(l1_nodes),
        "l2_count": len(l2_nodes),
        "l3_total": l3_total,
        "l3_with_actions": l3_with_actions_count,
        "l3_with_tags": l3_with_tags,
        "actions_total": actions_total,
        "actions_with_resources": actions_with_resources,
        "coverage": {
            "l3_actions": round(l3_with_actions_count / max(l3_total, 1), 2),
            "l3_tags": round(l3_with_tags / max(l3_total, 1), 2),
            "actions_resources": round(actions_with_resources / max(actions_total, 1), 2),
        },
        "missing_actions": [{"id": m["id"], "title": m.get("title", "")} for m in missing_actions[:50]],
        "by_l2": by_l2,
    }


@app.get("/admin/api/coverage-matrix")
async def admin_coverage_matrix(admin: dict = _Depends(get_current_admin)):
    """场景-动作覆盖率矩阵 — 按 L2 分组展示 L3 场景的动作覆盖情况"""
    sb = db()

    l2_nodes = sb.table("scenario_tree_nodes").select("id, title, parent_id").eq("level", 2).eq("is_active", True).order("sort_order").execute().data
    l3_nodes = sb.table("scenario_tree_nodes").select("id, title, parent_id").eq("level", 3).eq("is_active", True).order("sort_order").execute().data
    actions = sb.table("scenario_actions").select("scenario_id, is_published").eq("is_active", True).execute().data

    # 按 scenario_id 统计动作数
    action_count_map = {}
    published_count_map = {}
    for a in actions:
        sid = a["scenario_id"]
        action_count_map[sid] = action_count_map.get(sid, 0) + 1
        if a.get("is_published"):
            published_count_map[sid] = published_count_map.get(sid, 0) + 1

    # 按 L2 分组
    l1_nodes = sb.table("scenario_tree_nodes").select("id, title").eq("level", 1).eq("is_active", True).order("sort_order").execute().data
    l1_map = {n["id"]: n for n in l1_nodes}

    matrix = []
    for l2 in l2_nodes:
        l2_l3s = [n for n in l3_nodes if n.get("parent_id") == l2["id"]]
        l2_rows = []
        l2_with_actions = 0
        for l3 in l2_l3s:
            action_count = action_count_map.get(l3["id"], 0)
            published_count = published_count_map.get(l3["id"], 0)
            if action_count > 0:
                l2_with_actions += 1
            l2_rows.append({
                "l3_id": l3["id"],
                "l3_title": l3.get("title", ""),
                "action_count": action_count,
                "published_count": published_count,
                "has_actions": action_count > 0,
            })

        l1_title = ""
        if l2.get("parent_id") and l2["parent_id"] in l1_map:
            l1_title = l1_map[l2["parent_id"]].get("title", "")

        matrix.append({
            "l1_title": l1_title,
            "l2_id": l2["id"],
            "l2_title": l2.get("title", ""),
            "l3_total": len(l2_l3s),
            "l3_with_actions": l2_with_actions,
            "coverage": round(l2_with_actions / max(len(l2_l3s), 1), 2),
            "l3_items": l2_rows,
        })

    return {"matrix": matrix}


@app.post("/admin/api/actions/import")
async def admin_import_actions(file: UploadFile = File(...), admin: dict = _Depends(get_current_admin)):
    """Excel 批量导入动作"""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="服务端未安装 openpyxl，请执行 pip install openpyxl")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 格式")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 解析失败: {str(e)}")

    if "actions" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Excel 中未找到 'actions' 工作表")

    ws = wb["actions"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头

    sb = db()
    created = 0
    updated = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue  # 跳过空行

        try:
            scenario_id = str(row[0]).strip()
            title = str(row[1]).strip() if len(row) > 1 else ""
            description = str(row[2]).strip() if len(row) > 2 else ""

            # 验证场景存在
            scenario = sb.table("scenario_tree_nodes").select("id").eq("id", scenario_id).execute().data
            if not scenario:
                errors.append({"row": i, "error": f"场景 {scenario_id} 不存在"})
                continue

            # 解析 steps（用 | 分隔）
            steps_raw = str(row[3]).strip() if len(row) > 3 else ""
            steps = [s.strip() for s in steps_raw.split("|") if s.strip()] if steps_raw else []

            # 解析 tags（用 | 分隔）
            tags_raw = str(row[4]).strip() if len(row) > 4 else ""
            tags = [t.strip() for t in tags_raw.split("|") if t.strip()] if tags_raw else []

            difficulty = str(row[5]).strip() if len(row) > 5 else "medium"
            if difficulty not in ("easy", "medium", "hard", "expert"):
                difficulty = "medium"

            value_impact = int(row[6]) if len(row) > 6 and row[6] else 5
            experience_impact = int(row[7]) if len(row) > 7 and row[7] else 5
            value_impact = max(1, min(10, value_impact))
            experience_impact = max(1, min(10, experience_impact))

            data = {
                "scenario_id": scenario_id,
                "title": title,
                "description": description,
                "steps": steps,
                "tags": tags,
                "difficulty": difficulty,
                "value_impact": value_impact,
                "experience_impact": experience_impact,
                "is_active": True,
                "is_published": False,
            }

            try:
                sb.table("scenario_actions").insert(data).execute()
                created += 1
            except Exception:
                # 更新已有记录
                sb.table("scenario_actions").update(data).eq("scenario_id", scenario_id).eq("title", title).execute()
                updated += 1

        except Exception as e:
            errors.append({"row": i, "error": str(e)})

    return {"created": created, "updated": updated, "errors": errors}


@app.get("/admin/api/actions/template")
async def admin_actions_template(admin: dict = _Depends(get_current_admin)):
    """下载动作导入 Excel 模板"""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="服务端未安装 openpyxl")

    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "actions"

    headers = ["scenario_id", "title", "description", "steps (|分隔)", "tags (|分隔)", "difficulty", "value_impact (1-10)", "experience_impact (1-10)"]
    ws.append(headers)

    # 示例行
    ws.append(["S_L3_001", "进店观察", "进店后先观察店内陈列和客流", "观察陈列|观察客流|记录竞品|拍照记录", "S_陈列|K_观察力", "easy", 4, 6])
    ws.append(["S_L3_001", "老板沟通", "与烟酒店老板建立初步关系", "自我介绍|递名片|询问经营情况|了解进货渠道", "S_客情|K_沟通", "medium", 7, 5])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=actions_import_template.xlsx"},
    )


# ═══════════ Phase 1.5: 数据采信层 API ═══════════

@app.post("/admin/api/metrics")
async def admin_create_metric(body: dict, admin: dict = _Depends(get_current_admin)):
    """提交指标数据（手动录入单条）"""
    sb = db()
    required = ["user_id", "metric_type", "metric_value", "period_start", "period_end"]
    for k in required:
        if k not in body or not body[k]:
            raise HTTPException(status_code=400, detail=f"缺少必填字段: {k}")
    try:
        data = {
            "user_id": body["user_id"],
            "metric_type": body["metric_type"],
            "metric_value": float(body["metric_value"]),
            "period_start": body["period_start"],
            "period_end": body["period_end"],
            "data_source": body.get("data_source", "manual"),
            "notes": body.get("notes", ""),
        }
        result = sb.table("business_metrics").insert(data).execute()
        return {"ok": True, "id": result.data[0]["id"]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/admin/api/metrics/import")
async def admin_import_metrics(file: UploadFile = File(...), admin: dict = _Depends(get_current_admin)):
    """Excel 批量导入指标数据"""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="服务端未安装 openpyxl")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 格式")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 解析失败: {str(e)}")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    sb = db()
    created = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue
        try:
            data = {
                "user_id": str(row[0]).strip(),
                "metric_type": str(row[1]).strip(),
                "metric_value": float(row[2]) if row[2] else 0,
                "period_start": str(row[3]).strip(),
                "period_end": str(row[4]).strip(),
                "data_source": str(row[5]).strip() if len(row) > 5 and row[5] else "manual",
                "notes": str(row[6]).strip() if len(row) > 6 and row[6] else "",
            }
            sb.table("business_metrics").insert(data).execute()
            created += 1
        except Exception as e:
            errors.append({"row": i, "error": str(e)})

    return {"created": created, "errors": errors}


@app.get("/admin/api/metrics/template")
async def admin_metrics_template(admin: dict = _Depends(get_current_admin)):
    """下载指标导入 Excel 模板"""
    from fastapi.responses import StreamingResponse
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "metrics"
    ws.append(["user_id", "metric_type", "metric_value", "period_start", "period_end", "data_source", "notes"])
    ws.append(["00000000-0000-0000-0000-000000000001", "visit_count", 65, "2026-03-01", "2026-03-31", "manual", "3月拜访量"])
    ws.append(["00000000-0000-0000-0000-000000000001", "deal_amount", 150000, "2026-03-01", "2026-03-31", "manual", "3月成交额"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=metrics_import_template.xlsx"},
    )


@app.get("/admin/api/metrics")
async def admin_list_metrics(
    user_id: str | None = None,
    metric_type: str | None = None,
    period_start: str | None = None,
    period_end: str | None = None,
    page: int = 1,
    page_size: int = 50,
    admin: dict = _Depends(get_current_admin),
):
    """查询指标数据（支持过滤和分页）"""
    sb = db()
    query = sb.table("business_metrics").select("*")
    if user_id:
        query = query.eq("user_id", user_id)
    if metric_type:
        query = query.eq("metric_type", metric_type)
    if period_start:
        query = query.gte("period_start", period_start)
    if period_end:
        query = query.lte("period_end", period_end)
    query = query.order("period_start", desc=True).range((page - 1) * page_size, page * page_size - 1)
    result = query.execute()
    return {"metrics": result.data, "page": page, "page_size": page_size}


@app.get("/api/user/{user_id}/metrics")
async def get_user_metrics(user_id: str, metric_type: str | None = None, months: int = 6):
    """获取个人指标趋势（最近 N 个月）"""
    from datetime import timedelta
    sb = db()

    cutoff = (datetime.utcnow() - timedelta(days=30 * months)).strftime("%Y-%m-%d")
    query = sb.table("business_metrics").select("*").eq("user_id", user_id).gte("period_start", cutoff)
    if metric_type:
        query = query.eq("metric_type", metric_type)
    metrics = query.order("period_start").execute().data or []

    # 按月聚合
    monthly = {}
    for m in metrics:
        month = m["period_start"][:7]  # "2026-03"
        if month not in monthly:
            monthly[month] = {}
        mt = m["metric_type"]
        monthly[month][mt] = m["metric_value"]

    # 按指标类型分组时间序列
    series = {}
    for month, vals in monthly.items():
        for mt, val in vals.items():
            if mt not in series:
                series[mt] = []
            series[mt].append({"period": month, "value": float(val)})

    return {"metrics": metrics, "monthly": monthly, "series": series}


@app.get("/admin/api/team/list")
async def admin_list_teams(admin: dict = _Depends(get_current_admin)):
    """获取所有团队列表（树形结构）"""
    sb = db()
    teams = sb.table("team_assignments").select("id, user_id, team_name, team_level, parent_team_id, is_leader").order("team_level").execute().data or []

    # 按 team_name 聚合
    team_map = {}
    for t in teams:
        name = t["team_name"]
        if name not in team_map:
            team_map[name] = {"team_name": name, "team_level": t["team_level"], "members": [], "leader": None}
        user = sb.table("users").select("id, nickname").eq("id", t["user_id"]).execute().data
        member_info = {"user_id": t["user_id"], "is_leader": t["is_leader"]}
        if user:
            member_info["nickname"] = user[0].get("nickname", "")
        team_map[name]["members"].append(member_info)
        if t["is_leader"]:
            team_map[name]["leader"] = t["user_id"]

    return {"teams": list(team_map.values())}


@app.get("/admin/api/team/tree")
async def admin_team_tree(admin: dict = _Depends(get_current_admin)):
    """获取团队树形结构（用于下拉选择）"""
    sb = db()
    assignments = sb.table("team_assignments").select(
        "id, user_id, team_name, team_level, parent_team_id, is_leader"
    ).order("team_level").execute().data or []

    # 按 team_name 聚合，同时收集所有 team_name
    team_map = {}
    for t in assignments:
        name = t["team_name"]
        if name not in team_map:
            team_map[name] = {
                "id": t["id"],
                "name": name,
                "level": t["team_level"],
                "parent_id": t.get("parent_team_id"),
                "parent_name": None,
                "member_count": 0,
                "leader": None,
            }
        team_map[name]["member_count"] = team_map[name].get("member_count", 0) + 1
        if t["is_leader"] and t["user_id"]:
            user = sb.table("users").select("id, nickname").eq("id", t["user_id"]).execute().data
            if user:
                team_map[name]["leader"] = {"id": user[0]["id"], "nickname": user[0].get("nickname", "")}

    teams = list(team_map.values())

    # 设置 parent_name
    for team in teams:
        if team["parent_id"]:
            parent_assign = sb.table("team_assignments").select("team_name").eq("id", team["parent_id"]).execute().data
            if parent_assign:
                team["parent_name"] = parent_assign[0]["team_name"]

    return {"teams": teams}


@app.post("/admin/api/team")
async def admin_create_team(body: dict, admin: dict = _Depends(get_current_admin)):
    """创建新团队（先创建团队记录，再设置负责人）"""
    sb = db()
    team_name = body.get("team_name")
    team_level = body.get("team_level", 1)
    leader_id = body.get("leader_id")
    parent_team_id = body.get("parent_team_id")

    if not team_name:
        raise HTTPException(status_code=400, detail="缺少 team_name")

    # 检查团队是否已存在
    existing = sb.table("team_assignments").select("id").eq("team_name", team_name).execute().data
    if existing:
        raise HTTPException(status_code=400, detail=f"团队 '{team_name}' 已存在")

    data = {
        "team_name": team_name,
        "team_level": team_level,
    }
    if parent_team_id:
        data["parent_team_id"] = parent_team_id

    # 创建团队（没有 leader 时先创建空记录）
    if leader_id:
        data["user_id"] = leader_id
        data["is_leader"] = True

    result = sb.table("team_assignments").insert(data).execute()
    return {"ok": True, "id": result.data[0].get("id") if result.data else None}


@app.post("/admin/api/team/assign")
async def admin_assign_team(body: dict, admin: dict = _Depends(get_current_admin)):
    """设置用户团队归属"""
    sb = db()
    if "user_id" not in body or "team_name" not in body:
        raise HTTPException(status_code=400, detail="缺少 user_id 或 team_name")

    data = {
        "user_id": body["user_id"],
        "team_name": body["team_name"],
        "team_level": body.get("team_level", 1),
        "is_leader": body.get("is_leader", False),
    }

    # upsert
    existing = sb.table("team_assignments").select("id").eq("user_id", body["user_id"]).eq("team_level", data["team_level"]).execute().data
    if existing:
        sb.table("team_assignments").update(data).eq("id", existing[0]["id"]).execute()
    else:
        sb.table("team_assignments").insert(data).execute()

    return {"ok": True}


@app.post("/admin/api/team/set-parent")
async def admin_set_team_parent(body: dict, admin: dict = _Depends(get_current_admin)):
    """设置团队的父级团队（用于建立团队层级关系）"""
    sb = db()
    team_name = body.get("team_name")
    parent_team_name = body.get("parent_team_name")

    if not team_name:
        raise HTTPException(status_code=400, detail="缺少 team_name")

    if not parent_team_name:
        sb.table("team_assignments").update({"parent_team_id": None}).eq("team_name", team_name).execute()
        return {"ok": True, "message": "父级已清除"}

    parent_assign = sb.table("team_assignments").select("id, team_level").eq("team_name", parent_team_name).execute().data
    if not parent_assign:
        raise HTTPException(status_code=404, detail=f"父级团队 '{parent_team_name}' 不存在")

    parent_id = parent_assign[0]["id"]
    sb.table("team_assignments").update({"parent_team_id": parent_id}).eq("team_name", team_name).execute()

    return {"ok": True, "parent_team_id": parent_id, "parent_team_name": parent_team_name}


@app.get("/admin/api/team/{team_name}/ranking")
async def admin_team_radar(team_name: str, admin: dict = _Depends(get_current_admin)):
    """团队六维雷达（成员平均值）"""
    sb = db()

    # 获取团队成员
    assignments = sb.table("team_assignments").select("user_id").eq("team_name", team_name).execute().data or []
    user_ids = [a["user_id"] for a in assignments]

    if not user_ids:
        return {"team_name": team_name, "member_count": 0, "radar": {}, "members": []}

    # 计算每个成员的雷达
    all_radar = []
    members = []
    for uid in user_ids:
        try:
            profile = await get_user_profile(uid)
            all_radar.append(profile.get("radar", {}))
            user = sb.table("users").select("id, nickname").eq("id", uid).execute().data
            members.append({
                "user_id": uid,
                "nickname": user[0].get("nickname", "") if user else "",
                "radar": profile.get("radar", {}),
            })
        except Exception:
            pass

    if not all_radar:
        return {"team_name": team_name, "member_count": len(user_ids), "radar": {}, "members": members}

    # 平均值
    keys = ["scenarioCoverage", "painpointAnalysis", "valueMining", "skillMastery", "actionCompletion", "learningActivity"]
    avg_radar = {}
    for k in keys:
        vals = [r.get(k, 0) for r in all_radar if r.get(k) is not None]
        avg_radar[k] = round(sum(vals) / len(vals), 1) if vals else 0

    return {"team_name": team_name, "member_count": len(user_ids), "radar": avg_radar, "members": members}


@app.get("/admin/api/team/{team_name}/ranking")
async def admin_team_ranking(team_name: str, admin: dict = _Depends(get_current_admin)):
    """团队成员排行（按综合得分）"""
    sb = db()

    assignments = sb.table("team_assignments").select("user_id").eq("team_name", team_name).execute().data or []
    user_ids = [a["user_id"] for a in assignments]

    if not user_ids:
        return {"team_name": team_name, "ranking": []}

    ranking = []
    for uid in user_ids:
        try:
            profile = await get_user_profile(uid)
            radar = profile.get("radar", {})
            total_score = sum(radar.get(k, 0) for k in ["scenarioCoverage", "painpointAnalysis", "valueMining", "skillMastery", "actionCompletion", "learningActivity"])
            user = sb.table("users").select("id, nickname").eq("id", uid).execute().data
            ranking.append({
                "user_id": uid,
                "nickname": user[0].get("nickname", "") if user else "",
                "total_score": round(total_score, 1),
                "radar": radar,
            })
        except Exception:
            pass

    ranking.sort(key=lambda x: x["total_score"], reverse=True)
    for i, r in enumerate(ranking):
        r["rank"] = i + 1

    return {"team_name": team_name, "ranking": ranking}


@app.get("/admin/api/benchmarks")
async def admin_list_benchmarks(metric_type: str | None = None, admin: dict = _Depends(get_current_admin)):
    """获取基准数据"""
    sb = db()
    query = sb.table("metric_benchmarks").select("*")
    if metric_type:
        query = query.eq("metric_type", metric_type)
    data = query.order("metric_type, min_value").execute().data or []
    return {"benchmarks": data}


@app.post("/admin/api/benchmarks")
async def admin_upsert_benchmark(body: dict, admin: dict = _Depends(get_current_admin)):
    """创建或更新基准值"""
    sb = db()
    required = ["metric_type", "benchmark_level", "min_value", "max_value"]
    for k in required:
        if k not in body:
            raise HTTPException(status_code=400, detail=f"缺少必填字段: {k}")

    data = {
        "metric_type": body["metric_type"],
        "benchmark_level": body["benchmark_level"],
        "min_value": float(body["min_value"]),
        "max_value": float(body["max_value"]),
        "period_type": body.get("period_type", "monthly"),
        "description": body.get("description", ""),
    }

    existing = sb.table("metric_benchmarks").select("id").eq("metric_type", body["metric_type"]).eq("benchmark_level", body["benchmark_level"]).execute().data
    if existing:
        sb.table("metric_benchmarks").update(data).eq("id", existing[0]["id"]).execute()
    else:
        sb.table("metric_benchmarks").insert(data).execute()

    return {"ok": True}


@app.get("/api/user/{user_id}/profile/compare")
async def get_user_profile_compare(user_id: str):
    """个人 vs 基准对比"""
    sb = db()

    # 获取用户画像
    try:
        profile = await get_user_profile(user_id)
    except HTTPException:
        raise
    radar = profile.get("radar", {})

    # 获取基准（用六维雷达的 6 个 key 映射到 metric_type）
    # 简化映射：每个维度映射到一个综合指标
    benchmark_map = {
        "scenarioCoverage": "scenario_coverage",
        "painpointAnalysis": "high_value_action_rate",
        "valueMining": "avg_value_impact",
        "skillMastery": "action_completion_rate",
        "actionCompletion": "action_completion_rate",
        "learningActivity": "recent_activity",
    }

    benchmarks = sb.table("metric_benchmarks").select("*").execute().data or []
    bench_by_type = {}
    for b in benchmarks:
        bt = b["metric_type"]
        if bt not in bench_by_type:
            bench_by_type[bt] = {}
        bench_by_type[bt][b["benchmark_level"]] = {"min": b["min_value"], "max": b["max_value"]}

    # 构建对比数据
    comparison = {}
    for radar_key, bench_type in benchmark_map.items():
        user_val = radar.get(radar_key, 0)
        bench = bench_by_type.get(bench_type, {})
        comparison[radar_key] = {
            "user_value": user_val,
            "excellent": bench.get("excellent", {}).get("min", 8),
            "good": bench.get("good", {}).get("min", 6),
            "average": bench.get("average", {}).get("min", 4),
        }

    return {"radar": radar, "comparison": comparison}


@app.delete("/admin/api/metrics/{metric_id}")
async def admin_delete_metric(metric_id: int, admin: dict = _Depends(get_current_admin)):
    """删除指标记录"""
    sb = db()
    sb.table("business_metrics").delete().eq("id", metric_id).execute()
    return {"ok": True}


@app.delete("/admin/api/team/{assignment_id}")
async def admin_remove_team(assignment_id: int, admin: dict = _Depends(get_current_admin)):
    """移除团队成员"""
    sb = db()
    sb.table("team_assignments").delete().eq("id", assignment_id).execute()
    return {"ok": True}


# ═══════════ Phase 2: 诊断级 API ═══════════

@app.get("/admin/api/heatmap")
async def admin_org_heatmap(admin: dict = _Depends(get_current_admin)):
    """组织能力热力图 — X轴=六维能力，Y轴=团队，值=平均分"""
    sb = db()

    # 获取所有团队
    teams = sb.table("team_assignments").select("team_name, user_id").execute().data or []
    team_users: dict[str, list[str]] = {}
    for t in teams:
        if t["team_name"] not in team_users:
            team_users[t["team_name"]] = []
        team_users[t["team_name"]].append(t["user_id"])

    radar_keys = ["scenarioCoverage", "painpointAnalysis", "valueMining", "skillMastery", "actionCompletion", "learningActivity"]
    radar_labels = ["场景覆盖", "痛点分析", "价值挖掘", "技能掌握", "动作完成", "学习活跃"]

    heatmap = []
    for team_name, user_ids in team_users.items():
        row: dict[str, any] = {"team": team_name, "memberCount": len(user_ids), "dimensions": {}}
        all_radar = []
        for uid in user_ids:
            try:
                profile = await get_user_profile(uid)
                all_radar.append(profile.get("radar", {}))
            except Exception:
                pass

        for key in radar_keys:
            vals = [r.get(key, 0) for r in all_radar if r.get(key) is not None]
            avg = round(sum(vals) / len(vals), 1) if vals else 0
            row["dimensions"][key] = avg

        row["avgScore"] = round(sum(row["dimensions"].get(k, 0) for k in radar_keys) / len(radar_keys), 1)
        heatmap.append(row)

    heatmap.sort(key=lambda x: x["avgScore"], reverse=True)

    # 各维度全局统计
    dim_stats = {}
    for key in radar_keys:
        all_vals = [h["dimensions"].get(key, 0) for h in heatmap]
        dim_stats[key] = {
            "label": radar_labels[radar_keys.index(key)],
            "min": round(min(all_vals), 1) if all_vals else 0,
            "max": round(max(all_vals), 1) if all_vals else 0,
            "avg": round(sum(all_vals) / len(all_vals), 1) if all_vals else 0,
        }

    return {"heatmap": heatmap, "dimensions": radar_keys, "dimensionLabels": radar_labels, "dimensionStats": dim_stats}


@app.get("/api/diagnosis/alerts")
async def get_diagnosis_alerts(user_id: str):
    """异常检测 — 基于阈值规则和统计偏差"""
    sb = db()

    alerts = []

    # 获取用户指标
    metrics = sb.table("business_metrics").select("*").eq("user_id", user_id).order("period_start", desc=True).execute().data or []

    # 获取基准
    benchmarks = sb.table("metric_benchmarks").select("*").execute().data or []
    bench_map: dict[str, dict[str, float]] = {}
    for b in benchmarks:
        if b["metric_type"] not in bench_map:
            bench_map[b["metric_type"]] = {}
        bench_map[b["metric_type"]][b["benchmark_level"]] = float(b["min_value"])

    # 规则1: 指标低于基准平均值
    for m in metrics[:6]:  # 最近6条
        mt = m["metric_type"]
        avg_bench = bench_map.get(mt, {}).get("average", 0)
        if avg_bench > 0 and float(m["metric_value"]) < avg_bench * 0.5:
            alerts.append({
                "type": "low_metric",
                "severity": "warning",
                "metric_type": mt,
                "period": m["period_start"],
                "value": float(m["metric_value"]),
                "benchmark": avg_bench,
                "message": f"{mt} 低于团队平均 50% 以上（当前: {m['metric_value']}, 平均: {avg_bench}）",
            })

    # 规则2: 指标连续下降
    by_type: dict[str, list] = {}
    for m in metrics:
        mt = m["metric_type"]
        if mt not in by_type:
            by_type[mt] = []
        by_type[mt].append(float(m["metric_value"]))

    for mt, vals in by_type.items():
        if len(vals) >= 3:
            # 检查最近3期是否连续下降
            recent = vals[:3]
            if recent[0] > recent[1] > recent[2]:
                drop_pct = (recent[0] - recent[2]) / recent[0] * 100 if recent[0] > 0 else 0
                alerts.append({
                    "type": "declining_trend",
                    "severity": "danger" if drop_pct > 40 else "warning",
                    "metric_type": mt,
                    "drop_pct": round(drop_pct, 1),
                    "message": f"{mt} 连续3期下降，累计下降 {drop_pct:.0f}%",
                })

    # 规则3: 自评分与动作难度不匹配（认知偏差）
    progress = sb.table("user_action_progress").select("action_id, self_rating").eq("user_id", user_id).execute().data or []
    all_actions = sb.table("scenario_actions").select("id, difficulty").execute().data or []
    action_diff = {"easy": 3, "medium": 5, "hard": 7, "expert": 9}

    high_self_low_diff = 0
    total_with_rating = 0
    for p in progress:
        rating = p.get("self_rating", 5)
        action = next((a for a in all_actions if a["id"] == p["action_id"]), None)
        if action and rating is not None:
            total_with_rating += 1
            diff_val = action_diff.get(action.get("difficulty", "medium"), 5)
            if rating >= 8 and diff_val <= 4:  # 自评高但动作简单
                high_self_low_diff += 1

    if total_with_rating >= 5 and high_self_low_diff / total_with_rating > 0.4:
        alerts.append({
            "type": "cognitive_bias",
            "severity": "info",
            "ratio": round(high_self_low_diff / total_with_rating, 2),
            "message": f"存在认知偏差：{high_self_low_diff}/{total_with_rating} 个简单动作自评偏高（≥8分），建议关注实际业务结果",
        })

    # 规则4: 长期无动作完成（流失风险）
    from datetime import timedelta
    thirty_days_ago = (datetime.utcnow() - timedelta(days=30)).isoformat()
    recent_progress = [p for p in progress if p.get("completed_at") and p["completed_at"] > thirty_days_ago]
    if len(progress) > 3 and len(recent_progress) == 0:
        alerts.append({
            "type": "inactivity_risk",
            "severity": "danger",
            "days_inactive": 30,
            "message": "连续30天无动作完成记录，存在流失风险",
        })

    # 按严重度排序
    severity_order = {"danger": 0, "warning": 1, "info": 2}
    alerts.sort(key=lambda a: severity_order.get(a["severity"], 3))

    return {"alerts": alerts, "total": len(alerts)}


@app.get("/api/user/{user_id}/team-ranking")
async def get_user_team_ranking(user_id: str):
    """获取用户在团队中的排名"""
    sb = db()

    # 找到用户所属团队
    assignments = sb.table("team_assignments").select("team_name").eq("user_id", user_id).execute().data or []
    if not assignments:
        return {"in_team": False, "ranking": []}

    team_name = assignments[0]["team_name"]
    team_members = sb.table("team_assignments").select("user_id").eq("team_name", team_name).execute().data or []
    user_ids = [m["user_id"] for m in team_members]

    ranking = []
    for uid in user_ids:
        try:
            profile = await get_user_profile(uid)
            radar = profile.get("radar", {})
            total_score = sum(radar.get(k, 0) for k in ["scenarioCoverage", "painpointAnalysis", "valueMining", "skillMastery", "actionCompletion", "learningActivity"])
            user = sb.table("users").select("id, nickname").eq("id", uid).execute().data
            ranking.append({
                "user_id": uid,
                "nickname": user[0].get("nickname", "") if user else "",
                "total_score": round(total_score, 1),
            })
        except Exception:
            pass

    ranking.sort(key=lambda x: x["total_score"], reverse=True)

    my_rank = next((i + 1 for i, r in enumerate(ranking) if r["user_id"] == user_id), 0)

    # 隐私：只返回排名区间，不暴露其他成员姓名
    return {
        "in_team": True,
        "team_name": team_name,
        "my_rank": my_rank,
        "total_members": len(ranking),
        "my_score": next((r["total_score"] for r in ranking if r["user_id"] == user_id), 0),
        "top_score": ranking[0]["total_score"] if ranking else 0,
        "avg_score": round(sum(r["total_score"] for r in ranking) / len(ranking), 1) if ranking else 0,
    }


@app.get("/api/scenarios/{scenario_id}/process-deviation")
async def get_process_deviation(scenario_id: str, user_id: str):
    """流程偏差分析 — 用户的实际动作顺序 vs 标准顺序"""
    sb = db()

    # 标准顺序：按 sort_order 排列的已发布动作
    std_actions = sb.table("scenario_actions").select("id, title, sort_order").eq("scenario_id", scenario_id).eq("is_published", True).order("sort_order").execute().data or []

    # 用户实际完成顺序
    progress = sb.table("user_action_progress").select("action_id, completed_at, self_rating").eq("user_id", user_id).execute().data or []
    user_action_ids = {p["action_id"] for p in progress}
    completed_for_scenario = [p for p in progress if any(a["id"] == p["action_id"] for a in std_actions)]

    # 按完成时间排序
    completed_for_scenario.sort(key=lambda p: p.get("completed_at", ""))

    # 计算偏差
    std_order = {a["id"]: i for i, a in enumerate(std_actions)}
    skipped = [a for a in std_actions if a["id"] not in user_action_ids]
    completed_order = [p["action_id"] for p in completed_for_scenario]

    # 检测顺序偏差（用户完成顺序与标准顺序的差异）
    order_deviations = []
    for i, action_id in enumerate(completed_order):
        std_pos = std_order.get(action_id, -1)
        if std_pos >= 0 and std_pos != i:
            action_title = next((a["title"] for a in std_actions if a["id"] == action_id), action_id)
            order_deviations.append({
                "action_id": action_id,
                "title": action_title,
                "expected_position": std_pos + 1,
                "actual_position": i + 1,
            })

    return {
        "scenario_id": scenario_id,
        "standard_sequence": [{"id": a["id"], "title": a["title"], "sort_order": a["sort_order"]} for a in std_actions],
        "user_sequence": [{"action_id": p["action_id"], "title": next((a["title"] for a in std_actions if a["id"] == p["action_id"]), ""), "completed_at": p.get("completed_at", ""), "self_rating": p.get("self_rating", 5)} for p in completed_for_scenario],
        "completed_count": len(completed_for_scenario),
        "total_count": len(std_actions),
        "skipped": [{"id": a["id"], "title": a["title"]} for a in skipped],
        "order_deviations": order_deviations,
        "compliance_rate": round(len(completed_for_scenario) / max(len(std_actions), 1), 2),
    }


# ═══════════ Phase 3a: 预测级 API ═══════════

# ── 动作结果追踪 CRUD ──

@app.post("/api/user/{user_id}/outcomes")
async def create_outcome(user_id: str, body: dict):
    """提交动作结果"""
    sb = db()
    if "action_id" not in body or "outcome_type" not in body:
        raise HTTPException(status_code=400, detail="缺少 action_id 或 outcome_type")

    # 查找对应的 progress_id
    progress = sb.table("user_action_progress").select("id, scenario_id").eq("user_id", user_id).eq("action_id", body["action_id"]).execute().data
    progress_id = progress[0]["id"] if progress else None
    scenario_id = progress[0]["scenario_id"] if progress else body.get("scenario_id", "")

    data = {
        "user_id": user_id,
        "action_id": body["action_id"],
        "scenario_id": scenario_id,
        "progress_id": progress_id,
        "outcome_type": body["outcome_type"],
        "outcome_value": float(body.get("outcome_value", 0)),
        "outcome_date": body.get("outcome_date", datetime.utcnow().strftime("%Y-%m-%d")),
        "notes": body.get("notes", ""),
    }

    result = sb.table("action_outcomes").insert(data).execute()
    return {"ok": True, "id": result.data[0]["id"]}


@app.get("/api/user/{user_id}/outcomes")
async def list_user_outcomes(user_id: str, scenario_id: str | None = None):
    """获取用户动作结果列表"""
    sb = db()
    query = sb.table("action_outcomes").select("*").eq("user_id", user_id).order("outcome_date", desc=True)
    if scenario_id:
        query = query.eq("scenario_id", scenario_id)
    result = query.execute()
    return {"outcomes": result.data or []}


@app.get("/admin/api/outcomes")
async def admin_list_outcomes(
    user_id: str | None = None,
    scenario_id: str | None = None,
    outcome_type: str | None = None,
    page: int = 1,
    page_size: int = 50,
    admin: dict = _Depends(get_current_admin),
):
    """Admin 查看所有结果数据"""
    sb = db()
    query = sb.table("action_outcomes").select("*, users(nickname)").order("created_at", desc=True)
    if user_id:
        query = query.eq("user_id", user_id)
    if scenario_id:
        query = query.eq("scenario_id", scenario_id)
    if outcome_type:
        query = query.eq("outcome_type", outcome_type)
    start = (page - 1) * page_size
    end = page * page_size - 1
    result = query.range(start, end).execute()
    return {"outcomes": result.data or [], "page": page}


@app.get("/admin/api/outcomes/stats")
async def admin_outcomes_stats(admin: dict = _Depends(get_current_admin)):
    """结果统计：各类型数量、转化率趋势"""
    sb = db()
    outcomes = sb.table("action_outcomes").select("outcome_type, outcome_value, outcome_date").execute().data or []

    from collections import Counter
    type_counts = Counter(o["outcome_type"] for o in outcomes)
    total = len(outcomes)

    # 按月统计
    monthly: dict[str, int] = {}
    for o in outcomes:
        month = o.get("outcome_date", "")[:7]
        if month:
            monthly[month] = monthly.get(month, 0) + 1

    return {
        "total": total,
        "by_type": dict(type_counts),
        "monthly": monthly,
        "avg_value": round(sum(float(o.get("outcome_value", 0)) for o in outcomes) / max(total, 1), 2),
    }


# ── AI 数字教练 ──

@app.get("/api/user/{user_id}/coaching")
async def get_ai_coaching(user_id: str):
    """AI 数字教练 — 基于画像短板 + 指标趋势生成个性化辅导建议"""
    sb = db()

    # 获取用户画像
    try:
        profile = await get_user_profile(user_id)
    except HTTPException:
        raise

    radar = profile.get("radar", {})
    radar_base = profile.get("radarBase", radar)

    # 找出最弱的 2 个维度
    radar_keys = ["scenarioCoverage", "painpointAnalysis", "valueMining", "skillMastery", "actionCompletion", "learningActivity"]
    radar_labels = ["场景覆盖", "痛点分析", "价值挖掘", "技能掌握", "动作完成", "学习活跃"]
    sorted_dims = sorted(range(6), key=lambda i: radar.get(radar_keys[i], 0))
    weak_dims = sorted_dims[:2]
    strong_dims = sorted_dims[-2:]

    # 获取最弱维度对应的未完成动作
    learning_path = profile.get("learningPath", [])
    priority_actions = []
    for lp in learning_path[:3]:
        action = sb.table("scenario_actions").select("id, title, scenario_id, value_impact, difficulty").eq("id", lp["nextActionId"]).execute().data
        if action:
            a = action[0]
            scenario = sb.table("scenario_tree_nodes").select("title").eq("id", a["scenario_id"]).execute().data
            priority_actions.append({
                "action_id": a["id"],
                "title": a["title"],
                "scenario": scenario[0]["title"] if scenario else "",
                "value_impact": a["value_impact"],
                "difficulty": a["difficulty"],
            })

    # 获取最近指标趋势
    from datetime import timedelta
    three_months_ago = (datetime.utcnow() - timedelta(days=90)).strftime("%Y-%m-%d")
    metrics = sb.table("business_metrics").select("*").eq("user_id", user_id).gte("period_start", three_months_ago).order("period_start", desc=True).execute().data or []
    metric_summary = {}
    for m in metrics[:4]:
        mt = m["metric_type"]
        if mt not in metric_summary:
            metric_summary[mt] = float(m["metric_value"])

    # 获取团队排名（找同侪导师）
    try:
        ranking = await get_user_team_ranking(user_id)
        peer_info = f"团队排名: 第{ranking.get('my_rank', '?')}/{ranking.get('total_members', '?')}名"
    except Exception:
        peer_info = ""

    # 构建 coaching 文本（不依赖 LLM 也可用）
    weak_labels = [radar_labels[i] for i in weak_dims]
    strong_labels = [radar_labels[i] for i in strong_dims]

    recommendation = f"你在'{weak_labels[0]}'和'{weak_labels[1]}'维度得分较低，建议优先提升。"
    if priority_actions:
        recommendation += f"\n\n建议优先完成以下动作："
        for pa in priority_actions:
            recommendation += f"\n• {pa['scenario']} — {pa['title']}（价值度: {pa['value_impact']}/10）"
    if peer_info:
        recommendation += f"\n\n{peer_info}。"
    if metric_summary:
        recommendation += f"\n\n近期指标："
        type_labels = {"visit_count": "拜访量", "new_store": "新开店", "deal_amount": "成交额", "repeat_rate": "复购率"}
        for mt, val in metric_summary.items():
            recommendation += f"\n• {type_labels.get(mt, mt)}: {val}"

    # 尝试调用 LLM 生成更丰富的建议
    ai_recommendation = recommendation
    try:
        from app.core.llm.dashscope_client import call_llm
        system_prompt = """你是一位高端白酒销售培训教练。根据业务员的画像数据和指标趋势，给出简短、具体、可操作的辅导建议。
要求：
1. 用中文回复，语气亲切但专业
2. 突出最需要提升的维度
3. 给出 2-3 个具体的行动建议
4. 控制在 150 字以内"""

        user_prompt = f"""业务员画像数据：
最弱维度：{', '.join(weak_labels)}（得分：{radar.get(radar_keys[weak_dims[0]], 0)}, {radar.get(radar_keys[weak_dims[1]], 0)}）
最强维度：{', '.join(strong_labels)}（得分：{radar.get(radar_keys[strong_dims[0]], 0)}, {radar.get(radar_keys[strong_dims[1]], 0)}）
{peer_info}
{f'近期指标：{metric_summary}' if metric_summary else ''}
{f'建议动作：{[pa["title"] for pa in priority_actions]}' if priority_actions else ''}"""

        llm_result = call_llm(system_prompt, user_prompt)
        if llm_result and llm_result.reasoning:
            ai_recommendation = llm_result.reasoning
    except Exception:
        pass  # LLM 失败时使用规则生成的建议

    return {
        "weak_dimensions": weak_labels,
        "strong_dimensions": strong_labels,
        "priority_actions": priority_actions,
        "recommendation": ai_recommendation,
        "metric_summary": metric_summary,
    }


# ── 趋势预测 ──

@app.get("/api/user/{user_id}/prediction")
async def get_prediction(user_id: str):
    """基于历史数据的简单趋势预测（线性外推）"""
    sb = db()

    from datetime import timedelta
    six_months_ago = (datetime.utcnow() - timedelta(days=180)).strftime("%Y-%m-%d")
    metrics = sb.table("business_metrics").select("*").eq("user_id", user_id).gte("period_start", six_months_ago).order("period_start").execute().data or []

    # 按指标类型分组，计算趋势
    predictions = {}
    by_type: dict[str, list] = {}
    for m in metrics:
        mt = m["metric_type"]
        if mt not in by_type:
            by_type[mt] = []
        by_type[mt].append(float(m["metric_value"]))

    type_labels = {"visit_count": "拜访量", "new_store": "新开店", "deal_amount": "成交额", "repeat_rate": "复购率"}

    for mt, vals in by_type.items():
        if len(vals) < 2:
            predictions[mt] = {"label": type_labels.get(mt, mt), "trend": "stable", "current": vals[-1] if vals else 0, "predicted": vals[-1] if vals else 0, "confidence": 0}
            continue

        # 简单线性回归
        n = len(vals)
        x_mean = (n - 1) / 2
        y_mean = sum(vals) / n
        numerator = sum((i - x_mean) * (vals[i] - y_mean) for i in range(n))
        denominator = sum((i - x_mean) ** 2 for i in range(n))

        if denominator == 0:
            slope = 0
        else:
            slope = numerator / denominator

        predicted = round(vals[-1] + slope, 1)
        current = vals[-1]

        # 趋势方向
        if slope > 0.05 * max(y_mean, 1):
            trend = "up"
        elif slope < -0.05 * max(y_mean, 1):
            trend = "down"
        else:
            trend = "stable"

        # 置信度（基于 R² 的简化版）
        ss_res = sum((vals[i] - (y_mean + slope * (i - x_mean))) ** 2 for i in range(n))
        ss_tot = sum((v - y_mean) ** 2 for v in vals)
        r_squared = 1 - ss_res / ss_tot if ss_tot > 0 else 0
        confidence = round(max(0, min(1, r_squared)) * 100)

        predictions[mt] = {
            "label": type_labels.get(mt, mt),
            "trend": trend,
            "current": current,
            "predicted": predicted,
            "change_pct": round((predicted - current) / max(current, 0.01) * 100, 1),
            "confidence": confidence,
        }

    return {"predictions": predictions}


# ── What-If 模拟（Admin） ──

@app.get("/admin/api/whatif")
async def admin_whatif(
    action_id: str,
    target_value: int,
    target_experience: int,
    admin: dict = _Depends(get_current_admin),
):
    """What-If 模拟：将动作移到新象限后预测指标变化"""
    sb = db()

    action = sb.table("scenario_actions").select("*, scenario_tree_nodes(title)").eq("id", action_id).eq("is_active", True).execute().data
    if not action:
        raise HTTPException(status_code=404, detail="动作不存在")
    action = action[0]

    old_value = action["value_impact"]
    old_experience = action["experience_impact"]

    # 计算该动作被多少人完成
    completed_count = sb.table("user_action_progress").select("user_id", count="exact").eq("action_id", action_id).not_().is_("completed_at", "null").execute()
    completed_users = completed_count.count if completed_count.count else 0

    # 模拟影响
    value_change = target_value - old_value
    experience_change = target_experience - old_experience

    # 简化预测：价值度每增加1，预期成交额提升5%；体验度每增加1，预期复购率提升2%
    predicted_deal_change = f"{'+' if value_change > 0 else ''}{value_change * 5}%"
    predicted_return_change = f"{'+' if experience_change > 0 else ''}{experience_change * 2}%"

    # 该动作所在场景的其他动作分布
    scenario_actions = sb.table("scenario_actions").select("id, title, value_impact, experience_impact").eq("scenario_id", action["scenario_id"]).eq("is_active", True).execute().data or []

    return {
        "action": {
            "id": action["id"],
            "title": action["title"],
            "scenario": action.get("scenario_tree_nodes", {}).get("title", "") if isinstance(action.get("scenario_tree_nodes"), dict) else "",
        },
        "change": {
            "value": {"from": old_value, "to": target_value, "delta": value_change},
            "experience": {"from": old_experience, "to": target_experience, "delta": experience_change},
        },
        "impact": {
            "affected_users": completed_users,
            "predicted_deal_change": predicted_deal_change,
            "predicted_return_change": predicted_return_change,
        },
        "scenario_distribution": [
            {"id": a["id"], "title": a["title"], "value": a["value_impact"], "experience": a["experience_impact"]}
            for a in scenario_actions
        ],
    }


# ═══════════ Phase 3b: 产品化 API ═══════════

# ── 多租户管理 ──

@app.get("/admin/api/tenants")
async def admin_list_tenants(admin: dict = _Depends(get_current_admin)):
    """获取所有租户列表"""
    sb = admin_table()
    result = sb.table("tenant_configs").select("*").order("created_at").execute().data or []
    return {"tenants": result}


@app.post("/admin/api/tenants")
async def admin_create_tenant(body: dict, admin: dict = _Depends(get_current_admin)):
    """新增租户"""
    sb = admin_table()
    tenant_id = body.get("tenant_id", "").strip()
    display_name = body.get("display_name", "").strip()
    industry = body.get("industry", "liquor")
    if not tenant_id or not display_name:
        raise HTTPException(400, detail="tenant_id 和 display_name 不能为空")
    # 检查是否已存在
    existing = sb.table("tenant_configs").select("id").eq("tenant_id", tenant_id).execute().data
    if existing:
        raise HTTPException(400, detail=f"租户 {tenant_id} 已存在")
    now = datetime.utcnow().isoformat()
    sb.table("tenant_configs").insert({
        "tenant_id": tenant_id,
        "display_name": display_name,
        "industry": industry,
        "logo_url": body.get("logo_url", ""),
        "primary_color": body.get("primary_color", "#1a365d"),
        "accent_color": body.get("accent_color", "#d4af37"),
        "custom_metric_types": body.get("custom_metric_types", []),
        "custom_dimensions": body.get("custom_dimensions", []),
        "is_active": True,
        "created_at": now,
        "updated_at": now,
    }).execute()
    return {"ok": True, "tenant_id": tenant_id}


@app.get("/admin/api/tenant/config")
async def admin_tenant_config(admin: dict = _Depends(get_current_admin)):
    """获取当前租户配置"""
    sb = admin_table()
    result = sb.table("tenant_configs").select("*").eq("tenant_id", "default").execute().data
    return result[0] if result else {"tenant_id": "default", "display_name": "默认", "industry": "liquor"}


@app.put("/admin/api/tenant/config")
async def admin_update_tenant_config(body: dict, admin: dict = _Depends(get_current_admin)):
    """更新当前租户配置"""
    sb = admin_table()
    updatable = ["display_name", "industry", "logo_url", "primary_color", "accent_color", "custom_metric_types", "custom_dimensions"]
    data = {k: v for k, v in body.items() if k in updatable}
    data["updated_at"] = datetime.utcnow().isoformat()
    sb.table("tenant_configs").update(data).eq("tenant_id", "default").execute()
    return {"ok": True}


@app.get("/admin/api/industry-templates")
async def admin_list_templates(industry: str | None = None, admin: dict = _Depends(get_current_admin)):
    """获取行业模板"""
    sb = admin_table()
    query = sb.table("industry_templates").select("*").eq("is_active", True).order("sort_order")
    if industry:
        query = query.eq("industry", industry)
    return {"templates": query.execute().data or []}


@app.post("/admin/api/tenant/init-from-template")
async def admin_init_from_template(body: dict, admin: dict = _Depends(get_current_admin)):
    """从行业模板初始化租户配置"""
    sb = db()
    industry = body.get("industry", "liquor")
    template_type = body.get("template_type", "l1_structure")

    templates = sb.table("industry_templates").select("data").eq("industry", industry).eq("template_type", template_type).execute().data
    if not templates:
        raise HTTPException(status_code=404, detail=f"未找到 {industry} 行业的 {template_type} 模板")

    return {"ok": True, "data": templates[0]["data"]}


# ── 自动诊断报告 ──

@app.get("/admin/api/report/monthly")
async def admin_monthly_report(admin: dict = _Depends(get_current_admin)):
    """生成月度诊断报告数据"""
    sb = db()

    # 1. 数据健康快照
    l3_nodes = sb.table("scenario_tree_nodes").select("id").eq("level", 3).eq("is_active", True).execute().data or []
    actions = sb.table("scenario_actions").select("id, scenario_id, is_published").eq("is_active", True).execute().data or []
    published_actions = [a for a in actions if a.get("is_published")]

    scenario_ids_with_actions = {a["scenario_id"] for a in actions}
    l3_total = len(l3_nodes)
    l3_with_actions = len([n for n in l3_nodes if n["id"] in scenario_ids_with_actions])

    # 2. 用户活跃度
    users = sb.table("users").select("id, nickname, created_at").execute().data or []
    user_ids = [u["id"] for u in users]
    total_users = len(user_ids)

    active_users_30d = 0
    thirty_days_ago = (datetime.utcnow() - timedelta(days=30)).isoformat()
    for uid in user_ids:
        progress = sb.table("user_action_progress").select("id").eq("user_id", uid).gte("completed_at", thirty_days_ago).execute().data
        if progress:
            active_users_30d += 1

    # 3. 指标汇总
    metrics = sb.table("business_metrics").select("*").execute().data or []
    total_metrics = len(metrics)
    metrics_by_type: dict[str, list] = {}
    for m in metrics:
        mt = m["metric_type"]
        if mt not in metrics_by_type:
            metrics_by_type[mt] = []
        metrics_by_type[mt].append(float(m["metric_value"]))

    metric_summary = {}
    for mt, vals in metrics_by_type.items():
        metric_summary[mt] = {
            "count": len(vals),
            "avg": round(sum(vals) / len(vals), 1) if vals else 0,
            "max": round(max(vals), 1) if vals else 0,
            "min": round(min(vals), 1) if vals else 0,
        }

    # 4. 动作结果统计
    outcomes = sb.table("action_outcomes").select("outcome_type").execute().data or []
    from collections import Counter
    outcome_counts = Counter(o["outcome_type"] for o in outcomes)

    # 5. 团队数据（如果有）
    teams = sb.table("team_assignments").select("team_name, user_id").execute().data or []
    team_names = list(set(t["team_name"] for t in teams))

    # 6. 关键发现（自动生成）
    findings = []
    if l3_total > 0 and l3_with_actions / l3_total < 0.5:
        findings.append({"type": "warning", "text": f"场景动作覆盖率仅 {round(l3_with_actions / l3_total * 100)}%（{l3_with_actions}/{l3_total}），建议加速动作填充"})
    if total_users > 0 and active_users_30d / total_users < 0.3:
        findings.append({"type": "danger", "text": f"30天活跃率仅 {round(active_users_30d / total_users * 100)}%（{active_users_30d}/{total_users}），存在用户流失风险"})
    if total_metrics > 0:
        findings.append({"type": "info", "text": f"已录入 {total_metrics} 条指标数据，覆盖 {len(metrics_by_type)} 个指标类型"})
    if not findings:
        findings.append({"type": "info", "text": "系统运行正常"})

    # 7. 租户配置
    tenant = sb.table("tenant_configs").select("*").eq("tenant_id", "default").execute().data
    tenant_info = tenant[0] if tenant else {}

    return {
        "generated_at": datetime.utcnow().isoformat(),
        "tenant": tenant_info,
        "data_health": {
            "l3_total": l3_total,
            "l3_with_actions": l3_with_actions,
            "action_coverage": round(l3_with_actions / max(l3_total, 1), 2),
            "total_published_actions": len(published_actions),
        },
        "user_activity": {
            "total_users": total_users,
            "active_30d": active_users_30d,
            "active_rate": round(active_users_30d / max(total_users, 1), 2),
        },
        "metric_summary": metric_summary,
        "outcome_counts": dict(outcome_counts),
        "team_count": len(team_names),
        "findings": findings,
    }


# ── API 开放平台 ──

@app.post("/api/integration/wecom/sync")
async def wecom_sync(body: dict):
    """企微数据同步入口"""
    sb = db()
    records = body.get("records", [])

    if not records:
        return {"ok": True, "synced": 0}

    synced = 0
    errors = []

    for rec in records:
        try:
            # 根据 type 路由到不同的处理逻辑
            sync_type = rec.get("type", "")
            if sync_type == "metric":
                data = {
                    "user_id": rec["user_id"],
                    "metric_type": rec["metric_type"],
                    "metric_value": float(rec["metric_value"]),
                    "period_start": rec["period_start"],
                    "period_end": rec["period_end"],
                    "data_source": "wecom",
                    "notes": f"企微同步: {rec.get('source', '')}",
                }
                sb.table("business_metrics").insert(data).execute()
                synced += 1
            elif sync_type == "action_complete":
                from app.core.feature_flags import is_feature_enabled
                if is_feature_enabled("future.action_system"):
                    sb.table("user_action_progress").upsert(
                        {"user_id": rec["user_id"], "action_id": rec["action_id"], "scenario_id": rec.get("scenario_id", ""), "completed_at": rec.get("completed_at", datetime.utcnow().isoformat())},
                        on_conflict="user_id,action_id"
                    ).execute()
                    synced += 1
            else:
                errors.append({"record": rec, "error": f"未知同步类型: {sync_type}"})
        except Exception as e:
            errors.append({"record": rec, "error": str(e)})

    # 记录同步日志
    sb.table("integration_logs").insert({
        "tenant_id": "default",
        "integration_type": "wecom",
        "direction": "inbound",
        "status": "success" if not errors else "partial",
        "record_count": synced,
        "error_message": str(errors[:5]) if errors else "",
    }).execute()

    return {"ok": True, "synced": synced, "errors": len(errors)}


@app.post("/api/integration/crm/sync")
async def crm_sync(body: dict):
    """CRM 数据同步入口"""
    sb = db()
    records = body.get("records", [])

    synced = 0
    for rec in records:
        try:
            sync_type = rec.get("type", "")
            if sync_type == "metric":
                data = {
                    "user_id": rec["user_id"],
                    "metric_type": rec["metric_type"],
                    "metric_value": float(rec["metric_value"]),
                    "period_start": rec["period_start"],
                    "period_end": rec["period_end"],
                    "data_source": "crm",
                    "notes": f"CRM同步: {rec.get('source', '')}",
                }
                sb.table("business_metrics").insert(data).execute()
                synced += 1
        except Exception:
            pass

    sb.table("integration_logs").insert({
        "tenant_id": "default",
        "integration_type": "crm",
        "direction": "inbound",
        "status": "success",
        "record_count": synced,
    }).execute()

    return {"ok": True, "synced": synced}


@app.get("/admin/api/integration/logs")
async def admin_integration_logs(
    integration_type: str | None = None,
    limit: int = 50,
    admin: dict = _Depends(get_current_admin),
):
    """查看集成日志"""
    sb = db()
    query = sb.table("integration_logs").select("*").order("created_at", desc=True).limit(limit)
    if integration_type:
        query = query.eq("integration_type", integration_type)
    return {"logs": query.execute().data or []}


@app.get("/admin/api/export/data")
async def admin_export_data(
    data_type: str = "scenarios",
    format: str = "json",
    admin: dict = _Depends(get_current_admin),
):
    """导出诊断数据"""
    sb = db()

    if data_type == "scenarios":
        data = sb.table("scenario_tree_nodes").select("*").eq("is_active", True).order("level, sort_order").execute().data
    elif data_type == "actions":
        data = sb.table("scenario_actions").select("*").eq("is_active", True).execute().data
    elif data_type == "metrics":
        data = sb.table("business_metrics").select("*").order("period_start", desc=True).execute().data
    elif data_type == "outcomes":
        data = sb.table("action_outcomes").select("*").order("created_at", desc=True).execute().data
    elif data_type == "progress":
        data = sb.table("user_action_progress").select("*").order("completed_at", desc=True).execute().data
    else:
        data = []

    return {"data_type": data_type, "count": len(data), "exported_at": datetime.utcnow().isoformat(), "data": data}


# ── 练习 API ──

@app.get("/api/scenarios/{scenario_id}/practice")
async def get_practice_questions(scenario_id: str):
    """获取场景练习题目"""
    sb = db()
    questions = sb.table("practice_questions").select("*").eq("scenario_id", scenario_id).eq("is_active", True).order("question_order").execute().data or []
    if not questions:
        return {"questions": [], "scenarioId": scenario_id}
    result = []
    for q in questions:
        result.append({
            "id": q["id"],
            "order": q["question_order"],
            "situation": q["situation"],
            "options": q["options"],
            "correctIndex": q["correct_index"],
            "correctFeedback": q["correct_feedback"],
            "wrongFeedback": q["wrong_feedback"],
            "difficulty": q["difficulty"],
        })
    return {"questions": result, "scenarioId": scenario_id}


@app.post("/api/practice/submit")
async def submit_practice(body: dict):
    """提交练习结果"""
    sb = db()
    user_id = body.get("userId")
    scenario_id = body.get("scenarioId")
    answers = body.get("answers", [])
    time_spent = body.get("timeSpent", 0)

    if not user_id or not scenario_id:
        raise HTTPException(status_code=400, detail="userId 和 scenarioId 不能为空")

    # 获取题目和正确答案
    questions = sb.table("practice_questions").select("*").eq("scenario_id", scenario_id).eq("is_active", True).order("question_order").execute().data or []
    if not questions:
        raise HTTPException(status_code=404, detail="该场景暂无练习题目")

    correct = 0
    for i, q in enumerate(questions):
        if i < len(answers) and answers[i] == q["correct_index"]:
            correct += 1

    total = len(questions)
    score_pct = round(correct / total * 100) if total else 0
    passed = score_pct >= 80

    result = sb.table("practice_sessions").insert({
        "user_id": user_id,
        "scenario_id": scenario_id,
        "total_questions": total,
        "correct_answers": correct,
        "score_pct": score_pct,
        "passed": passed,
        "time_spent_sec": time_spent,
        "answers": answers,
        "completed_at": datetime.utcnow().isoformat(),
    }).execute()

    return {
        "sessionId": result.data[0]["id"] if result.data else None,
        "score": score_pct,
        "correct": correct,
        "total": total,
        "passed": passed,
        "levelProgress": 0.5 if passed else 0,
    }


# ═══════════════════════════════════════════════════════════
#  访谈分析模块 — Interview Analysis
# ═══════════════════════════════════════════════════════════

from app.core.llm.interview_analyzer import analyze_interview, analyze_batch, build_user_query
from app.core.interview_writer import write_analysis_results


@app.get("/admin/api/interviews")
async def admin_list_interviews(
    analysis_status: str | None = None,
    review_status: str | None = None,
    page: int = 1, page_size: int = 20,
    admin: dict = _Depends(get_current_admin),
    _=require_feature("admin.interview_analysis"),
):
    """获取访谈列表"""
    sb = db()
    query = sb.table("interview_records").select("*", count="exact")
    if analysis_status:
        query = query.eq("analysis_status", analysis_status)
    if review_status:
        query = query.eq("review_status", review_status)
    # 访谈列表不做角色过滤（管理功能，应展示所有访谈）
    start = (page - 1) * page_size
    end = page * page_size - 1
    result = query.order("created_at", desc=True).range(start, end).execute()
    items = [_row_to_camel(r) for r in (result.data or [])]
    return {"items": items, "total": result.count, "page": page, "page_size": page_size}


@app.get("/admin/api/interviews/{interview_id}")
async def admin_get_interview(
    interview_id: int,
    admin: dict = _Depends(get_current_admin),
    _=require_feature("admin.interview_analysis"),
):
    """获取访谈详情（含分析结果）"""
    sb = db()
    row = sb.table("interview_records").select("*").eq("id", interview_id).execute().data
    if not row:
        raise HTTPException(status_code=404, detail="访谈不存在")
    return _row_to_camel(row[0])


@app.post("/admin/api/interviews")
async def admin_create_interview(
    body: dict,
    admin: dict = _Depends(get_current_admin),
    _=require_feature("admin.interview_analysis"),
):
    """创建访谈记录"""
    data = _body_to_snake(body)
    data["created_by"] = admin["id"]
    data["analysis_status"] = "pending"
    data["review_status"] = "pending"
    data["write_status"] = "pending"
    sb = db()
    result = sb.table("interview_records").insert(data).execute()
    if not result.data:
        raise HTTPException(status_code=500, detail="创建失败")
    return _row_to_camel(result.data[0])


@app.put("/admin/api/interviews/{interview_id}")
async def admin_update_interview(
    interview_id: int,
    body: dict,
    admin: dict = _Depends(get_current_admin),
    _=require_feature("admin.interview_analysis"),
):
    """更新访谈元信息（标题/被访者/日期等）"""
    sb = db()
    row = sb.table("interview_records").select("id").eq("id", interview_id).execute().data
    if not row:
        raise HTTPException(status_code=404, detail="访谈不存在")
    data = _body_to_snake(body)
    data["updated_at"] = datetime.now().isoformat()
    # 只允许更新元信息字段
    allowed = {"title", "content", "format_type", "interviewee", "role_id", "interview_date", "is_active"}
    data = {k: v for k, v in data.items() if k in allowed}
    result = sb.table("interview_records").update(data).eq("id", interview_id).execute()
    return _row_to_camel(result.data[0])


@app.delete("/admin/api/interviews/{interview_id}")
async def admin_delete_interview(
    interview_id: int,
    admin: dict = _Depends(get_current_admin),
    _=require_feature("admin.interview_analysis"),
):
    """删除访谈记录"""
    sb = db()
    row = sb.table("interview_records").select("id").eq("id", interview_id).execute().data
    if not row:
        raise HTTPException(status_code=404, detail="访谈不存在")
    sb.table("interview_records").update({"is_active": False}).eq("id", interview_id).execute()
    return {"ok": True}


@app.post("/admin/api/interviews/{interview_id}/analyze")
async def admin_analyze_interview(
    interview_id: int,
    body: dict,
    admin: dict = _Depends(get_current_admin),
    _=require_feature("admin.interview_analysis"),
):
    """触发 LLM 分析访谈"""
    sb = db()
    row = sb.table("interview_records").select("*").eq("id", interview_id).execute().data
    if not row:
        raise HTTPException(status_code=404, detail="访谈不存在")

    record = row[0]
    content = record.get("content", "")
    if not content.strip():
        raise HTTPException(status_code=400, detail="访谈内容为空")

    # 更新状态为 analyzing
    sb.table("interview_records").update({
        "analysis_status": "analyzing",
        "updated_at": datetime.now().isoformat(),
    }).eq("id", interview_id).execute()

    try:
        # 加载标签字典
        tag_dict = db().table("tags").select("*").eq("is_active", True).execute().data or []

        # 加载场景树结构（L2 + L3）
        scenario_tree = (
            db().table("scenario_tree_nodes")
            .select("id, level, parent_id, title")
            .in_("level", [2, 3])
            .eq("is_active", True)
            .execute().data or []
        )

        # 获取角色名
        role_name = None
        if record.get("role_id"):
            role_row = db().table("roles").select("name").eq("id", record["role_id"]).execute().data
            if role_row:
                role_name = role_row[0]["name"]

        # 获取租户名
        tenant_name = "国窖1573"
        try:
            slug = _tenant_schema.get()
            tc = admin_table().table("tenant_configs").select("display_name").eq("tenant_id", slug).execute().data
            if tc:
                tenant_name = tc[0]["display_name"]
        except Exception:
            pass

        # 执行分析
        analysis = analyze_interview(
            content=content,
            tag_dict=tag_dict,
            scenario_tree=scenario_tree,
            tenant_name=tenant_name,
            interviewee=record.get("interviewee"),
            role_name=role_name,
            interview_date=record.get("interview_date"),
        )

        # 保存结果
        update_data = {
            "analysis_status": "completed" if analysis["ok"] else "failed",
            "analysis_result": analysis["result"],
            "analysis_model": analysis["model"],
            "analysis_error": analysis["error"],
            "review_status": "reviewing" if analysis["ok"] else "pending",
            "analyzed_at": datetime.now().isoformat(),
            "updated_at": datetime.now().isoformat(),
        }
        sb.table("interview_records").update(update_data).eq("id", interview_id).execute()

        return {
            "ok": analysis["ok"],
            "interviewId": interview_id,
            "analysisStatus": update_data["analysis_status"],
            "analysisResult": analysis["result"],
            "modelUsed": analysis["model"],
            "latencyMs": analysis["latencyMs"],
            "warnings": analysis["warnings"],
        }

    except Exception as e:
        logger.error(f"Interview analysis failed: {e}")
        sb.table("interview_records").update({
            "analysis_status": "failed",
            "analysis_error": str(e),
            "updated_at": datetime.now().isoformat(),
        }).eq("id", interview_id).execute()
        raise HTTPException(status_code=500, detail=f"分析失败: {str(e)}")


@app.post("/admin/api/interviews/{interview_id}/review")
async def admin_review_interview(
    interview_id: int,
    body: dict,
    admin: dict = _Depends(get_current_admin),
    _=require_feature("admin.interview_analysis"),
):
    """提交审阅结果（管理员编辑后的提取数据）"""
    sb = db()
    row = sb.table("interview_records").select("id, analysis_result").eq("id", interview_id).execute().data
    if not row:
        raise HTTPException(status_code=404, detail="访谈不存在")

    data = _body_to_snake(body)
    review_status = data.get("review_status", "confirmed")
    reviewer_note = data.get("reviewer_note", "")
    edited_result = data.get("edited_result")

    update_data = {
        "review_status": review_status,
        "reviewer_note": reviewer_note,
        "reviewed_at": datetime.now().isoformat(),
        "updated_at": datetime.now().isoformat(),
    }
    # 如果管理员提供了编辑后的结果，覆盖原始分析结果
    if edited_result:
        update_data["analysis_result"] = edited_result

    sb.table("interview_records").update(update_data).eq("id", interview_id).execute()
    return {"ok": True, "interviewId": interview_id, "reviewStatus": review_status}


@app.post("/admin/api/interviews/{interview_id}/confirm-write")
async def admin_confirm_write_interview(
    interview_id: int,
    admin: dict = _Depends(get_current_admin),
    _=require_feature("admin.interview_analysis"),
):
    """将审阅后的分析结果写入草稿（场景树 + 四维度 + 行动）"""
    sb = db()
    row = sb.table("interview_records").select(
        "id, analysis_result, role_id"
    ).eq("id", interview_id).execute().data
    if not row:
        raise HTTPException(status_code=404, detail="访谈不存在")

    record = row[0]
    analysis_result = record.get("analysis_result")
    if not analysis_result:
        raise HTTPException(status_code=400, detail="没有分析结果可写入，请先完成分析")

    try:
        result = write_analysis_results(
            sb=sb,
            interview_id=interview_id,
            analysis_result=analysis_result,
            admin_id=admin["id"],
            role_id=record.get("role_id"),
        )

        # 持久化写入结果到数据库
        write_status = "completed" if not result["errors"] else "partial"
        sb.table("interview_records").update({
            "write_status": write_status,
            "write_result": result,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }).eq("id", interview_id).execute()

        return {
            "ok": True,
            "interviewId": interview_id,
            "writeStatus": "completed" if not result["errors"] else "partial",
            "created": {
                "scenarios": len(result["scenarios"]),
                "dimensions": len(result["dimensions"]),
                "actions": len(result["actions"]),
            },
            "writeResult": result,
        }

    except Exception as e:
        logger.error(f"Confirm write failed: {e}")
        sb.table("interview_records").update({
            "write_status": "failed",
            "updated_at": datetime.now().isoformat(),
        }).eq("id", interview_id).execute()
        raise HTTPException(status_code=500, detail=f"写入失败: {str(e)}")


# ──────────────────────────────────────────────
# v2: 文件上传
# ──────────────────────────────────────────────
@app.post("/admin/api/interviews/upload")
async def admin_upload_interview(
    file: UploadFile = File(...),
    title: str | None = None,
    interviewee: str | None = None,
    role_id: int | None = None,
    interview_date: date | None = None,
    admin: dict = _Depends(get_current_admin),
    _=require_feature("admin.interview_analysis"),
):
    """上传文件创建访谈记录（支持 md/json/txt）"""
    ALLOWED = {".md", ".json", ".txt"}
    ext = ""
    if file.filename:
        dot = file.filename.rfind(".")
        if dot >= 0:
            ext = file.filename[dot:].lower()

    if ext not in ALLOWED:
        raise HTTPException(status_code=400, detail=f"不支持的文件格式 {ext}，仅支持 {', '.join(ALLOWED)}")

    try:
        raw = await file.read()
        if ext == ".json":
            payload = json.loads(raw)
            if isinstance(payload, dict):
                content = payload.get("content", payload.get("text", json.dumps(payload, ensure_ascii=False)))
            else:
                content = str(payload)
        else:
            content = raw.decode("utf-8", errors="replace")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件读取失败: {str(e)}")

    if not content.strip():
        raise HTTPException(status_code=400, detail="文件内容为空")

    file_title = title or (file.filename or "未命名访谈")
    sb = db()

    data = {
        "title": file_title,
        "content": content.strip(),
        "interviewee": interviewee,
        "role_id": role_id,
        "interview_date": interview_date,
        "format_type": "structured" if ext == ".json" else "mixed",
        "source_files": json.dumps([{
            "filename": file.filename or "unknown",
            "size": len(raw),
            "uploadedAt": datetime.now(timezone.utc).isoformat(),
        }], ensure_ascii=False),
        "analysis_status": "pending",
        "review_status": "pending",
        "write_status": "pending",
        "created_by": admin.get("id"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    res = sb.table("interview_records").insert(data).execute()
    if not res.data:
        raise HTTPException(status_code=500, detail="创建失败")

    return {"ok": True, "interview": res.data[0]}


# ──────────────────────────────────────────────
# v2: 汇总分析
# ──────────────────────────────────────────────
@app.post("/admin/api/interviews/batch-analyze")
async def admin_batch_analyze(
    body: dict,
    admin: dict = _Depends(get_current_admin),
    _=require_feature("admin.interview_analysis"),
):
    """选择多条访谈进行汇总分析"""
    data = _body_to_snake(body)
    interview_ids = data.get("interview_ids", [])
    interview_ids = [int(x) for x in interview_ids]
    if not interview_ids or len(interview_ids) < 2:
        raise HTTPException(status_code=400, detail="请至少选择 2 条访谈")

    sb = db()
    rows = (
        sb.table("interview_records")
        .select("id, title, content, interviewee, role_id, interview_date")
        .in_("id", interview_ids)
        .eq("is_active", True)
        .execute().data or []
    )

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="有效访谈不足 2 条")

    # 加载上下文
    tag_dict = sb.table("tags").select("*").eq("is_active", True).execute().data or []
    scenario_tree = (
        sb.table("scenario_tree_nodes")
        .select("id, level, parent_id, title")
        .in_("level", [2, 3])
        .eq("is_active", True)
        .execute().data or []
    )

    tenant_name = "国窖1573"
    try:
        slug = _tenant_schema.get()
        tc = admin_table().table("tenant_configs").select("display_name").eq("tenant_id", slug).execute().data
        if tc:
            tenant_name = tc[0]["display_name"]
    except Exception:
        pass

    analysis = analyze_batch(rows, tag_dict, scenario_tree, tenant_name)

    if not analysis["ok"]:
        return {"ok": False, "error": analysis.get("error", "汇总分析失败")}

    # 保存汇总结果到 interview_summaries
    title = f"汇总分析 — {len(rows)} 条访谈 ({datetime.now().strftime('%m/%d %H:%M')})"
    # interview_ids 是 integer[]，需要用 PostgreSQL 数组字面量格式
    ids_array_str = "{" + ",".join(str(x) for x in interview_ids) + "}"
    # 通过 rpc 调用 raw SQL 处理类型
    from app.core.database.connection import get_connection
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """INSERT INTO interview_summaries (title, interview_ids, summary_result, import_status, created_by)
           VALUES (%s, %s, %s, 'pending', %s) RETURNING id""",
        (title, ids_array_str, Json(analysis["result"]), admin.get("id")),
    )
    new_id = cur.fetchone()["id"]
    cur.close()
    conn.close()
    res = sb.table("interview_summaries").select("*").eq("id", new_id).execute()

    return {
        "ok": True,
        "summaryId": res.data[0]["id"] if res.data else None,
        "summaryTitle": title,
        "result": analysis["result"],
        "modelUsed": analysis.get("model", ""),
        "latencyMs": analysis.get("latencyMs", 0),
        "warnings": analysis.get("warnings", []),
    }


# ──────────────────────────────────────────────
# v2: 汇总记录 CRUD
# ──────────────────────────────────────────────
@app.get("/admin/api/interview-summaries")
async def admin_list_summaries(
    admin: dict = _Depends(get_current_admin),
    _=require_feature("admin.interview_analysis"),
):
    """列出所有汇总记录"""
    sb = db()
    rows = (
        sb.table("interview_summaries")
        .select("*")
        .eq("is_active", True)
        .order("created_at", desc=True)
        .execute().data or []
    )
    return {"ok": True, "summaries": [_row_to_camel(r) for r in rows]}


@app.get("/admin/api/interview-summaries/{summary_id}")
async def admin_get_summary(
    summary_id: int,
    admin: dict = _Depends(get_current_admin),
    _=require_feature("admin.interview_analysis"),
):
    """获取汇总详情"""
    sb = db()
    rows = sb.table("interview_summaries").select("*").eq("id", summary_id).execute().data
    if not rows:
        raise HTTPException(status_code=404, detail="汇总不存在")
    return {"ok": True, "summary": _row_to_camel(rows[0])}


@app.put("/admin/api/interview-summaries/{summary_id}")
async def admin_update_summary(
    summary_id: int,
    body: dict,
    admin: dict = _Depends(get_current_admin),
    _=require_feature("admin.interview_analysis"),
):
    """编辑汇总结果（管理员修改后的结果）"""
    sb = db()
    rows = sb.table("interview_summaries").select("id").eq("id", summary_id).execute().data
    if not rows:
        raise HTTPException(status_code=404, detail="汇总不存在")

    data = _body_to_snake(body)
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if "edit_result" in data:
        update_data["edit_result"] = data["edit_result"]
    if "title" in data:
        update_data["title"] = data["title"]

    sb.table("interview_summaries").update(update_data).eq("id", summary_id).execute()
    return {"ok": True, "summaryId": summary_id}


@app.delete("/admin/api/interview-summaries/{summary_id}")
async def admin_delete_summary(
    summary_id: int,
    admin: dict = _Depends(get_current_admin),
    _=require_feature("admin.interview_analysis"),
):
    """删除汇总记录"""
    sb = db()
    sb.table("interview_summaries").update({
        "is_active": False,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }).eq("id", summary_id).execute()
    return {"ok": True}


# ──────────────────────────────────────────────
# v2: 批量导入
# ──────────────────────────────────────────────
@app.post("/admin/api/interview-summaries/{summary_id}/confirm-import")
async def admin_confirm_import_summary(
    summary_id: int,
    body: dict | None = None,
    admin: dict = _Depends(get_current_admin),
    _=require_feature("admin.interview_analysis"),
):
    """将汇总结果批量导入到场景树/四维度/行动（草稿状态）"""
    sb = db()
    rows = sb.table("interview_summaries").select("*").eq("id", summary_id).execute().data
    if not rows:
        raise HTTPException(status_code=404, detail="汇总不存在")

    summary = rows[0]

    # 优先使用管理员编辑后的结果，否则使用 LLM 原始结果
    analysis_result = summary.get("edit_result") or summary.get("summary_result")
    if not analysis_result:
        raise HTTPException(status_code=400, detail="无可用分析结果")

    # 获取角色 ID（从关联访谈中取第一个）
    role_id = None
    interview_ids = summary.get("interview_ids", [])
    if interview_ids:
        iv_row = sb.table("interview_records").select("role_id").eq("id", interview_ids[0]).execute().data
        if iv_row and iv_row[0].get("role_id"):
            role_id = iv_row[0]["role_id"]

    try:
        result = write_analysis_results(
            sb=sb,
            summary_id=summary_id,
            analysis_result=analysis_result,
            admin_id=admin.get("id", 0),
            role_id=role_id,
        )

        # 同步更新关联访谈的 write_status
        if interview_ids:
            status = "completed" if not result.get("errors") else "partial"
            sb.table("interview_records").update({
                "write_status": status,
                "write_result": {"summaryId": summary_id},
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }).in_("id", interview_ids).execute()

        return {
            "ok": True,
            "summaryId": summary_id,
            "scenariosCreated": len(result.get("scenarios", [])),
            "dimensionsCreated": len(result.get("dimensions", [])),
            "actionsCreated": len(result.get("actions", [])),
            "errors": result.get("errors", []),
        }

    except Exception as e:
        logger.error(f"Summary import failed: {e}")
        sb.table("interview_summaries").update({
            "import_status": "failed",
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }).eq("id", summary_id).execute()
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


# 启动说明
# 注册评估模块路由
app.include_router(assessment.router)

if __name__ == "__main__":
    import uvicorn
    print("=" * 50)
    print("国窖1573智能销售兵工厂 — 后端 API v2.1")
    print("数据源: Supabase (降级: enriched-data.json)")
    print("=" * 50)
    print("API 文档: http://localhost:8000/docs")
    print("健康检查:   http://localhost:8000/health")
    print()
    uvicorn.run(app, host="0.0.0.0", port=8000, reload=True)
